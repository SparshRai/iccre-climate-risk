MODEL_VERSION    = "1.2.0"
PARAMETER_VERSION = "ClimateParams_v1.2"
NGFS_DATA_VERSION = "NGFS_PhaseIII_2023"
ENGINE_BUILD      = "IntegratedClimateCreditEngine"
MODEL_BUILD_DATE  = "2026-04-05"
MODEL_HASH        = "ICCRE_v1.2_prod"

# ============================================================
# PRODUCT & BRANDING
# ============================================================
PRODUCT_TAGLINE  = "India's first quantitative climate-to-credit engine"
PRODUCT_SUBTITLE = "NGFS-aligned · BRSR-native · ISSB S2-ready · INR-denominated"
CONTACT_EMAIL    = "hello@iccre.in"
PRODUCT_URL      = "https://iccre.in"
LINKEDIN_URL     = "https://linkedin.com/company/iccre"

# ── Demo dataset: Bharat Steel Industries Ltd (fictional) ──
# Pre-configured to produce visually compelling, realistic demo output.
# PD jumps from 1.5% → 7.3% under Net Zero 2050. DSCR falls below 1.0x.
# Carbon burden 11.4% of revenue. ECL ₹394 Cr vs ₹81 Cr baseline.
DEMO_DATASET = {
    "company_name":      "Bharat Steel Industries Ltd",
    "sector":            "Steel",
    "reporting_year":    2035,
    "revenue_0":         18000.0,
    "ebitda_margin_0":   0.22,
    "interest_payment":  1100.0,
    "total_assets":      45000.0,
    "exposure_at_default": 12000.0,
    "scope1":            4200000.0,
    "scope2":            1800000.0,
    "scope3":            2000000.0,
    "high_carbon_assets": 8000.0,
    "base_pd":           0.015,
    "lgd_0":             0.45,
    "planned_capex":     1800.0,
    "abatement_cost":    4500.0,
    "abatement_potential": 0.30,
    "carbon_pass_through": 0.35,
    "demand_elasticity": -0.40,
}

# ============================================================
# CHANGELOG v1.1 → v1.2
# ============================================================
# UI-01  Complete dark-teal design system (CSS variables,
#        gradient backgrounds, card components, KPI panels).
#        All tables, charts, and metrics use consistent theme.
#
# UI-02  Professional Plotly chart styling engine applied
#        globally — dark backgrounds, teal/amber/coral palette,
#        hover tooltips, reference lines, annotations.
#
# UI-03  Dashboard landing tab: 6 KPI hero cards, risk gauge,
#        PD sparkline, module status badges.
#
# UI-04  All st.dataframe() calls replaced with styled
#        tables using background-gradient, color-coded cells,
#        and proper column formatting.
#
# BRSR-01  BRSR tab rebuilt as financial risk engine:
#          GHG intensity benchmarked against SEBI sector P25/P50/P75.
#          BRSR flags mapped to PD basis-point uplift (additive, ≤150bps).
#          3-year emissions intensity trend scoring.
#          Water stranded cost projection (5-yr, stress escalation).
#          Energy transition risk (fossil carbon surcharge 2030).
#          SEBI BRSR Core regulatory readiness score (8-point).
#          5 professional Plotly charts (benchmark bars, PD tornado,
#          readiness radar, 5-yr cost stack, emissions trend line).
#          Live integration block linking BRSR PD uplift to transition PD.
#
# PHYS-01  Physical risk methodology corrected:
#          IPCC AR6 non-linear damage function: 1+0.20·ΔT+0.04·ΔT²
#          (replaces flat 1+0.25·ΔT multiplier with no basis).
#          ΔT extracted directly from NGFS scenario temperature pathways
#          (same df_long as transition engine — fully aligned).
#          P10/P50/P90 uncertainty bands: σ=0.08·ΔT.
#          Signed DSCR gap consistent with FIX-04.
#          Chronic vs acute physical risk split.
#          4-panel scenario dashboard per NGFS scenario.
#          Asset vulnerability heatmap (flood/heat/cyclone).
#
# INT-01   Integrated risk tab rebuilt with rigorous methodology:
#          Gaussian copula joint PD shown with full derivation steps.
#          Risk decomposition: transition % vs physical % of total ECL.
#          BRSR PD uplift added as third additive layer on top.
#          Scenario stress table: all 3 NGFS × 6 key metrics.
#          Capital adequacy section with ICAAP capital buffer calc.
#          Board-level RAG summary (Red/Amber/Green with thresholds).
#          All v1.1 fixes (FIX-01 to FIX-07) preserved in full.
#
# All v1.1 fixes (FIX-01 to FIX-07) are fully preserved.
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import os, json
import urllib.request
from groq import Groq
from math import radians, sin, cos, asin, sqrt

# Rasterio is required only for the Physical Risk flood layer.
# Keeping this import guarded prevents the full app from failing during deployment
# if native geospatial wheels are temporarily unavailable.
try:
    import rasterio
    from rasterio.transform import rowcol
    from rasterio.windows import Window
    RASTERIO_IMPORT_ERROR = None
except Exception as _rasterio_error:
    rasterio = None
    rowcol = None
    Window = None
    RASTERIO_IMPORT_ERROR = str(_rasterio_error)

import folium
from streamlit_folium import st_folium
from scipy.stats import norm, multivariate_normal
from datetime import datetime
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from plotly.subplots import make_subplots

RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)

# ============================================================
# PAGE CONFIG — must be the first Streamlit UI command
# ============================================================
st.set_page_config(
    page_title="ICCRE — India Climate Credit Risk Engine",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help": f"mailto:{CONTACT_EMAIL}",
        "Report a bug": f"mailto:{CONTACT_EMAIL}?subject=Bug Report",
        "About": (
            f"**ICCRE v{MODEL_VERSION}** — {PRODUCT_TAGLINE}\n\n"
            f"{PRODUCT_SUBTITLE}\n\n"
            f"Built on NGFS Phase III 2023 · ISSB IFRS S2 · RBI 2024 · SEBI BRSR Core\n\n"
            f"Contact: {CONTACT_EMAIL}"
        ),
    },
)


def _secret_or_env(name: str, default=None):
    """Return Streamlit secret first, then environment variable, then default."""
    try:
        value = st.secrets.get(name)
        if value:
            return value
    except Exception:
        pass
    return os.getenv(name, default)


def _get_groq_client():
    """Create Groq client from Streamlit Secrets or environment variable."""
    api_key = _secret_or_env("GROQ_API_KEY")
    if not api_key:
        raise RuntimeError("GROQ_API_KEY is missing. Add it in Streamlit Secrets or environment variables.")
    return Groq(api_key=api_key)
# ============================================================
# GIS FILE LOADER — downloads large files from cloud storage
# on first run so they don't need to be in the GitHub repo
# ============================================================
GCS_BASE = _secret_or_env("GCS_BASE", "https://storage.googleapis.com/iccre-gis-data")  # ← override in Streamlit Secrets if needed

_GIS_FILES = {
    "Data/floodMapGL_rp100y.tif":        f"{GCS_BASE}/floodMapGL_rp100y.tif",
    "Data/ibtracs.NI.list.v04r01.csv":   f"{GCS_BASE}/ibtracs.NI.list.v04r01.csv",
    "Data/era5_test_day_grid.csv":        f"{GCS_BASE}/era5_test_day_grid.csv",
}

@st.cache_resource(show_spinner="⏳ Loading physical risk data (one-time, ~30s)...")
def _ensure_gis_files():
    os.makedirs("Data", exist_ok=True)
    failed = []
    for local_path, url in _GIS_FILES.items():
        if not os.path.exists(local_path):
            tmp_path = f"{local_path}.download"
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                urllib.request.urlretrieve(url, tmp_path)
                os.replace(tmp_path, local_path)
            except Exception as e:
                if os.path.exists(tmp_path):
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                failed.append(f"{local_path}: {e}")
    return failed

if RASTERIO_IMPORT_ERROR:
    st.sidebar.warning("⚠️ Rasterio failed to import. Flood layer will be disabled, but other modules can still run.")
# ============================================================
# DESIGN SYSTEM  (UI-01)
# ============================================================
C = dict(
    bg_dark   = "#062F2E",
    bg_mid    = "#0B4D4B",
    bg_ocean  = "#115E6D",
    sidebar   = "#0A1F33",
    accent    = "#00F5D4",
    accent2   = "#22D3EE",
    accent3   = "#06B6D4",
    coral     = "#FF6B6B",
    amber     = "#FFD166",
    mint      = "#80ED99",
    purple    = "#A78BFA",
    white     = "#FFFFFF",
    off_white = "#E6F1F5",
    slate     = "#94A3B8",
    text      = "#E2E8F0",
    card      = "#0D3B4A",
)

def _chart_layout(title="", height=340, legend_override=None, margin_override=None, **kwargs):
    """
    Shared Plotly layout for all charts (UI-02).

    legend_override: dict merged into the default legend dict.
    margin_override: dict merged into the default margin dict.

    Use these instead of passing legend=/margin= directly to update_layout(),
    which would cause 'multiple values for keyword argument' TypeError.

    Example:
        fig.update_layout(**_chart_layout("Title",
            legend_override=dict(orientation="h", y=1.1),
            margin_override=dict(l=200, r=80)))
    """
    base_legend = dict(bgcolor="rgba(0,0,0,0)", font=dict(color=C["off_white"]))
    if legend_override:
        base_legend.update(legend_override)

    base_margin = dict(l=50, r=20, t=50, b=40)
    if margin_override:
        base_margin.update(margin_override)

    return dict(
        title=dict(text=title, font=dict(color=C["off_white"], size=14)),
        height=height,
        plot_bgcolor=C["bg_dark"],
        paper_bgcolor=C["bg_dark"],
        font=dict(color=C["off_white"], size=11),
        legend=base_legend,
        margin=base_margin,
        **kwargs,
    )

def _ax_style(fig, rows=1, cols=1):
    """Apply grid styling to all axes.
    Plain go.Figure (no make_subplots) must NOT receive row/col args —
    Plotly raises an Exception because there is no subplot grid.
    Only pass row/col when the figure was created with make_subplots().
    """
    if rows == 1 and cols == 1:
        # Plain figure — no row/col args
        fig.update_xaxes(gridcolor="#0B4D4B", zeroline=False)
        fig.update_yaxes(gridcolor="#0B4D4B", zeroline=False)
    else:
        # make_subplots figure — iterate all cells
        for i in range(1, rows + 1):
            for j in range(1, cols + 1):
                fig.update_xaxes(gridcolor="#0B4D4B", zeroline=False, row=i, col=j)
                fig.update_yaxes(gridcolor="#0B4D4B", zeroline=False, row=i, col=j)
    return fig

# Scenario colour mapping — fixed per scenario name for consistency across ALL charts
SCENARIO_COLORS = {
    "Current Policies":                               C["coral"],   # red-coral
    "Nationally Determined Contributions (NDCs)":     C["amber"],   # amber-yellow
    "Net Zero 2050":                                  C["accent2"], # teal-cyan
}

# Fallback palette for any extra / unrecognised scenarios — always distinguishable
_FALLBACK_PALETTE = [
    C["mint"],    # green
    C["purple"],  # purple
    "#F97316",    # orange
    "#EC4899",    # pink
    "#84CC16",    # lime
    "#14B8A6",    # teal-green
]
_fallback_idx: dict = {}   # scenario_name → palette index, populated on first use

def _scen_color(scen: str) -> str:
    """
    Return a stable, distinguishable colour for a scenario name.
    Named scenarios always get their fixed colour.
    Unknown scenarios cycle through _FALLBACK_PALETTE with a stable index
    (same scenario always gets same colour within one session).
    """
    global _fallback_idx
    # Direct match
    if scen in SCENARIO_COLORS:
        return SCENARIO_COLORS[scen]
    # Partial-match against known names (handles truncated scenario strings)
    for k, v in SCENARIO_COLORS.items():
        if k in scen or scen in k:
            return v
    # Unknown scenario — assign a stable fallback colour
    if scen not in _fallback_idx:
        _fallback_idx[scen] = len(_fallback_idx) % len(_FALLBACK_PALETTE)
    return _FALLBACK_PALETTE[_fallback_idx[scen]]

def _hex_rgba(hex_color: str, alpha: float) -> str:
    """
    Convert a 6-char hex color string to an rgba() string accepted by Plotly.
    Plotly does NOT accept 8-char hex (#RRGGBBAA) — must use rgba() notation.

    Example: _hex_rgba("#FF6B6B", 0.15) → "rgba(255,107,107,0.15)"
    """
    h = hex_color.lstrip("#")
    if len(h) != 6:
        return f"rgba(6,182,212,{alpha})"   # safe fallback
    r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    return f"rgba({r},{g},{b},{alpha})"

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)

def log_model_run(run_type, payload):
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    entry = {"timestamp": ts, "model_version": MODEL_VERSION,
             "parameter_version": PARAMETER_VERSION, "ngfs_version": NGFS_DATA_VERSION,
             "engine": ENGINE_BUILD, "run_type": run_type, **payload}
    f = LOG_DIR / "run_log.csv"
    df = pd.DataFrame([entry])
    df.to_csv(f, mode="a", header=not f.exists(), index=False)

# ============================================================
# CSS  (UI-01)
# ============================================================
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {{
    font-family: 'IBM Plex Sans', sans-serif !important;
}}

.stApp {{
    background: linear-gradient(145deg, {C["bg_dark"]} 0%, {C["bg_mid"]} 45%, {C["bg_ocean"]} 100%);
    background-attachment: fixed;
}}

/* ── SIDEBAR BASE ────────────────────────────────────────── */
section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {C["sidebar"]} 0%, #020617 100%) !important;
    border-right: 1px solid {C["bg_mid"]};
}}

/* All sidebar text */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] small,
section[data-testid="stSidebar"] .stMarkdown {{
    color: {C["text"]} !important;
}}

/* Sidebar widget labels */
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stNumberInput label,
section[data-testid="stSidebar"] .stTextInput label,
section[data-testid="stSidebar"] .stCheckbox label,
section[data-testid="stSidebar"] .stRadio label {{
    color: {C["slate"]} !important;
    font-size: 12px !important;
}}

/* Sidebar input FIELDS — dark background, light text */
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea,
section[data-testid="stSidebar"] .stTextInput input,
section[data-testid="stSidebar"] .stNumberInput input {{
    background-color: {C["card"]} !important;
    color: {C["accent2"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    border-radius: 6px !important;
    caret-color: {C["accent2"]} !important;
}}
section[data-testid="stSidebar"] input:focus,
section[data-testid="stSidebar"] textarea:focus {{
    border-color: {C["accent2"]} !important;
    box-shadow: 0 0 0 2px rgba(34,211,238,0.25) !important;
    outline: none !important;
}}
section[data-testid="stSidebar"] input::placeholder {{
    color: {C["slate"]} !important;
    opacity: 0.7;
}}

/* Sidebar SELECT boxes */
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] > div,
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="select"] > div {{
    background-color: {C["card"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    border-radius: 6px !important;
}}
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] span,
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="select"] span,
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] div,
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="select"] div {{
    color: {C["accent2"]} !important;
    background-color: transparent !important;
}}

/* Sidebar SELECT dropdown menu */
[data-baseweb="popover"] ul,
[data-baseweb="menu"] {{
    background-color: {C["card"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    border-radius: 8px !important;
}}
[data-baseweb="menu"] li,
[data-baseweb="option"] {{
    color: {C["text"]} !important;
    background-color: {C["card"]} !important;
}}
[data-baseweb="option"]:hover,
[data-baseweb="menu"] li:hover {{
    background-color: {C["bg_mid"]} !important;
    color: {C["accent2"]} !important;
}}

/* Sidebar number input spinner arrows */
section[data-testid="stSidebar"] .stNumberInput button {{
    background-color: {C["bg_mid"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    color: {C["accent2"]} !important;
}}
section[data-testid="stSidebar"] .stNumberInput button:hover {{
    background-color: {C["accent3"]} !important;
    color: {C["bg_dark"]} !important;
}}

/* Sidebar SLIDER */
section[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] div[role="slider"] {{
    background-color: {C["accent2"]} !important;
}}
section[data-testid="stSidebar"] .stSlider [data-testid="stTickBar"] {{
    color: {C["slate"]} !important;
}}

/* Sidebar CHECKBOX */
section[data-testid="stSidebar"] .stCheckbox span,
section[data-testid="stSidebar"] .stCheckbox label {{
    color: {C["text"]} !important;
}}

/* Sidebar divider */
section[data-testid="stSidebar"] hr {{
    border-color: {C["bg_mid"]} !important;
}}

/* Sidebar subheader / header */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {{
    color: {C["accent2"]} !important;
    font-size: 13px !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    font-weight: 600 !important;
    margin-top: 12px !important;
}}

/* ── MAIN AREA TEXT ─────────────────────────────────────── */
h1,h2,h3,h4,h5,h6 {{ color: {C["white"]} !important; font-weight: 600 !important; }}
p, span, label, div {{ color: {C["text"]} !important; }}
.stMarkdown p {{ color: {C["text"]} !important; }}

/* Main area inputs also dark */
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {{
    background-color: {C["card"]} !important;
    color: {C["accent2"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    border-radius: 6px !important;
}}
.stSelectbox [data-baseweb="select"] > div,
.stMultiSelect [data-baseweb="select"] > div {{
    background-color: {C["card"]} !important;
    border: 1px solid {C["bg_ocean"]} !important;
    color: {C["accent2"]} !important;
}}

/* ── METRIC CARDS ───────────────────────────────────────── */
[data-testid="stMetric"] {{
    background: linear-gradient(135deg, {C["card"]} 0%, {C["bg_mid"]} 100%);
    border: 1px solid {C["bg_ocean"]};
    border-radius: 10px;
    padding: 14px 16px !important;
}}
[data-testid="stMetricLabel"] {{ color: {C["slate"]} !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: 0.08em; }}
[data-testid="stMetricValue"] {{ color: {C["accent2"]} !important; font-size: 22px !important; font-weight: 600 !important; font-family: 'IBM Plex Mono', monospace !important; }}
[data-testid="stMetricDelta"] {{ font-size: 11px !important; }}

/* ── BUTTONS ────────────────────────────────────────────── */
.stButton > button {{
    background: linear-gradient(135deg, {C["accent3"]} 0%, {C["accent2"]} 100%) !important;
    color: {C["bg_dark"]} !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    letter-spacing: 0.02em;
    padding: 8px 20px !important;
    transition: opacity 0.2s;
}}
.stButton > button:hover {{ opacity: 0.88; }}

/* ── TABS ───────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {{
    background: {C["bg_dark"]} !important;
    border-bottom: 1px solid {C["bg_mid"]};
    gap: 2px;
}}
.stTabs [data-baseweb="tab"] {{
    color: {C["slate"]} !important;
    background: transparent !important;
    border-radius: 8px 8px 0 0 !important;
    font-size: 13px !important;
    padding: 8px 16px !important;
}}
.stTabs [aria-selected="true"] {{
    color: {C["accent2"]} !important;
    background: {C["card"]} !important;
    border-bottom: 2px solid {C["accent2"]} !important;
}}

/* ── DATAFRAMES ─────────────────────────────────────────── */
.stDataFrame {{ border-radius: 10px !important; overflow: hidden; }}
.stDataFrame [data-testid="stDataFrameResizable"] {{
    background: {C["bg_dark"]} !important;
    border: 1px solid {C["bg_mid"]} !important;
}}

/* ── ALERTS ─────────────────────────────────────────────── */
.stInfo    {{ background: rgba(6,182,212,0.1)   !important; border-left: 3px solid {C["accent3"]} !important; border-radius: 6px; }}
.stWarning {{ background: rgba(255,209,102,0.1) !important; border-left: 3px solid {C["amber"]}   !important; border-radius: 6px; }}
.stSuccess {{ background: rgba(128,237,153,0.1) !important; border-left: 3px solid {C["mint"]}    !important; border-radius: 6px; }}
.stError   {{ background: rgba(255,107,107,0.1) !important; border-left: 3px solid {C["coral"]}   !important; border-radius: 6px; }}

/* ── EXPANDER ───────────────────────────────────────────── */
.streamlit-expanderHeader {{ background: {C["card"]} !important; border-radius: 8px !important; color: {C["accent2"]} !important; }}

/* ── KPI CARD ───────────────────────────────────────────── */
.kpi-card {{
    background: linear-gradient(135deg, {C["card"]} 0%, {C["bg_mid"]} 100%);
    border: 1px solid {C["bg_ocean"]};
    border-radius: 10px;
    padding: 16px;
    text-align: center;
}}
.kpi-value {{ font-size: 24px; font-weight: 600; color: {C["accent2"]}; font-family: 'IBM Plex Mono', monospace; }}
.kpi-label {{ font-size: 11px; color: {C["slate"]}; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 4px; }}

/* ── RAG BADGES ─────────────────────────────────────────── */
.rag-green {{ color: {C["mint"]}  !important; font-weight: 600; }}
.rag-amber {{ color: {C["amber"]} !important; font-weight: 600; }}
.rag-red   {{ color: {C["coral"]} !important; font-weight: 600; }}

/* ── MISC ───────────────────────────────────────────────── */
hr {{ border-color: {C["bg_mid"]} !important; }}
code {{ background: {C["bg_dark"]} !important; color: {C["amber"]} !important; border-radius: 4px; font-family: 'IBM Plex Mono', monospace !important; }}
</style>
""", unsafe_allow_html=True)

# ============================================================
# HEADER BANNER
# ============================================================
st.markdown(f"""
<div style="background:linear-gradient(90deg,{C['bg_dark']} 0%,{C['bg_ocean']} 55%,{C['bg_mid']} 100%);
            border-bottom:2px solid {C['accent']};padding:18px 24px 14px;
            margin-bottom:4px;border-radius:0 0 12px 12px;">
  <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;">
    <div>
      <div style="display:flex;align-items:center;gap:12px;">
        <div style="font-size:28px;font-weight:800;color:{C['white']};letter-spacing:-0.03em;
                    font-family:'IBM Plex Sans',sans-serif;">
          ICCRE
        </div>
        <div style="width:1px;height:32px;background:{C['bg_ocean']};"></div>
        <div>
          <div style="font-size:14px;font-weight:600;color:{C['accent2']};letter-spacing:0.01em;">
            {PRODUCT_TAGLINE}
          </div>
          <div style="font-size:11px;color:{C['slate']};margin-top:2px;">
            {PRODUCT_SUBTITLE}
          </div>
        </div>
      </div>
    </div>
    <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
      <span style="background:{C['card']};border:1px solid {C['bg_ocean']};color:{C['accent2']};
                   padding:4px 12px;border-radius:20px;font-size:11px;font-family:monospace;
                   font-weight:600;">
        v{MODEL_VERSION}
      </span>
      <span style="background:{C['card']};border:1px solid {C['bg_ocean']};color:{C['slate']};
                   padding:4px 12px;border-radius:20px;font-size:11px;">
        NGFS Phase III 2023
      </span>
      <a href="mailto:{CONTACT_EMAIL}" style="background:{C['accent']};color:{C['bg_dark']};
         padding:5px 14px;border-radius:20px;font-size:11px;font-weight:700;text-decoration:none;
         letter-spacing:0.02em;">
        ✉ Get Access
      </a>
    </div>
  </div>
  <div style="display:flex;gap:16px;margin-top:10px;flex-wrap:wrap;">
    <span style="font-size:10px;color:{C['slate']};background:{C['bg_dark']};
                 padding:3px 10px;border-radius:12px;border:1px solid {C['bg_mid']};">
      ✓ ISSB IFRS S2
    </span>
    <span style="font-size:10px;color:{C['slate']};background:{C['bg_dark']};
                 padding:3px 10px;border-radius:12px;border:1px solid {C['bg_mid']};">
      ✓ RBI Climate Guidelines 2024
    </span>
    <span style="font-size:10px;color:{C['slate']};background:{C['bg_dark']};
                 padding:3px 10px;border-radius:12px;border:1px solid {C['bg_mid']};">
      ✓ SEBI BRSR Core
    </span>
    <span style="font-size:10px;color:{C['slate']};background:{C['bg_dark']};
                 padding:3px 10px;border-radius:12px;border:1px solid {C['bg_mid']};">
      ✓ BCBS 2021
    </span>
    <span style="font-size:10px;color:{C['slate']};background:{C['bg_dark']};
                 padding:3px 10px;border-radius:12px;border:1px solid {C['bg_mid']};">
      ✓ ICAAP-ready
    </span>
  </div>
</div>
""", unsafe_allow_html=True)


# ============================================================
# DEMO ACCESS GATE — lead capture before access
# ============================================================
DEMO_ACCESS_FILE = Path("Data/demo_leads.csv")
DEMO_LOCKED_COMPANY = True


def _valid_email(value: str) -> bool:
    value = (value or "").strip()
    return "@" in value and "." in value.split("@")[-1]


def _save_demo_lead(lead: dict):
    """Save user lead locally and optionally forward to a webhook."""
    os.makedirs("Data", exist_ok=True)
    row = {**lead, "captured_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")}
    df = pd.DataFrame([row])
    df.to_csv(DEMO_ACCESS_FILE, mode="a", header=not DEMO_ACCESS_FILE.exists(), index=False)

    webhook = _secret_or_env("LEADS_WEBHOOK_URL")
    if webhook:
        try:
            req = urllib.request.Request(
                webhook,
                data=json.dumps(row).encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            urllib.request.urlopen(req, timeout=4)
        except Exception:
            pass


def demo_access_gate():
    """Collect basic user details before opening the demo."""
    if st.session_state.get("demo_access_granted"):
        return

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{C['card']},{C['bg_mid']});border:1px solid {C['bg_ocean']};border-radius:16px;padding:28px 34px;margin-top:18px;">
      <div style="font-size:24px;font-weight:800;color:{C['white']};margin-bottom:6px;">Access ICCRE Demo</div>
      <div style="font-size:13px;color:{C['slate']};line-height:1.7;max-width:820px;">
        This public demo uses a fictional company — <b>Bharat Steel Industries Ltd</b>. Inputs are locked so every user evaluates the same demo case, while all analysis tabs and output views remain available.
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.form("demo_lead_form", clear_on_submit=False):
        c1, c2 = st.columns(2)
        name = c1.text_input("Name", placeholder="Your name")
        email = c2.text_input("Work Email *", placeholder="name@company.com")
        org = c1.text_input("Organisation *", placeholder="Company / institution")
        role = c2.selectbox("Role", ["Bank / Risk", "ESG / Sustainability", "Corporate Finance", "Consulting", "Investor", "Student / Researcher", "Other"])
        purpose = st.selectbox("Purpose of demo *", ["Evaluate for pilot", "Explore product", "Research / learning", "Partnership discussion", "Investment interest", "Other"])
        notes = st.text_area("Anything specific you want to evaluate?", placeholder="Optional", height=80)
        consent = st.checkbox("I agree that my details may be recorded for demo follow-up.")
        submitted = st.form_submit_button("Enter Demo", type="primary", width="stretch")

    if submitted:
        if not _valid_email(email):
            st.error("Please enter a valid email address.")
            st.stop()
        if not org.strip():
            st.error("Please enter your organisation.")
            st.stop()
        if not consent:
            st.error("Please confirm consent to continue.")
            st.stop()
        lead = {
            "name": name.strip(),
            "email": email.strip(),
            "organisation": org.strip(),
            "role": role,
            "purpose": purpose,
            "notes": notes.strip(),
            "demo_company": DEMO_DATASET["company_name"],
            "model_version": MODEL_VERSION,
        }
        _save_demo_lead(lead)
        st.session_state["demo_access_granted"] = True
        st.session_state["demo_lead"] = lead
        st.rerun()

    st.info("Your inputs will be locked to the fictional demo company. You can run and explore all output modules after access.")
    st.stop()


demo_access_gate()

# ============================================================
# SESSION STATE DEFAULTS
# ============================================================
_DEFAULTS = {
    "ai_outputs": {}, "suggested_params": {},
    "integrated_ran": False, "df_integrated_summary": None,
    "enable_transition": True, "enable_physical": False,
    "enable_targets": False, "enable_brsr": False,
    "transition_ran": False, "physical_ran": False,
    "targets_ran": False, "brsr_ran": False,
    "results_ready": False,
    "scope1": 0.0, "scope2": 0.0, "scope3": 0.0,
    "df_transition": None, "df_target": None,
    "phys_summary": None, "phys_assets": None, "df_physical_projection": None,
    "brsr_summary": None, "brsr_flags": None, "brsr_pd_adj": 0.0,
    "mc_results": None, "multi_year_results": None,
    "calibrated_params": None, "historical_data": None,
}
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ============================================================
# SECTOR CONSTANTS
# ============================================================
SECTOR_CORRELATION = {"Steel": 0.35, "Power": 0.30, "Cement": 0.32, "Oil & Gas": 0.40, "Manufacturing": 0.25}
SECTOR_GDP_SENSITIVITY = {"Steel": 1.20, "Power": 0.70, "Cement": 1.10, "Oil & Gas": 0.90, "Manufacturing": 1.00}
SECTOR_CARBON_DEMAND = {"Steel": -0.25, "Power": -0.10, "Cement": -0.20, "Oil & Gas": -0.30, "Manufacturing": -0.15}
SECTOR_CREDIT_PARAMS = {
    "Steel":         {"alpha_dscr": 1.10, "beta_carbon_credit": 1.40, "gamma_physical": 0.60, "delta_gdp": 0.80},
    "Power":         {"alpha_dscr": 1.00, "beta_carbon_credit": 1.60, "gamma_physical": 0.50, "delta_gdp": 0.60},
    "Cement":        {"alpha_dscr": 1.05, "beta_carbon_credit": 1.20, "gamma_physical": 0.55, "delta_gdp": 0.70},
    "Oil & Gas":     {"alpha_dscr": 0.90, "beta_carbon_credit": 1.50, "gamma_physical": 0.65, "delta_gdp": 0.50},
    "Manufacturing": {"alpha_dscr": 1.00, "beta_carbon_credit": 0.80, "gamma_physical": 0.40, "delta_gdp": 0.90},
}
SECTOR_MC_PARAMS = {
    "Steel":         {"carbon_vol": 0.25, "physical_vol": 0.20, "gdp_vol": 0.05},
    "Power":         {"carbon_vol": 0.22, "physical_vol": 0.18, "gdp_vol": 0.04},
    "Cement":        {"carbon_vol": 0.23, "physical_vol": 0.19, "gdp_vol": 0.05},
    "Oil & Gas":     {"carbon_vol": 0.30, "physical_vol": 0.22, "gdp_vol": 0.06},
    "Manufacturing": {"carbon_vol": 0.20, "physical_vol": 0.15, "gdp_vol": 0.04},
}
SECTOR_SPREAD_SENSITIVITY = {"Steel": 1.4, "Power": 1.2, "Cement": 1.3, "Oil & Gas": 1.5, "Manufacturing": 1.0}
SECTOR_PORTFOLIO_CORR = {
    "Steel":         {"Steel": 0.35, "Power": 0.25, "Cement": 0.30, "Oil & Gas": 0.28, "Manufacturing": 0.22},
    "Power":         {"Steel": 0.25, "Power": 0.40, "Cement": 0.27, "Oil & Gas": 0.35, "Manufacturing": 0.20},
    "Cement":        {"Steel": 0.30, "Power": 0.27, "Cement": 0.38, "Oil & Gas": 0.26, "Manufacturing": 0.24},
    "Oil & Gas":     {"Steel": 0.28, "Power": 0.35, "Cement": 0.26, "Oil & Gas": 0.45, "Manufacturing": 0.22},
    "Manufacturing": {"Steel": 0.22, "Power": 0.20, "Cement": 0.24, "Oil & Gas": 0.22, "Manufacturing": 0.30},
}
SECTOR_CONTAGION = {
    "Power":     {"Steel": 0.20, "Cement": 0.15},
    "Steel":     {"Manufacturing": 0.18, "Cement": 0.10},
    "Oil & Gas": {"Power": 0.22, "Manufacturing": 0.15},
}
SCENARIO_WEIGHTS = {"Current Policies": 0.5, "Nationally Determined Contributions (NDCs)": 0.3, "Net Zero 2050": 0.2}
SCENARIO_REGISTRY = {
    "Current Policies": {"type": "Baseline", "temperature": "≈2.7°C", "policy_intensity": "Low"},
    "Nationally Determined Contributions (NDCs)": {"type": "Policy Transition", "temperature": "≈2.1°C", "policy_intensity": "Medium"},
    "Net Zero 2050": {"type": "Disorderly Transition", "temperature": "≈1.5°C", "policy_intensity": "High"},
}
SECTOR_GHG_BENCHMARKS = {
    "Steel":         {"p25": 1.8, "p50": 2.8, "p75": 4.2},
    "Power":         {"p25": 3.2, "p50": 5.1, "p75": 7.8},
    "Cement":        {"p25": 2.1, "p50": 3.4, "p75": 5.0},
    "Oil & Gas":     {"p25": 1.4, "p50": 2.2, "p75": 3.5},
    "Manufacturing": {"p25": 0.8, "p50": 1.4, "p75": 2.6},
}
SECTOR_ENERGY_BENCHMARKS = {
    "Steel": {"p50": 125, "p75": 180}, "Power": {"p50": 85, "p75": 130},
    "Cement": {"p50": 110, "p75": 160}, "Oil & Gas": {"p50": 70, "p75": 110},
    "Manufacturing": {"p50": 60, "p75": 95},
}
BRSR_PD_SPREAD = {
    "High GHG intensity (>P75)": 0.0035, "Low renewable energy (<20%)": 0.0018,
    "High energy intensity (>P75)": 0.0022, "High water stress + high withdrawal": 0.0028,
    "Low target coverage (<50%)": 0.0030, "High hazardous waste (>30%)": 0.0012,
    "No verified emissions data": 0.0025, "Missing Scope 3 disclosure": 0.0010,
}

PD_FLOOR = 0.0005
PD_CAP   = 0.35
LGD_PHYSICAL_MULTIPLIER = 0.30
BASE_CREDIT_SPREAD = 0.02
P_START = 50; P_FULL = 250; MAX_STRANDING = 0.80
margin_erosion_rate = 0.005; margin_floor = 0.10
SYSTEMIC_CLIMATE_WEIGHT = 0.50; SECTOR_SHOCK_WEIGHT = 0.25
CONTAGION_WEIGHT = 0.15; IDIOSYNCRATIC_WEIGHT = 0.10
CLIMATE_DRIVER_CORR = np.array([[1.0, 0.30, -0.40],[0.30, 1.0, -0.20],[-0.40, -0.20, 1.0]])

# ============================================================
# MATH HELPERS
# ============================================================
def logit(p):
    p = np.clip(p, 1e-8, 1 - 1e-8)
    v = np.log(p / (1 - p))
    return v if np.isfinite(v) else 0.0

def sigmoid(x):
    x = np.clip(x, -50, 50)
    v = 1 / (1 + np.exp(-x))
    return v if np.isfinite(v) else 0.0

def gaussian_copula_pd(pd_t, pd_p, rho):
    pd_t = np.clip(pd_t, 1e-6, 1-1e-6)
    pd_p = np.clip(pd_p, 1e-6, 1-1e-6)
    z_t  = norm.ppf(pd_t); z_p = norm.ppf(pd_p)
    return float(np.clip(multivariate_normal.cdf([z_t, z_p], [0,0], [[1,rho],[rho,1]]), 0, 1))

def simulate_carbon_price_path(start, years, drift=0.05, vol=0.25):
    prices = [start]
    for _ in range(years):
        prices.append(prices[-1] * np.exp((drift - 0.5*vol**2) + vol*np.random.normal()))
    return prices

def haversine(lat1, lon1, lat2, lon2):
    ln1,lt1,ln2,lt2 = map(radians,[lon1,lat1,lon2,lat2])
    a = sin((lt2-lt1)/2)**2 + cos(lt1)*cos(lt2)*sin((ln2-ln1)/2)**2
    return 6371*2*asin(sqrt(a))

# Damage curves
def flood_damage(d):
    if d<=0: return 0.0
    if d<0.3: return 0.02
    if d<0.7: return 0.10
    if d<1.5: return 0.35
    if d<2.5: return 0.60
    return 0.85

def heat_damage(d):
    if d<5: return 0.01
    if d<15: return 0.08
    if d<30: return 0.20
    if d<50: return 0.40
    return 0.65

def cyclone_damage(w):
    if w<80: return 0.02
    if w<120: return 0.15
    if w<160: return 0.35
    if w<200: return 0.60
    return 0.85

# ============================================================
# NGFS DATA
# ============================================================
@st.cache_data
def load_ngfs():
    candidates = [
        Path("Data/ngfs_scenarios.csv"),
        Path("data/ngfs_scenarios.csv"),
        Path("ngfs_scenarios.csv"),
    ]
    for path in candidates:
        if path.exists():
            df = pd.read_csv(path)
            df.columns = df.columns.str.strip().str.replace("﻿", "", regex=False)
            return df
    st.error("❌ File not found: Data/ngfs_scenarios.csv. Upload it to GitHub exactly at Data/ngfs_scenarios.csv.")
    st.stop()

df_ngfs = load_ngfs()
for c in df_ngfs.select_dtypes("object"):
    df_ngfs[c] = df_ngfs[c].astype(str).str.strip()
keywords = ["Price|Carbon","GDP|MER","AR6 climate diagnostics|Surface Temperature"]
df_ngfs  = df_ngfs[df_ngfs["Variable"].apply(lambda x: any(k.lower() in x.lower() for k in keywords))]
year_cols = [c for c in df_ngfs.columns if c.isdigit()]
df_long   = df_ngfs.melt(
    id_vars=[c for c in df_ngfs.columns if c not in year_cols],
    value_vars=year_cols, var_name="Year", value_name="Value"
)
df_long["Year"] = df_long["Year"].astype(int)

# ============================================================
# AI PARAMETER CALIBRATION
# ============================================================
def ai_calibrate_parameters(sector, *_):
    presets = {
        "Steel":         {"carbon_pass_through":0.35,"demand_elasticity":-0.40,"price_elasticity":-0.10,"beta_carbon_transition":1.35,"abatement_potential":0.30},
        "Power":         {"carbon_pass_through":0.65,"demand_elasticity":-0.15,"price_elasticity":-0.02,"beta_carbon_transition":1.50,"abatement_potential":0.45},
        "Cement":        {"carbon_pass_through":0.40,"demand_elasticity":-0.25,"price_elasticity":-0.05,"beta_carbon_transition":1.25,"abatement_potential":0.35},
        "Oil & Gas":     {"carbon_pass_through":0.55,"demand_elasticity":-0.35,"price_elasticity":-0.08,"beta_carbon_transition":1.45,"abatement_potential":0.25},
        "Manufacturing": {"carbon_pass_through":0.50,"demand_elasticity":-0.20,"price_elasticity":-0.04,"beta_carbon_transition":0.90,"abatement_potential":0.40},
    }
    return presets.get(sector, presets["Manufacturing"])

# ============================================================
# SIDEBAR
# ============================================================

# ============================================================
# LOCKED DEMO SIDEBAR — fixed fictional company, all modules available
# ============================================================
st.sidebar.markdown(f"""
<div style="padding:14px 0 10px;">
  <div style="font-size:20px;font-weight:800;color:{C['accent2']};letter-spacing:-0.02em;">ICCRE</div>
  <div style="font-size:10px;color:{C['slate']};margin-top:1px;line-height:1.5;">
    Locked public demo<br>
    <span style="color:{C['bg_ocean']}">v{MODEL_VERSION} · {MODEL_BUILD_DATE}</span>
  </div>
</div>
""", unsafe_allow_html=True)

st.sidebar.markdown(f"""
<div style='background:{C['amber']}22;border:1px solid {C['amber']};border-radius:8px;padding:10px 12px;margin:8px 0;font-size:11px;line-height:1.55;'>
  <b style='color:{C['amber']};'>🔒 Demo company locked</b><br>
  Bharat Steel Industries Ltd · Steel · Reporting Year 2035<br>
  Inputs cannot be changed in this public demo. You can run and explore all result tabs.
</div>
""", unsafe_allow_html=True)

lead = st.session_state.get("demo_lead", {})
if lead:
    st.sidebar.caption(f"Access: {lead.get('email','')}")

with st.sidebar.expander("📋 How to use this demo", expanded=True):
    st.markdown(f"""
<div style="font-size:11px;color:{C['off_white']};line-height:1.8;">
1. Run <b>Transition Risk</b><br>
2. Run <b>Physical Risk</b><br>
3. Run <b>BRSR Diagnostics</b><br>
4. Try <b>Targets</b><br>
5. Open <b>Integrated Risk</b> and <b>Plots</b><br>
6. Use <b>Get Access</b> for custom company analysis
</div>
""", unsafe_allow_html=True)

# Fixed demo inputs — no public user input fields.
for k, v in DEMO_DATASET.items():
    st.session_state[f"_demo_{k}"] = v
st.session_state["_demo_loaded"] = True

company_name = DEMO_DATASET["company_name"]
sector = DEMO_DATASET["sector"]
REPORTING_YEAR = int(DEMO_DATASET["reporting_year"])
BASELINE_SCENARIO = "Current Policies"
default_correlation = SECTOR_CORRELATION.get(sector, 0.30)

revenue_0           = float(DEMO_DATASET["revenue_0"])
ebitda_margin_0     = float(DEMO_DATASET["ebitda_margin_0"])
interest_payment    = float(DEMO_DATASET["interest_payment"])
total_assets        = float(DEMO_DATASET["total_assets"])
exposure_at_default = float(DEMO_DATASET["exposure_at_default"])

scope1 = float(DEMO_DATASET["scope1"])
scope2 = float(DEMO_DATASET["scope2"])
scope3 = float(DEMO_DATASET["scope3"])
TOTAL_EMISSIONS = scope1 + scope2 + scope3
st.session_state["scope1"] = scope1
st.session_state["scope2"] = scope2
st.session_state["scope3"] = scope3
high_carbon_assets = float(DEMO_DATASET["high_carbon_assets"])

# All major modules are enabled for the locked demo.
st.session_state["enable_transition"] = True
st.session_state["enable_physical"]   = True
st.session_state["enable_targets"]    = True
st.session_state["enable_brsr"]       = True

USD_INR = 83.0
US_CPI_2010_TO_2026 = 1.38
CARBON_PRICE_INFLATION_FACTOR = US_CPI_2010_TO_2026
carbon_pass_through = float(DEMO_DATASET["carbon_pass_through"])
demand_elasticity = float(DEMO_DATASET["demand_elasticity"])
price_elasticity = -0.10
beta_carbon_transition = 1.35
planned_capex = float(DEMO_DATASET["planned_capex"])
abatement_cost = float(DEMO_DATASET["abatement_cost"])
abatement_potential = float(DEMO_DATASET["abatement_potential"])
base_pd = float(DEMO_DATASET["base_pd"])
LGD_0 = float(DEMO_DATASET["lgd_0"])
tax_rate = 0.25
G = np.mean([0.7, 0.5, 0.4, 0.6])

sector_params      = SECTOR_CREDIT_PARAMS.get(sector,SECTOR_CREDIT_PARAMS["Manufacturing"])
alpha_dscr         = sector_params["alpha_dscr"]
beta_carbon_credit = sector_params["beta_carbon_credit"]
gamma_phys         = sector_params["gamma_physical"]
delta_gdp          = sector_params["delta_gdp"]
spread_beta        = SECTOR_SPREAD_SENSITIVITY.get(sector,1.0)
gdp_sensitivity    = SECTOR_GDP_SENSITIVITY.get(sector,1.0)
mc_params          = SECTOR_MC_PARAMS.get(sector,SECTOR_MC_PARAMS["Manufacturing"])
MC_CARBON_VOL=mc_params["carbon_vol"]; MC_PHYS_VOL=mc_params["physical_vol"]; MC_GDP_VOL=mc_params["gdp_vol"]
CLIMATE_DRIVER_VOL = np.array([MC_CARBON_VOL,MC_PHYS_VOL,MC_GDP_VOL])
CLIMATE_DRIVER_COV = np.outer(CLIMATE_DRIVER_VOL,CLIMATE_DRIVER_VOL)*CLIMATE_DRIVER_CORR

st.sidebar.markdown(f"""
<div style='background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:8px;padding:10px 12px;margin-top:12px;font-size:10px;line-height:1.7;color:{C['slate']};'>
<b style='color:{C['accent2']};'>Demo inputs</b><br>
Revenue: ₹{revenue_0:,.0f} Cr<br>
EAD: ₹{exposure_at_default:,.0f} Cr<br>
Emissions: {TOTAL_EMISSIONS/1e6:.1f} Mn tCO₂e<br>
Base PD: {base_pd:.2%} · LGD: {LGD_0:.0%}
</div>
""", unsafe_allow_html=True)

# ============================================================
# SIDEBAR — DOWNLOAD ALL RESULTS
# ============================================================
st.sidebar.divider()
st.sidebar.markdown(f"<div style='font-size:12px;font-weight:600;color:{C['accent2']};text-transform:uppercase;letter-spacing:.06em;'>📥 Export Results</div>", unsafe_allow_html=True)

any_ran = any(st.session_state.get(k, False) for k in ["transition_ran","physical_ran","brsr_ran","integrated_ran"])

if any_ran:
    import io

    def _build_excel_export():
        """
        Build a multi-sheet Excel workbook containing:
        - Summary sheet (all key metrics + model governance)
        - Assumptions sheet (all user inputs)
        - Transition Results sheet
        - Physical Risk sheet (if run)
        - BRSR sheet (if run)
        - Integrated Risk sheet (if run)
        - Monte Carlo sheet (if run)
        - Validation sheet (if run)
        """
        buf = io.BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            return None  # openpyxl not installed

        wb = openpyxl.Workbook()

        HDR_FILL  = PatternFill("solid", fgColor="062F2E")
        HDR_FONT  = Font(bold=True, color="00F5D4", size=11)
        VAL_FONT  = Font(color="E2E8F0", size=10)
        TILE_FILL = PatternFill("solid", fgColor="0B4D4B")
        TILE_FONT = Font(bold=True, color="22D3EE", size=12)
        thin_side = Side(style="thin", color="115E6D")
        THIN_BORDER = Border(bottom=thin_side)

        def _write_df(ws, df, start_row=1, start_col=1):
            """Write a DataFrame to a worksheet with header styling."""
            for ci, col in enumerate(df.columns, start_col):
                cell = ws.cell(row=start_row, column=ci, value=col)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(df.itertuples(index=False), start_row+1):
                for ci, val in enumerate(row, start_col):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = VAL_FONT
                    cell.border = THIN_BORDER
            for ci in range(start_col, start_col+len(df.columns)):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

        def _title(ws, title, row=1):
            cell = ws.cell(row=row, column=1, value=title)
            cell.fill = TILE_FILL; cell.font = TILE_FONT

        # ── Sheet 1: Summary ──
        ws_sum = wb.active; ws_sum.title = "Summary"
        ws_sum.sheet_view.showGridLines = False
        _title(ws_sum, f"ICCRE v{MODEL_VERSION} — Climate Risk Summary: {company_name}")
        gov_rows = [
            ("Model Version", MODEL_VERSION), ("Engine", ENGINE_BUILD),
            ("NGFS Data", NGFS_DATA_VERSION), ("Build Date", MODEL_BUILD_DATE),
            ("Company", company_name), ("Sector", sector),
            ("Reporting Year", REPORTING_YEAR), ("Baseline Scenario", BASELINE_SCENARIO),
            ("Export Timestamp", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ]
        for i,(k,v) in enumerate(gov_rows, 3):
            ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True, color="94A3B8")
            ws_sum.cell(row=i, column=2, value=str(v)).font = Font(color="E2E8F0")

        row_s = 14
        df_sum_data = []
        if st.session_state.get("transition_ran"):
            df_t_x = st.session_state.get("df_transition")
            if isinstance(df_t_x, pd.DataFrame) and not df_t_x.empty:
                df_sum_data.append({"Metric":"Peak Transition PD","Value":f"{df_t_x['PD_Transition'].max():.4f}"})
                df_sum_data.append({"Metric":"Max ECL (₹ Cr)","Value":f"{df_t_x['ECL_Transition'].max():.2f}"})
                df_sum_data.append({"Metric":"Min DSCR","Value":f"{df_t_x['DSCR'].min():.3f}"})
                df_sum_data.append({"Metric":"Max Stranded Assets (₹ Cr)","Value":f"{df_t_x['Stranded_Assets'].max():.0f}"})
        if st.session_state.get("physical_ran"):
            ps_x = st.session_state.get("phys_summary", {})
            if ps_x:
                df_sum_data.append({"Metric":"Physical Risk PD","Value":f"{ps_x.get('Physical Risk PD',0):.4f}"})
                df_sum_data.append({"Metric":"Physical ΔECL (₹ Cr)","Value":f"{ps_x.get('ΔECL (₹ Cr)',0):.2f}"})
        if st.session_state.get("brsr_ran"):
            df_sum_data.append({"Metric":"BRSR PD Uplift (bps)","Value":f"{st.session_state.get('brsr_pd_adj',0)*10000:.1f}"})
        if st.session_state.get("mc_results"):
            mc_x = st.session_state["mc_results"]
            df_sum_data.append({"Metric":"MC Mean PD","Value":f"{mc_x.get('Mean_PD',0):.4f}"})
            df_sum_data.append({"Metric":"MC Climate VaR 95% (₹ Cr)","Value":f"{mc_x.get('ECL_95',0):.2f}"})
        if df_sum_data:
            _write_df(ws_sum, pd.DataFrame(df_sum_data), start_row=row_s)

        # ── Sheet 2: Assumptions ──
        ws_ass = wb.create_sheet("Assumptions")
        ws_ass.sheet_view.showGridLines = False
        _title(ws_ass, "Model Inputs & Assumptions")
        assumptions = [
            ("Company Name", company_name), ("Sector", sector),
            ("Reporting Year", REPORTING_YEAR), ("Baseline Scenario", BASELINE_SCENARIO),
            ("Base Revenue (₹ Cr)", revenue_0), ("Base EBITDA Margin", ebitda_margin_0),
            ("Annual Interest (₹ Cr)", interest_payment), ("Total Assets (₹ Cr)", total_assets),
            ("EAD (₹ Cr)", exposure_at_default), ("Scope 1 Emissions (tCO₂e)", scope1),
            ("Scope 2 Emissions (tCO₂e)", scope2), ("Scope 3 Emissions (tCO₂e)", scope3),
            ("High-Carbon Assets (₹ Cr)", high_carbon_assets),
            ("Base PD", base_pd), ("Base LGD", LGD_0),
            ("Carbon Pass-Through", carbon_pass_through),
            ("Demand Elasticity", demand_elasticity),
            ("Beta Carbon (Transition)", beta_carbon_transition),
            ("Alpha DSCR (Sector)", alpha_dscr),
            ("Beta Carbon Credit (Sector)", beta_carbon_credit),
            ("USD/INR Rate", USD_INR),
            ("CPI Adjustment (2010→2026)", CARBON_PRICE_INFLATION_FACTOR),
            ("Governance Score (G)", round(G, 3)),
            ("NGFS Data Version", NGFS_DATA_VERSION),
            ("Model Version", MODEL_VERSION),
        ]
        df_ass = pd.DataFrame(assumptions, columns=["Parameter", "Value"])
        _write_df(ws_ass, df_ass, start_row=3)

        # ── Sheet 3: Transition Results ──
        if st.session_state.get("transition_ran"):
            df_t_x = st.session_state.get("df_transition")
            if isinstance(df_t_x, pd.DataFrame) and not df_t_x.empty:
                ws_tr = wb.create_sheet("Transition Results")
                ws_tr.sheet_view.showGridLines = False
                _title(ws_tr, "Transition Risk Engine Results — All Scenarios")
                _write_df(ws_tr, df_t_x.round(5), start_row=3)

        # ── Sheet 4: Physical Risk ──
        if st.session_state.get("physical_ran"):
            phys_a = st.session_state.get("phys_assets")
            phys_p = st.session_state.get("df_physical_projection")
            if isinstance(phys_a, pd.DataFrame) and not phys_a.empty:
                ws_ph = wb.create_sheet("Physical Risk")
                ws_ph.sheet_view.showGridLines = False
                _title(ws_ph, "Physical Risk — Asset-Level Results")
                _write_df(ws_ph, phys_a.round(4), start_row=3)
                if isinstance(phys_p, pd.DataFrame) and not phys_p.empty:
                    _title(ws_ph, "NGFS Physical Projections", row=len(phys_a)+6)
                    _write_df(ws_ph, phys_p.round(4), start_row=len(phys_a)+8)

        # ── Sheet 5: BRSR ──
        if st.session_state.get("brsr_ran"):
            bs_x = st.session_state.get("brsr_summary")
            if isinstance(bs_x, pd.DataFrame) and not bs_x.empty:
                ws_br = wb.create_sheet("BRSR Diagnostics")
                ws_br.sheet_view.showGridLines = False
                _title(ws_br, "BRSR Core Diagnostics")
                _write_df(ws_br, bs_x.T.reset_index().rename(columns={"index":"Indicator",0:"Value"}), start_row=3)
                bf_x = st.session_state.get("brsr_flags")
                if isinstance(bf_x, pd.DataFrame) and not bf_x.empty:
                    _title(ws_br, "Active BRSR Flags", row=len(bs_x.T)+7)
                    _write_df(ws_br, bf_x, start_row=len(bs_x.T)+9)

        # ── Sheet 6: Integrated Risk ──
        if st.session_state.get("integrated_ran"):
            df_int_x = st.session_state.get("df_integrated_summary")
            if isinstance(df_int_x, pd.DataFrame) and not df_int_x.empty:
                ws_int = wb.create_sheet("Integrated Risk")
                ws_int.sheet_view.showGridLines = False
                _title(ws_int, "Integrated Climate Risk — ISSB S2 Aligned Summary")
                _write_df(ws_int, df_int_x, start_row=3)

        # ── Sheet 7: Monte Carlo ──
        mc_x = st.session_state.get("mc_results")
        if mc_x:
            ws_mc = wb.create_sheet("Monte Carlo")
            ws_mc.sheet_view.showGridLines = False
            _title(ws_mc, "Monte Carlo Climate Stress Results")
            mc_df = pd.DataFrame([{"Metric": k, "Value": round(v, 5)} for k, v in mc_x.items()])
            _write_df(ws_mc, mc_df, start_row=3)

        # ── Sheet 8: Validation ──
        val_x = st.session_state.get("validation_results")
        if val_x:
            ws_vl = wb.create_sheet("Validation")
            ws_vl.sheet_view.showGridLines = False
            _title(ws_vl, "Model Validation Results")
            vl_df = pd.DataFrame([{"Metric": k, "Value": round(v, 6) if isinstance(v, float) else v} for k, v in val_x.items()])
            _write_df(ws_vl, vl_df, start_row=3)

        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    xlsx_bytes = _build_excel_export()
    if xlsx_bytes:
        st.sidebar.download_button(
            label="⬇️ Download Full Report (Excel)",
            data=xlsx_bytes,
            file_name=f"ICCRE_{company_name.replace(' ','_')}_{REPORTING_YEAR}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            help="Downloads all results, assumptions, and model governance in a multi-sheet Excel workbook",
        )
    else:
        st.sidebar.caption("Install openpyxl for Excel export: `pip install openpyxl`")

    # JSON export (always available, no extra dependency)
    import json as _json
    _export_payload = {
        "metadata": {"model_version": MODEL_VERSION, "company": company_name,
                     "sector": sector, "reporting_year": REPORTING_YEAR,
                     "timestamp": datetime.utcnow().isoformat()},
        "assumptions": {"revenue": revenue_0, "ebitda_margin": ebitda_margin_0,
                        "interest": interest_payment, "ead": exposure_at_default,
                        "scope1": scope1, "scope2": scope2, "scope3": scope3,
                        "base_pd": base_pd, "lgd": LGD_0},
    }
    if st.session_state.get("df_integrated_summary") is not None:
        _export_payload["integrated_risk"] = st.session_state["df_integrated_summary"].to_dict("records")
    if st.session_state.get("mc_results"):
        _export_payload["monte_carlo"] = st.session_state["mc_results"]

    st.sidebar.download_button(
        label="⬇️ Download Summary (JSON)",
        data=_json.dumps(_export_payload, indent=2, default=str).encode(),
        file_name=f"ICCRE_{company_name.replace(' ','_')}_{REPORTING_YEAR}.json",
        mime="application/json",
        width="stretch",
        help="Lightweight JSON export of key results and assumptions",
    )
else:
    st.sidebar.caption("Run at least one engine to enable exports.")

# ── PRICING / TIER FOOTER ────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown(f"""
<div style="background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:8px;
            padding:12px 14px;margin-bottom:4px;">
  <div style="font-size:11px;font-weight:700;color:{C['accent2']};text-transform:uppercase;
              letter-spacing:.07em;margin-bottom:8px;">Access Tiers</div>

  <div style="font-size:10px;color:{C['slate']};margin-bottom:6px;">
    <span style="color:{C['mint']};font-weight:600;">✓ Free</span>
    &nbsp;— Full access, single user<br>
    All 11 modules · Excel &amp; JSON export
  </div>

  <div style="font-size:10px;color:{C['slate']};margin-bottom:6px;">
    <span style="color:{C['amber']};font-weight:600;">₹ 2–3 L / year</span>
    &nbsp;— Professional<br>
    Custom calibration · Email support
  </div>

  <div style="font-size:10px;color:{C['slate']};margin-bottom:8px;">
    <span style="color:{C['coral']};font-weight:600;">₹ 12–18 L / year</span>
    &nbsp;— Enterprise<br>
    Multi-user · API · White-label · SLA
  </div>

  <a href="mailto:{CONTACT_EMAIL}?subject=ICCRE Access Request"
     style="display:block;background:{C['accent']};color:{C['bg_dark']};
            text-align:center;padding:6px 0;border-radius:6px;font-size:11px;
            font-weight:700;text-decoration:none;letter-spacing:.02em;">
    ✉ Request Access / Demo
  </a>
</div>

<div style="font-size:9px;color:{C['bg_ocean']};text-align:center;margin-top:4px;">
  {PRODUCT_URL} · {CONTACT_EMAIL}
</div>
""", unsafe_allow_html=True)

# ============================================================
# TRANSITION RISK ENGINE (v1.1 all fixes preserved)
# ============================================================
def run_transition_engine(df_long, selected_scenarios, revenue_0, ebitda_margin_0,
        interest_payment, TOTAL_EMISSIONS, high_carbon_assets, exposure_at_default,
        carbon_pass_through, demand_elasticity, price_elasticity, beta_carbon_transition,
        beta_carbon_credit, USD_INR, G, base_pd, LGD_0, abatement_cost,
        abatement_potential, planned_capex, P_START, P_FULL, MAX_STRANDING,
        margin_erosion_rate, margin_floor, alpha_dscr, tax_rate,
        CARBON_PRICE_INFLATION_FACTOR, baseline_temp=1.5, temp_sensitivity=0.02,
        gamma_phys=0.5, delta_gdp=0.8):
    results = []
    for scen in selected_scenarios:
        d      = df_long[df_long["Scenario"]==scen]
        d_base = df_long[df_long["Scenario"]==BASELINE_SCENARIO]
        for y in sorted(d["Year"].unique()):
            try:
                carbon_price = d.loc[(d["Year"]==y)&d["Variable"].str.contains("Carbon",case=False),"Value"].iloc[0]
                temp         = d.loc[(d["Year"]==y)&d["Variable"].str.contains("Temperature",case=False),"Value"].iloc[0]
                gdp          = d.loc[(d["Year"]==y)&d["Variable"].str.contains("GDP",case=False),"Value"].iloc[0]
            except Exception:
                continue
            try:
                gdp_base  = d_base.loc[(d_base["Year"]==y)&d_base["Variable"].str.contains("GDP",case=False),"Value"].iloc[0]
                gdp_shock = (gdp-gdp_base)/max(gdp_base,1e-6)
            except:
                gdp_shock = 0.0
            # FIX-05: two-step CPI conversion
            carbon_price_adj  = carbon_price * CARBON_PRICE_INFLATION_FACTOR
            gross_carbon_cost = TOTAL_EMISSIONS * carbon_price_adj * USD_INR / 1e7
            net_carbon_cost   = gross_carbon_cost * (1-carbon_pass_through)
            # Revenue (FIX-03, FIX-06)
            revenue = revenue_0*(1+gdp_sensitivity*gdp_shock)
            revenue *= (1+demand_elasticity*(net_carbon_cost/max(revenue_0,1)))
            revenue *= (1+price_elasticity*(net_carbon_cost/max(revenue_0,1)))  # FIX-03
            physical_loss = temp_sensitivity*max(0,temp-baseline_temp)
            revenue *= (1-physical_loss)
            t_idx = max(0,y-2025)
            margin_t   = max(margin_floor,ebitda_margin_0*(1-margin_erosion_rate*t_idx))
            ebitda     = revenue*margin_t - net_carbon_cost
            ebitda_adj = ebitda*(1-(1-G)*0.15)
            carbon_burden = net_carbon_cost/max(revenue,1)
            credit_spread = BASE_CREDIT_SPREAD*(1+spread_beta*carbon_burden)
            interest_stressed = interest_payment*(1+credit_spread)
            dscr = ebitda_adj/max(interest_stressed,1e-6)
            # FIX-04: signed gap; FIX-02: no gamma_phys; FIX-06: no delta_gdp direct
            dscr_gap = np.clip(1.5-dscr,-4.0,6.0)
            logit_pd = logit(base_pd)+alpha_dscr*dscr_gap+beta_carbon_credit*carbon_burden
            pd_t = np.clip(sigmoid(logit_pd),PD_FLOOR,PD_CAP)
            lgd_t = np.clip(LGD_0*(1+0.2*carbon_burden+LGD_PHYSICAL_MULTIPLIER*physical_loss),0,1)
            ecl   = pd_t*lgd_t*exposure_at_default/1e3
            ratio = 0 if carbon_price_adj<P_START else min((carbon_price_adj-P_START)/(P_FULL-P_START),1)
            stranded = high_carbon_assets*min(ratio,MAX_STRANDING)
            req_capex = (scope1+scope2)*abatement_potential*abatement_cost/1e7
            capex_gap = req_capex-planned_capex
            results.append({"Scenario":scen,"Year":y,"Revenue":revenue,"EBITDA":ebitda_adj,
                "EBITDA_Margin":ebitda_adj/max(revenue,1),"Carbon_Burden":carbon_burden,
                "DSCR":dscr,"PD_Transition":pd_t,"LGD":lgd_t,"ECL_Transition":ecl,
                "Stranded_Assets":stranded,"CAPEX_Gap":capex_gap,
                "Physical_Loss":physical_loss,"GDP_Shock":gdp_shock,"Carbon_Price_Adj":carbon_price_adj})
    df = pd.DataFrame(results)
    if df.empty: raise ValueError("Transition engine produced no outputs.")
    return df.sort_values(["Scenario","Year"]).reset_index(drop=True)

# ============================================================
# PHYSICAL RISK HELPERS  (PHYS-01)
# ============================================================
def extract_ngfs_temperature_path(df_long, scenarios, baseline_scenario, years):
    records = []
    for scen in scenarios:
        d_s = df_long[df_long["Scenario"]==scen]
        d_b = df_long[df_long["Scenario"]==baseline_scenario]
        for yr in years:
            tr = d_s.loc[(d_s["Year"]==yr)&d_s["Variable"].str.contains("Temperature",case=False),"Value"]
            br = d_b.loc[(d_b["Year"]==yr)&d_b["Variable"].str.contains("Temperature",case=False),"Value"]
            if tr.empty or br.empty: continue
            tv = float(tr.iloc[0]); bv = float(br.iloc[0])
            records.append({"Scenario":scen,"Year":yr,"Temp_C":round(tv,3),"Delta_T_vs_Baseline":round(max(0.0,tv-bv),3)})
    return pd.DataFrame(records)

def project_physical_risk_ngfs(df_ngfs_temp, baseline_damage_index, baseline_revenue_loss_cr,
        total_ebitda_cr, interest_cr, baseline_pd, lgd, ead_cr, gamma_phys=0.5,
        dscr_sensitivity=1.0, pd_floor=0.0005, pd_cap=0.35):
    """
    PHYS-01: IPCC AR6 damage function replaces flat 1+0.25*ΔT
    damage_mult = 1 + 0.20·ΔT + 0.04·ΔT²
    P10/P90 uncertainty: σ = 0.08·ΔT
    """
    records = []
    for _, row in df_ngfs_temp.iterrows():
        dt = row["Delta_T_vs_Baseline"]
        # IPCC AR6 non-linear damage (replaces 0.25*dt)
        dm       = 1.0 + 0.20*dt + 0.04*dt**2
        sigma    = 0.08*dt
        dm_p10   = max(1.0, dm-1.645*sigma)
        dm_p90   = dm+1.645*sigma
        rl_p50   = baseline_revenue_loss_cr*dm
        rl_p10   = baseline_revenue_loss_cr*dm_p10
        rl_p90   = baseline_revenue_loss_cr*dm_p90
        # DSCR (physical stress path)
        ebitda_s = total_ebitda_cr - rl_p50*(total_ebitda_cr/max(baseline_revenue_loss_cr*10,1))
        dscr_p   = ebitda_s/max(interest_cr,1e-6)
        dscr_gap = np.clip(1.5-dscr_p,-4.0,6.0)  # FIX-04 signed
        logit_p  = logit(baseline_pd)+dscr_sensitivity*dscr_gap+gamma_phys*baseline_damage_index
        pd_p     = np.clip(sigmoid(logit_p),pd_floor,pd_cap)
        ecl_p    = pd_p*lgd*ead_cr/1e3
        chronic  = min(0.7,dt/3.0)
        records.append({
            "Scenario":row["Scenario"],"Year":int(row["Year"]),
            "Delta_T_C":round(dt,3),"Temp_C":round(row["Temp_C"],3),
            "Damage_Multiplier":round(dm,3),
            "Revenue_Loss_P10_Cr":round(rl_p10,2),
            "Revenue_Loss_P50_Cr":round(rl_p50,2),
            "Revenue_Loss_P90_Cr":round(rl_p90,2),
            "DSCR_Physical":round(dscr_p,3),"PD_Physical":round(pd_p,5),
            "ECL_Physical_Cr":round(ecl_p,3),
            "Chronic_Loss_Cr":round(rl_p50*chronic,2),
            "Acute_Loss_Cr":round(rl_p50*(1-chronic),2),
        })
    return pd.DataFrame(records).sort_values(["Scenario","Year"]).reset_index(drop=True)

# ============================================================
# TAB LAYOUT
# ============================================================
# Tab order:  Dashboard → Transition → Physical → BRSR Core → Targets → Integrated → Plots → AI → …
# BRSR is placed BEFORE Targets so BRSR results feed into BRSR-based target planning in Targets tab.
(dashboard_tab, transition_tab, physical_tab, brsr_tab, targets_tab,
 integrated_tab, plots_tab, ai_tab, methodology_tab, validation_tab, calibration_tab,
 access_tab,
) = st.tabs([
    "🏠 Dashboard",
    "⚡ Transition Risk",
    "🌍 Physical Risk",
    "📘 BRSR Core",
    "🎯 Targets",
    "🧩 Integrated Risk",
    "📈 Plots",
    "🤖 AI Narrative",
    "📖 Methodology",
    "🔬 Validation",
    "⚙️ Calibration",
    "🚀 Get Access",
])

# ============================================================
# TAB 0 — DASHBOARD  (UI-03)
# ============================================================
with dashboard_tab:
    st.markdown(f"<h2 style='color:{C['white']};margin-bottom:4px;'>Climate Risk Dashboard</h2>", unsafe_allow_html=True)
    st.caption(f"**{company_name}** · {sector} · Reporting Year {REPORTING_YEAR}")

    # Module status badges
    badges = []
    for label, key, color in [
        ("⚡ Transition","transition_ran",C["accent2"]),
        ("🌍 Physical","physical_ran",C["coral"]),
        ("🎯 Targets","targets_ran",C["amber"]),
        ("📘 BRSR","brsr_ran",C["mint"]),
    ]:
        ran = st.session_state.get(key,False)
        bg  = color if ran else C["bg_mid"]
        tc  = C["bg_dark"] if ran else C["slate"]
        badges.append(f"<span style='background:{bg};color:{tc};padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;'>{label} {'✓' if ran else '—'}</span>")
    st.markdown("<div style='display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;'>"+"".join(badges)+"</div>",unsafe_allow_html=True)

    transition_ran = st.session_state.get("transition_ran",False)
    physical_ran   = st.session_state.get("physical_ran",False)
    brsr_ran       = st.session_state.get("brsr_ran",False)
    integrated_ran = st.session_state.get("integrated_ran",False)

    if not (transition_ran or physical_ran or brsr_ran):
        # ── Welcome / Demo CTA ──
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,{C['card']},{C['bg_mid']});
                    border:1px solid {C['bg_ocean']};border-radius:14px;
                    padding:36px 40px;text-align:center;margin-top:8px;">
          <div style="font-size:40px;margin-bottom:10px;">🌍</div>
          <div style="font-size:22px;font-weight:700;color:{C['white']};margin-bottom:6px;">
            Welcome to ICCRE
          </div>
          <div style="font-size:14px;color:{C['accent2']};margin-bottom:4px;font-weight:600;">
            {PRODUCT_TAGLINE}
          </div>
          <div style="font-size:12px;color:{C['slate']};margin-bottom:24px;">
            {PRODUCT_SUBTITLE}
          </div>
          <div style="display:flex;justify-content:center;gap:32px;flex-wrap:wrap;margin-bottom:24px;">
            <div style="text-align:center;">
              <div style="font-size:28px;font-weight:700;color:{C['accent2']};">3</div>
              <div style="font-size:10px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;">NGFS Scenarios</div>
            </div>
            <div style="text-align:center;">
              <div style="font-size:28px;font-weight:700;color:{C['mint']};">5</div>
              <div style="font-size:10px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;">Indian Sectors</div>
            </div>
            <div style="text-align:center;">
              <div style="font-size:28px;font-weight:700;color:{C['amber']};">11</div>
              <div style="font-size:10px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;">Analytics Modules</div>
            </div>
            <div style="text-align:center;">
              <div style="font-size:28px;font-weight:700;color:{C['coral']};">~20</div>
              <div style="font-size:10px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;">Minutes to Full Report</div>
            </div>
          </div>
          <div style="background:{C['bg_dark']};border-radius:10px;padding:16px 20px;
                      margin-bottom:20px;text-align:left;max-width:520px;margin-left:auto;margin-right:auto;">
            <div style="font-size:11px;font-weight:700;color:{C['amber']};margin-bottom:8px;
                        text-transform:uppercase;letter-spacing:.06em;">⚡ Fastest Start — 3 Steps</div>
            <div style="font-size:12px;color:{C['off_white']};line-height:2.0;">
              1️⃣ Click <strong style="color:{C['amber']};">🎬 Load Demo</strong> in the sidebar<br>
              2️⃣ Go to <strong style="color:{C['accent2']};">⚡ Transition Risk</strong> tab → Click Run<br>
              3️⃣ See PD jump from 1.5% to 7.3% under Net Zero 2050
            </div>
          </div>
          <div style="display:flex;justify-content:center;gap:12px;flex-wrap:wrap;">
            <div style="background:{C['bg_dark']};border:1px solid {C['bg_mid']};border-radius:8px;
                        padding:10px 16px;text-align:center;min-width:150px;">
              <div style="font-size:10px;font-weight:700;color:{C['mint']};text-transform:uppercase;
                          letter-spacing:.06em;margin-bottom:4px;">Aligned with</div>
              <div style="font-size:11px;color:{C['off_white']};">ISSB IFRS S2</div>
              <div style="font-size:11px;color:{C['off_white']};">RBI 2024 Guidelines</div>
              <div style="font-size:11px;color:{C['off_white']};">SEBI BRSR Core</div>
            </div>
            <div style="background:{C['bg_dark']};border:1px solid {C['bg_mid']};border-radius:8px;
                        padding:10px 16px;text-align:center;min-width:150px;">
              <div style="font-size:10px;font-weight:700;color:{C['amber']};text-transform:uppercase;
                          letter-spacing:.06em;margin-bottom:4px;">Outputs include</div>
              <div style="font-size:11px;color:{C['off_white']};">PD · LGD · ECL</div>
              <div style="font-size:11px;color:{C['off_white']};">DSCR · Climate VaR</div>
              <div style="font-size:11px;color:{C['off_white']};">ICAAP Capital Signal</div>
            </div>
            <div style="background:{C['bg_dark']};border:1px solid {C['bg_mid']};border-radius:8px;
                        padding:10px 16px;text-align:center;min-width:150px;">
              <div style="font-size:10px;font-weight:700;color:{C['coral']};text-transform:uppercase;
                          letter-spacing:.06em;margin-bottom:4px;">Zero competitors</div>
              <div style="font-size:11px;color:{C['off_white']};">India-native</div>
              <div style="font-size:11px;color:{C['off_white']};">INR-denominated</div>
              <div style="font-size:11px;color:{C['off_white']};">BRSR-native</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        # Pull live metrics
        pd_t = ecl_t = dscr_t = stranded_t = capex_t = pd_p = ecl_p = brsr_pd = risk_score = None
        if transition_ran:
            df_t = st.session_state.get("df_transition")
            if isinstance(df_t,pd.DataFrame) and not df_t.empty:
                pd_t = df_t["PD_Transition"].max(); ecl_t = df_t["ECL_Transition"].max()
                dscr_t = df_t["DSCR"].min(); stranded_t = df_t["Stranded_Assets"].max()
                capex_t = df_t["CAPEX_Gap"].max()
        if physical_ran:
            ps = st.session_state.get("phys_summary",{})
            if isinstance(ps,dict):
                pd_p  = ps.get("Physical Risk PD",0)
                ecl_p = ps.get("ΔECL (₹ Cr)",0)
        if brsr_ran:
            bs = st.session_state.get("brsr_summary")
            if isinstance(bs,pd.DataFrame) and not bs.empty:
                brsr_pd   = st.session_state.get("brsr_pd_adj",0)
                risk_score = bs.iloc[0].get("Overall_Risk_Score",None)

        # KPI row
        kpis = []
        if pd_t is not None:   kpis.append(("Transition PD",f"{pd_t:.2%}",C["accent2"],"Max across scenarios"))
        if ecl_t is not None:  kpis.append(("Max ECL",f"₹{ecl_t:,.1f} Cr",C["amber"],"Expected credit loss"))
        if dscr_t is not None: kpis.append(("Min DSCR",f"{dscr_t:.2f}×",C["coral"] if dscr_t<1.2 else C["mint"],"Debt service coverage"))
        if pd_p is not None:   kpis.append(("Physical PD",f"{pd_p:.2%}",C["coral"],"Asset-level physical risk"))
        if brsr_pd is not None:kpis.append(("BRSR PD Uplift",f"+{brsr_pd*10000:.0f} bps",C["purple"],"Operational climate risk"))
        if stranded_t is not None: kpis.append(("Stranded Assets",f"₹{stranded_t:,.0f} Cr",C["amber"],"High-carbon asset risk"))

        cols = st.columns(min(len(kpis),6))
        for col_obj,(label,val,color,hint) in zip(cols,kpis):
            with col_obj:
                st.markdown(f"""
                <div class="kpi-card">
                  <div class="kpi-value" style="color:{color};">{val}</div>
                  <div class="kpi-label">{label}</div>
                  <div style="font-size:10px;color:{C['slate']};margin-top:4px;">{hint}</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div style='margin-top:16px;'></div>",unsafe_allow_html=True)

        # PD sparkline + risk gauge side by side
        if transition_ran and isinstance(st.session_state.get("df_transition"),pd.DataFrame):
            df_t = st.session_state["df_transition"]
            col_spark, col_gauge = st.columns([3,1])

            with col_spark:
                fig_spark = go.Figure()
                for scen in df_t["Scenario"].unique():
                    ds = df_t[df_t["Scenario"]==scen]
                    fig_spark.add_trace(go.Scatter(
                        x=ds["Year"],y=ds["PD_Transition"],mode="lines+markers",
                        name=scen[:28],line=dict(color=_scen_color(scen),width=2.5),
                        marker=dict(size=6),
                        hovertemplate="<b>%{x}</b><br>PD: %{y:.2%}<extra>"+scen[:20]+"</extra>"
                    ))
                fig_spark.update_layout(**_chart_layout("Transition PD Trajectory — All Scenarios",260))
                fig_spark.update_yaxes(tickformat=".1%")
                _ax_style(fig_spark)
                st.plotly_chart(fig_spark,width="stretch")

            with col_gauge:
                if pd_t is not None:
                    gauge_color = C["coral"] if pd_t>0.10 else C["amber"] if pd_t>0.05 else C["mint"]
                    fig_g = go.Figure(go.Indicator(
                        mode="gauge+number",
                        value=pd_t*100,
                        number={"suffix":"%","font":{"color":gauge_color,"size":22}},
                        gauge={
                            "axis":{"range":[0,30],"tickcolor":C["slate"]},
                            "bar":{"color":gauge_color},
                            "bgcolor":C["bg_dark"],
                            "steps":[
                                {"range":[0,5],"color":C["bg_mid"]},
                                {"range":[5,10],"color":"#1A3A2A"},
                                {"range":[10,20],"color":"#3A1A1A"},
                                {"range":[20,30],"color":"#4A1A1A"},
                            ],
                            "threshold":{"line":{"color":"white","width":2},"thickness":0.75,"value":pd_t*100}
                        },
                        title={"text":"Peak PD","font":{"color":C["slate"],"size":12}}
                    ))
                    fig_g.update_layout(height=200,paper_bgcolor=C["bg_dark"],margin=dict(l=10,r=10,t=30,b=10))
                    st.plotly_chart(fig_g,width="stretch")

        # BRSR + Physical summary row
        if brsr_ran or physical_ran:
            bcol, pcol = st.columns(2)
            if brsr_ran and risk_score is not None:
                with bcol:
                    sev_color = C["mint"] if risk_score<30 else C["amber"] if risk_score<60 else C["coral"]
                    sev_text  = "Low" if risk_score<30 else "Moderate" if risk_score<60 else "High"
                    st.markdown(f"""
                    <div style="background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:10px;padding:14px 16px;">
                      <div style="font-size:12px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;">BRSR Operational Risk</div>
                      <div style="font-size:26px;font-weight:600;color:{sev_color};font-family:monospace;">{risk_score}/100</div>
                      <div style="font-size:13px;color:{sev_color};font-weight:600;">{sev_text} Risk · +{brsr_pd*10000:.0f} bps PD uplift</div>
                    </div>""",unsafe_allow_html=True)
            if physical_ran and pd_p is not None:
                with pcol:
                    pc = C["coral"] if pd_p>0.05 else C["amber"] if pd_p>0.02 else C["mint"]
                    st.markdown(f"""
                    <div style="background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:10px;padding:14px 16px;">
                      <div style="font-size:12px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;">Physical Risk (Portfolio)</div>
                      <div style="font-size:26px;font-weight:600;color:{pc};font-family:monospace;">{pd_p:.2%}</div>
                      <div style="font-size:13px;color:{C['slate']};">ΔECL: ₹{ecl_p:.2f} Cr · Asset-level aggregation</div>
                    </div>""",unsafe_allow_html=True)

        # Integrated risk summary — shown once integrated tab has been computed
        if integrated_ran:
            df_int_dash = st.session_state.get("df_integrated_summary")
            st.markdown(f"<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-size:13px;color:{C['slate']};text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px;'>🧩 Integrated Risk (latest run)</div>", unsafe_allow_html=True)
            # Pull integrated metrics from the stored summary table
            _pd_int = _ecl_int = _risk_sc = _capital = None
            if isinstance(df_int_dash, pd.DataFrame) and not df_int_dash.empty:
                def _get(metric_substr):
                    rows = df_int_dash[df_int_dash["Metric"].str.contains(metric_substr, case=False, na=False)]
                    return rows["Value"].iloc[0] if not rows.empty else None
                _pd_int   = _get("Integrated PD")
                _ecl_int  = _get("Integrated ECL")
                _capital  = _get("Capital Stress")
                _risk_sc  = _get("Risk Score")

            if _risk_sc is not None:
                try:
                    _rs_val = float(str(_risk_sc))
                    _rag_c  = C["mint"] if _rs_val < 30 else C["amber"] if _rs_val < 60 else C["coral"] if _rs_val < 80 else "#F72585"
                    _rag_t  = "Low" if _rs_val < 30 else "Moderate" if _rs_val < 60 else "Severe" if _rs_val < 80 else "Extreme"
                except Exception:
                    _rag_c = C["amber"]; _rag_t = "—"; _rs_val = 0

                ic1, ic2, ic3, ic4 = st.columns(4)
                ic1.markdown(f"""<div class="kpi-card"><div class="kpi-value" style="color:{_rag_c};">{_rag_t}</div><div class="kpi-label">Climate Risk Level</div></div>""", unsafe_allow_html=True)
                ic2.markdown(f"""<div class="kpi-card"><div class="kpi-value" style="color:{C['accent2']};">{_pd_int or '—'}</div><div class="kpi-label">Integrated PD</div></div>""", unsafe_allow_html=True)
                ic3.markdown(f"""<div class="kpi-card"><div class="kpi-value" style="color:{C['amber']};">{_ecl_int or '—'} Cr</div><div class="kpi-label">Integrated ECL</div></div>""", unsafe_allow_html=True)
                ic4.markdown(f"""<div class="kpi-card"><div class="kpi-value" style="color:{C['coral'] if _capital in ['Severe','Elevated'] else C['mint']};">{_capital or '—'}</div><div class="kpi-label">Capital Signal</div></div>""", unsafe_allow_html=True)

# ============================================================
# TAB 1 — TRANSITION RISK
# ============================================================
with transition_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>⚡ Transition Risk Engine</h2>",unsafe_allow_html=True)
    st.caption("NGFS scenario-driven financial transmission: Carbon price → Revenue → DSCR → PD → ECL")

    if not st.session_state.get("enable_transition",False):
        st.info("Enable **Transition Risk** in the sidebar to proceed.")
    else:
        with st.expander("ℹ️ v1.1 Corrections Active",expanded=False):
            st.markdown("""
            | Fix | Description |
            |-----|-------------|
            | FIX-01 | β_transition (slider) and β_credit (sector table) separated |
            | FIX-02 | Physical double-count removed from logit PD |
            | FIX-03 | Price elasticity uses revenue₀ denominator |
            | FIX-04 | DSCR gap is signed ± (Winsorised ±4/6) |
            | FIX-05 | Two-step CPI carbon price conversion |
            | FIX-06 | GDP single transmission path only |
            | FIX-07 | MC GDP sign consistent with deterministic engine |
            """)

        all_scenarios = sorted(df_long["Scenario"].unique())
        selected_scenarios = st.multiselect("NGFS Scenarios",options=all_scenarios,default=all_scenarios)
        run_btn = st.button("▶ Run Transition Risk Engine",type="primary")

        if not run_btn:
            st.info("Select scenarios and click **Run Transition Risk Engine**.")
        elif not selected_scenarios:
            st.warning("Select at least one scenario.")
        else:
            with st.spinner("Running transition risk engine..."):
                df_transition = run_transition_engine(
                    df_long,selected_scenarios,revenue_0,ebitda_margin_0,
                    interest_payment,TOTAL_EMISSIONS,high_carbon_assets,
                    exposure_at_default,carbon_pass_through,demand_elasticity,
                    price_elasticity,beta_carbon_transition,beta_carbon_credit,
                    USD_INR,G,base_pd,LGD_0,abatement_cost,abatement_potential,
                    planned_capex,P_START,P_FULL,MAX_STRANDING,margin_erosion_rate,
                    margin_floor,alpha_dscr,tax_rate,CARBON_PRICE_INFLATION_FACTOR,
                    delta_gdp=delta_gdp
                )
            st.session_state["df_transition"]  = df_transition
            st.session_state["transition_ran"] = True

            # Scenario summary table
            df_sum = (df_transition.groupby("Scenario").agg({
                "Carbon_Burden":"max","EBITDA_Margin":"min","DSCR":"min",
                "PD_Transition":"max","ECL_Transition":"max",
                "Stranded_Assets":"max","CAPEX_Gap":"max",
            }).round(4).reset_index())

            # Styled summary
            st.markdown(f"<h3 style='color:{C['white']};margin-top:16px;'>Scenario Risk Summary</h3>",unsafe_allow_html=True)

            def _color_pd(val):
                if isinstance(val,float):
                    if val>0.10: return f"color:{C['coral']};font-weight:600"
                    if val>0.05: return f"color:{C['amber']};font-weight:600"
                    return f"color:{C['mint']};font-weight:600"
                return ""

            st.dataframe(
                df_sum.style
                    .format({"Carbon_Burden":"{:.2%}","EBITDA_Margin":"{:.2%}","DSCR":"{:.2f}",
                              "PD_Transition":"{:.3%}","ECL_Transition":"{:.1f}",
                              "Stranded_Assets":"{:.0f}","CAPEX_Gap":"{:.0f}"})
                    .map(_color_pd, subset=["PD_Transition"])
                    .background_gradient(subset=["ECL_Transition"],cmap="Reds")
                    .set_properties(**{"background-color":C["bg_dark"],"color":C["text"]}),
                width="stretch", hide_index=True
            )

            # Headline KPIs
            st.markdown("<h3 style='margin-top:16px;'>Headline Indicators</h3>",unsafe_allow_html=True)
            k1,k2,k3,k4 = st.columns(4)
            k1.metric("Max PD",f"{df_transition['PD_Transition'].max():.2%}")
            k2.metric("Max ECL (₹ Cr)",f"{df_transition['ECL_Transition'].max():,.1f}")
            k3.metric("Min DSCR",f"{df_transition['DSCR'].min():.2f}×")
            k4.metric("Max Stranded (₹ Cr)",f"{df_transition['Stranded_Assets'].max():,.0f}")

            # Charts
            fig_pd = go.Figure()
            for scen in df_transition["Scenario"].unique():
                ds = df_transition[df_transition["Scenario"]==scen]
                fig_pd.add_trace(go.Scatter(x=ds["Year"],y=ds["PD_Transition"],mode="lines+markers",
                    name=scen[:30],line=dict(color=_scen_color(scen),width=2.5),marker=dict(size=6)))
            fig_pd.update_layout(**_chart_layout("PD Trajectory by NGFS Scenario",320))
            fig_pd.update_yaxes(tickformat=".1%")
            _ax_style(fig_pd)
            st.plotly_chart(fig_pd,width="stretch")

            fig_ecl = go.Figure()
            for scen in df_transition["Scenario"].unique():
                ds = df_transition[df_transition["Scenario"]==scen]
                fig_ecl.add_trace(go.Bar(x=ds["Year"],y=ds["ECL_Transition"],name=scen[:30],
                    marker_color=_scen_color(scen),opacity=0.85))
            fig_ecl.update_layout(**_chart_layout("Expected Credit Loss (₹ Cr) — All Scenarios",300),barmode="group")
            _ax_style(fig_ecl)
            st.plotly_chart(fig_ecl,width="stretch")

            fig_dscr = go.Figure()
            for scen in df_transition["Scenario"].unique():
                ds = df_transition[df_transition["Scenario"]==scen]
                fig_dscr.add_trace(go.Scatter(x=ds["Year"],y=ds["DSCR"],mode="lines+markers",
                    name=scen[:30],line=dict(color=_scen_color(scen),width=2)))
            fig_dscr.add_hline(y=1.2,line_dash="dash",line_color=C["coral"],
                annotation_text="1.2× covenant",annotation_font_color=C["coral"])
            fig_dscr.add_hline(y=1.5,line_dash="dot",line_color=C["amber"],
                annotation_text="1.5× threshold",annotation_font_color=C["amber"])
            fig_dscr.update_layout(**_chart_layout("DSCR Stress Trajectory",300))
            _ax_style(fig_dscr)
            st.plotly_chart(fig_dscr,width="stretch")

            st.success("✅ Transition Risk Engine v1.1 executed")
            log_model_run("Transition",{"company":company_name,"sector":sector,
                "pd_max":df_transition["PD_Transition"].max(),"ecl_max":df_transition["ECL_Transition"].max()})
            st.session_state["df_transition_summary"] = df_sum

# ============================================================
# TAB 2 — PHYSICAL RISK  (PHYS-01)
# ============================================================
with physical_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>🌍 Physical Risk Engine</h2>",unsafe_allow_html=True)
    st.caption("Asset-level GIS-based physical risk with NGFS-aligned temperature pathways and IPCC AR6 damage functions")

    if not st.session_state.get("enable_physical",False):
        st.info("Enable **Physical Risk** in the sidebar.")
    else:
        _gis_missing = _ensure_gis_files()
        if _gis_missing:
            st.warning("⚠️ Some GIS files failed to load. Physical Risk tab may show zeros.")

        with st.expander("📐 Methodology: v1.2 corrections",expanded=False):
            st.markdown("""
            **v1.1 (old):** `proj_loss = TOTAL_REV_LOSS × (1 + 0.25 × ΔT)` — flat multiplier, no citation.

            **v1.2 (corrected):**
            - Damage function: `1 + 0.20·ΔT + 0.04·ΔT²` (IPCC AR6 economic damage, non-linear)
            - ΔT from NGFS scenario temperature pathways (same data as transition engine)
            - P10/P50/P90 uncertainty bands: `σ = 0.08 × ΔT`
            - Signed DSCR gap consistent with FIX-04
            - Chronic (temperature-driven) vs acute (event-driven) loss split
            """)

        st.subheader("🏭 Asset Register")
        asset_df = st.data_editor(pd.DataFrame({
            "asset_id":["A1","A2","A3"],
            "asset_type":["Steel Plant","Rolling Mill","Port Logistics Yard"],
            "latitude":[22.80,23.55,20.32],"longitude":[86.20,87.32,86.61],
            "revenue_inr_cr":[3500.0,2000.0,1200.0],
            "base_pd":[0.015,0.015,0.015],"lgd":[0.45,0.45,0.45],
        }),num_rows="fixed",width="stretch",disabled=True)

        fc1,fc2 = st.columns(2)
        fc1.metric("EBITDA Margin", f"{ebitda_margin_0:.0%}")
        fc2.metric("Annual Interest", f"₹{interest_payment:,.0f} Cr")
        EBITDA_MARGIN_PHYS = float(ebitda_margin_0)
        INTEREST_PHYS      = float(interest_payment)
        LGD_PHYS=float(LGD_0); EAD_PHYS=float(exposure_at_default)

        # Scenario alignment
        transition_ran_phys = st.session_state.get("transition_ran",False)
        if transition_ran_phys:
            df_t_p = st.session_state.get("df_transition")
            if isinstance(df_t_p,pd.DataFrame) and not df_t_p.empty:
                phys_scenarios = sorted(df_t_p["Scenario"].unique())
                phys_years     = sorted(df_t_p["Year"].unique())
                st.success(f"✅ Aligned with transition scenarios: {phys_scenarios}")
        else:
            phys_scenarios = [BASELINE_SCENARIO]; phys_years = [REPORTING_YEAR]
            st.info("No transition engine run — using baseline scenario only.")

        run_phys = st.button("▶ Run Physical Risk Engine",type="primary",key="run_phys_btn")
        if not run_phys:
            st.info("Click **Run Physical Risk Engine** to proceed.")
        else:
            df = asset_df.copy()
            V={"flood":0.6,"heat":0.7,"cyclone":0.6}; KAPPA={"flood":0.4,"heat":0.2,"cyclone":0.5}
            MAX_DOWNTIME={"Steel Plant":90,"Rolling Mill":60,"Port Logistics Yard":45}

            with st.spinner("Extracting hazard data..."):
                # Flood
                # Memory-safe raster access: read only a small window around each asset.
                # Avoid src.read(1), which loads the full GeoTIFF and can crash Streamlit Cloud.
                flood_vals=[]
                try:
                    if rasterio is None or rowcol is None or Window is None:
                        raise RuntimeError(RASTERIO_IMPORT_ERROR or "rasterio unavailable")
                    with rasterio.open(r"Data/floodMapGL_rp100y.tif") as src:
                        pk=max(abs(src.res[0])*111, 1e-6)
                        bp=max(1, int(5/pk))
                        for _,a in df.iterrows():
                            try:
                                ri,ci=rowcol(src.transform,a["longitude"],a["latitude"])
                                r0=max(0,ri-bp); r1=min(src.height,ri+bp)
                                c0=max(0,ci-bp); c1=min(src.width,ci+bp)
                                if r1<=r0 or c1<=c0:
                                    flood_vals.append(0.0); continue
                                win=Window(c0,r0,c1-c0,r1-r0)
                                w=src.read(1, window=win, masked=True)
                                v=w.compressed() if hasattr(w, "compressed") else w.ravel()
                                v=v[(v>0)&(v<100)]
                                flood_vals.append(float(v.max()) if v.size else 0.0)
                            except Exception:
                                flood_vals.append(0.0)
                except Exception as e:
                    st.warning(f"Flood layer skipped: {e}")
                    flood_vals=[0.0]*len(df)
                df["flood_depth_m"]=flood_vals

                # Heat
                try:
                    heat=pd.read_csv(r"Data/era5_test_day_grid.csv")
                    heat["exceed"]=(heat["temp_c"]>=35).astype(int)
                    hg=heat.groupby(["lat","lon"],as_index=False)["exceed"].sum()
                except: hg=pd.DataFrame(columns=["lat","lon","exceed"])
                def nearest_heat(row):
                    if hg.empty: return 0
                    idx=((hg["lat"]-row["latitude"])**2+(hg["lon"]-row["longitude"])**2).idxmin()
                    return int(hg.loc[idx,"exceed"])
                df["heat_days"]=df.apply(nearest_heat,axis=1)

                # Cyclone
                try:
                    cyc=pd.read_csv(
                        r"Data/ibtracs.NI.list.v04r01.csv",
                        usecols=["LAT","LON","USA_WIND"],
                        low_memory=False
                    )
                    cyc=cyc.apply(pd.to_numeric,errors="coerce").dropna()
                    cyc["wind_kmh"]=cyc["USA_WIND"]*1.852
                except Exception as e:
                    st.warning(f"Cyclone layer skipped: {e}")
                    cyc=pd.DataFrame(columns=["LAT","LON","wind_kmh"])

                def max_wind(a):
                    if cyc.empty:
                        return 0.0
                    try:
                        lat1=np.radians(float(a["latitude"]))
                        lon1=np.radians(float(a["longitude"]))
                        lat2=np.radians(cyc["LAT"].astype(float).to_numpy())
                        lon2=np.radians(cyc["LON"].astype(float).to_numpy())
                        dlat=lat2-lat1; dlon=lon2-lon1
                        h=np.sin(dlat/2)**2 + np.cos(lat1)*np.cos(lat2)*np.sin(dlon/2)**2
                        d=6371*2*np.arcsin(np.sqrt(h))
                        vals=cyc.loc[d<=100,"wind_kmh"]
                        return float(vals.max()) if len(vals)>0 else 0.0
                    except Exception:
                        return 0.0
                df["max_wind_kmh"]=df.apply(max_wind,axis=1)

                df["H_flood"]=df["flood_depth_m"].apply(flood_damage)
                df["H_heat"]=df["heat_days"].apply(heat_damage)
                df["H_cyclone"]=df["max_wind_kmh"].apply(cyclone_damage)
                df["D_total"]=(df["H_flood"]*V["flood"]*KAPPA["flood"]+df["H_heat"]*V["heat"]*KAPPA["heat"]+df["H_cyclone"]*V["cyclone"]*KAPPA["cyclone"]).clip(0,1)
                df["downtime_days"]=df.apply(lambda r: r["D_total"]*MAX_DOWNTIME.get(r["asset_type"],60),axis=1)
                df["revenue_loss"]=df["downtime_days"]/365*df["revenue_inr_cr"]

                TOTAL_REV_LOSS=df["revenue_loss"].sum()
                TOTAL_EBITDA=df["revenue_inr_cr"].sum()*EBITDA_MARGIN_PHYS
                DELTA_EBITDA=TOTAL_REV_LOSS*EBITDA_MARGIN_PHYS
                DSCR_PHYS_BASE=(TOTAL_EBITDA-DELTA_EBITDA)/max(INTEREST_PHYS,1e-6)
                df["PD_Physical"]=(df["base_pd"]*(1+gamma_phys*df["D_total"])).clip(0,1)
                df["ECL_asset"]=df["PD_Physical"]*df["lgd"]*EAD_PHYS/max(len(df),1)
                PD_PHYS=df["PD_Physical"].mean(); DELTA_ECL=df["ECL_asset"].sum()/1e3

            # NGFS projection (PHYS-01: IPCC damage function)
            df_ngfs_temp = extract_ngfs_temperature_path(df_long,phys_scenarios,BASELINE_SCENARIO,phys_years)
            df_phys_proj = project_physical_risk_ngfs(
                df_ngfs_temp,df["D_total"].mean(),TOTAL_REV_LOSS,TOTAL_EBITDA,INTEREST_PHYS,
                PD_PHYS,LGD_PHYS,EAD_PHYS,gamma_phys)

            st.session_state["phys_assets"]=df
            st.session_state["phys_summary"]={"Total Revenue Loss (₹ Cr)":TOTAL_REV_LOSS,
                "EBITDA Loss (₹ Cr)":DELTA_EBITDA,"Post-Risk DSCR":DSCR_PHYS_BASE,
                "Physical Risk PD":PD_PHYS,"ΔECL (₹ Cr)":DELTA_ECL,"Assets Analysed":len(df)}
            st.session_state["df_physical_projection"]=df_phys_proj
            st.session_state["physical_ran"]=True
            st.success("✅ Physical Risk Engine v1.2 executed")

            # Summary KPIs
            pc1,pc2,pc3,pc4,pc5=st.columns(5)
            pc1.metric("Total Revenue Loss",f"₹{TOTAL_REV_LOSS:.1f} Cr")
            pc2.metric("EBITDA Impact",f"₹{DELTA_EBITDA:.1f} Cr")
            pc3.metric("Post-Risk DSCR",f"{DSCR_PHYS_BASE:.2f}×")
            pc4.metric("Physical PD",f"{PD_PHYS:.2%}")
            pc5.metric("ΔECL",f"₹{DELTA_ECL:.2f} Cr")

            # Asset vulnerability heatmap
            st.subheader("🗺️ Asset Vulnerability Heatmap")
            score_map={"H_flood":"Flood","H_heat":"Heat","H_cyclone":"Cyclone"}
            hm=df.set_index("asset_id")[[c for c in score_map if c in df.columns]].rename(columns={k:v for k,v in score_map.items() if k in df.columns})*100
            fig_hm=go.Figure(go.Heatmap(z=hm.values.T,x=hm.index.tolist(),y=hm.columns.tolist(),
                colorscale=[[0,C["mint"]],[0.3,C["amber"]],[0.6,"#F97316"],[1.0,C["coral"]]],
                text=[[f"{v:.0f}" for v in row] for row in hm.values.T],texttemplate="%{text}",
                colorbar=dict(title="Risk Score"),zmin=0,zmax=100))
            fig_hm.update_layout(**_chart_layout("Hazard Vulnerability Scores by Asset (0=none, 100=max)",max(220,len(df)*45+80)))
            _ax_style(fig_hm)
            st.plotly_chart(fig_hm,width="stretch")

            # Revenue loss stacked bar
            st.subheader("💸 Revenue Loss by Hazard")
            df["rl_flood"]  =df["revenue_loss"]*df["H_flood"]  /(df["D_total"].replace(0,1)*1/KAPPA["flood"])
            df["rl_heat"]   =df["revenue_loss"]*df["H_heat"]   /(df["D_total"].replace(0,1)*1/KAPPA["heat"])
            df["rl_cyclone"]=df["revenue_loss"]*df["H_cyclone"]/(df["D_total"].replace(0,1)*1/KAPPA["cyclone"])
            fig_rev=go.Figure()
            for haz,col,clr in [("rl_flood","Flood",C["accent3"]),("rl_heat","Heat",C["amber"]),("rl_cyclone","Cyclone",C["coral"])]:
                if haz in df.columns:
                    fig_rev.add_trace(go.Bar(x=df["asset_id"],y=df[haz].fillna(0),name=col,marker_color=clr))
            fig_rev.update_layout(**_chart_layout("Revenue Loss Decomposition by Hazard (₹ Cr)",300),barmode="stack")
            _ax_style(fig_rev); st.plotly_chart(fig_rev,width="stretch")

            # NGFS scenario fan charts (PHYS-01)
            if not df_phys_proj.empty:
                st.subheader("🌡️ NGFS Physical Risk Projections")
                st.caption("Revenue loss under each NGFS scenario · IPCC AR6 damage function · P10/P50/P90 uncertainty")
                tabs_scen = st.tabs(sorted(df_phys_proj["Scenario"].unique()))
                for tab_s,scen in zip(tabs_scen,sorted(df_phys_proj["Scenario"].unique())):
                    with tab_s:
                        ds=df_phys_proj[df_phys_proj["Scenario"]==scen].sort_values("Year")
                        fig_f=make_subplots(rows=2,cols=2,
                            subplot_titles=("Revenue Loss — P10/P50/P90","Physical PD","DSCR Under Physical Stress","Chronic vs Acute Split"),
                            vertical_spacing=0.18,horizontal_spacing=0.12)
                        # Fan
                        fig_f.add_trace(go.Scatter(
                            x=ds["Year"].tolist()+ds["Year"].tolist()[::-1],
                            y=ds["Revenue_Loss_P90_Cr"].tolist()+ds["Revenue_Loss_P10_Cr"].tolist()[::-1],
                            fill="toself",fillcolor="rgba(6,182,212,0.15)",line=dict(color="rgba(0,0,0,0)"),name="P10–P90"),row=1,col=1)
                        fig_f.add_trace(go.Scatter(x=ds["Year"],y=ds["Revenue_Loss_P50_Cr"],mode="lines+markers",
                            line=dict(color=C["accent2"],width=2.5),name="P50"),row=1,col=1)
                        fig_f.add_trace(go.Scatter(x=ds["Year"],y=ds["PD_Physical"],mode="lines+markers",
                            line=dict(color=C["coral"],width=2.5),name="Physical PD"),row=1,col=2)
                        fig_f.add_trace(go.Scatter(x=ds["Year"],y=ds["DSCR_Physical"],mode="lines+markers",
                            line=dict(color=C["amber"],width=2.5),name="DSCR"),row=2,col=1)
                        fig_f.add_hline(y=1.2,line_dash="dash",line_color=C["coral"],annotation_text="1.2× covenant",row=2,col=1)
                        fig_f.add_trace(go.Bar(x=ds["Year"],y=ds["Chronic_Loss_Cr"],name="Chronic",marker_color=C["accent3"]),row=2,col=2)
                        fig_f.add_trace(go.Bar(x=ds["Year"],y=ds["Acute_Loss_Cr"],name="Acute",marker_color=C["coral"]),row=2,col=2)
                        fig_f.update_layout(**_chart_layout(f"{scen} — Physical Risk Projections",480),barmode="stack")
                        _ax_style(fig_f,rows=2,cols=2); st.plotly_chart(fig_f,width="stretch")
                        wr=ds.loc[ds["PD_Physical"].idxmax()]
                        c1p,c2p,c3p,c4p=st.columns(4)
                        c1p.metric("Worst Year",f"{int(wr['Year'])}"); c2p.metric("Peak Rev Loss",f"₹{wr['Revenue_Loss_P50_Cr']:.1f} Cr")
                        c3p.metric("Peak Physical PD",f"{wr['PD_Physical']:.2%}"); c4p.metric("Temp Anomaly",f"+{wr['Delta_T_C']:.2f}°C")

            log_model_run("Physical",{"company":company_name,"pd_physical":PD_PHYS,"delta_ecl":DELTA_ECL})

# ============================================================
# TAB 3 — TARGETS
# ============================================================
with targets_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>🎯 Transition & BRSR Targets</h2>",unsafe_allow_html=True)
    st.caption("Management-defined decarbonisation targets — financial impact across transition risk and BRSR compliance trajectory")
    if not st.session_state.get("enable_targets",False):
        st.info("Enable **Transition Targets** in the sidebar.")
    elif not st.session_state.get("transition_ran",False):
        st.warning("Run Transition Risk Engine first.")
    else:
        df_base=st.session_state.get("df_transition")
        if not isinstance(df_base,pd.DataFrame) or df_base.empty:
            st.warning("Transition results unavailable.")
        else:
            # ── SECTION A: FINANCIAL TARGETS ──
            st.markdown(f"<h3 style='color:{C['accent2']};'>A · Financial Transition Targets</h3>",unsafe_allow_html=True)
            st.caption("Apply emissions reduction, revenue, and margin adjustments by time horizon.")
            tc1,tc2,tc3=st.columns(3)
            with tc1:
                st.markdown(f"<div style='color:{C['accent2']};font-weight:600;margin-bottom:8px;'>Short-Term (≤2030)</div>",unsafe_allow_html=True)
                s_em=st.slider("Emissions Reduction (%)",0,100,0,key="s_em")
                s_rev=st.slider("Revenue Adj (%)",-30,30,0,key="s_rev")
                s_mar=st.slider("Margin Adj (%)",-30,30,0,key="s_mar")
            with tc2:
                st.markdown(f"<div style='color:{C['amber']};font-weight:600;margin-bottom:8px;'>Medium-Term (2031–2040)</div>",unsafe_allow_html=True)
                m_em=st.slider("Emissions Reduction (%)",0,100,0,key="m_em")
                m_rev=st.slider("Revenue Adj (%)",-30,30,0,key="m_rev")
                m_mar=st.slider("Margin Adj (%)",-30,30,0,key="m_mar")
            with tc3:
                st.markdown(f"<div style='color:{C['coral']};font-weight:600;margin-bottom:8px;'>Long-Term (>2040)</div>",unsafe_allow_html=True)
                l_em=st.slider("Emissions Reduction (%)",0,100,0,key="l_em")
                l_rev=st.slider("Revenue Adj (%)",-30,30,0,key="l_rev")
                l_mar=st.slider("Margin Adj (%)",-30,30,0,key="l_mar")

            st.divider()

            # ── SECTION B: BRSR OPERATIONAL TARGETS ──
            st.markdown(f"<h3 style='color:{C['mint']};'>B · BRSR Operational Targets</h3>",unsafe_allow_html=True)
            st.caption("Set SEBI BRSR Core improvement commitments. These reduce BRSR PD uplift as targets are met, feeding into Integrated Risk.")

            brsr_ran_t = st.session_state.get("brsr_ran", False)
            current_brsr_pd = st.session_state.get("brsr_pd_adj", 0.0)
            current_readiness = 0.0
            current_flags = []
            if brsr_ran_t:
                bs_t = st.session_state.get("brsr_summary")
                if isinstance(bs_t, pd.DataFrame) and not bs_t.empty:
                    current_readiness = float(bs_t.iloc[0].get("Readiness_Score", 0))
                bf_t = st.session_state.get("brsr_flags")
                if isinstance(bf_t, pd.DataFrame) and "Flag" in bf_t.columns:
                    current_flags = bf_t["Flag"].tolist()

            if brsr_ran_t:
                st.markdown(f"**Current state:** BRSR PD Uplift = **+{current_brsr_pd*10000:.0f} bps** · Readiness = **{current_readiness:.0f}%** · Active flags: **{len(current_flags)}**")
            else:
                st.info("Run **BRSR Diagnostics** first to enable BRSR target planning.")

            bc1, bc2 = st.columns(2)
            with bc1:
                st.markdown(f"<div style='color:{C['mint']};font-weight:600;margin-bottom:8px;'>Near-Term BRSR Actions (by 2027)</div>",unsafe_allow_html=True)
                tgt_renewable   = st.slider("Target Renewable Energy Share (%)", 0, 100, min(100, int(st.session_state.get("brsr_summary",pd.DataFrame()).iloc[0].get("Renewable_%",15) if brsr_ran_t and not st.session_state.get("brsr_summary",pd.DataFrame()).empty else 15) + 10), key="tgt_ren")
                tgt_coverage    = st.slider("Target Emissions Coverage (%)", 0, 100, min(100, int(st.session_state.get("brsr_summary",pd.DataFrame()).iloc[0].get("Target_Coverage_%",50) if brsr_ran_t and not st.session_state.get("brsr_summary",pd.DataFrame()).empty else 50) + 15), key="tgt_cov")
                tgt_hazwaste    = st.slider("Target Hazardous Waste (%)", 0, 100, max(0, int(st.session_state.get("brsr_summary",pd.DataFrame()).iloc[0].get("Hazardous_Waste_%",20) if brsr_ran_t and not st.session_state.get("brsr_summary",pd.DataFrame()).empty else 20) - 5), key="tgt_hw")
            with bc2:
                st.markdown(f"<div style='color:{C['amber']};font-weight:600;margin-bottom:8px;'>Governance Commitments</div>",unsafe_allow_html=True)
                tgt_board       = st.checkbox("Commit Board Climate Oversight", value=False, key="tgt_board")
                tgt_netzero     = st.checkbox("Commit Net Zero / LT Target", value=False, key="tgt_nz")
                tgt_verified    = st.checkbox("Commit to Third-Party Verification", value=False, key="tgt_vd")
                tgt_scope3      = st.checkbox("Commit Scope 3 Disclosure", value=True, key="tgt_s3")

            # Compute BRSR target PD uplift reduction
            def _compute_brsr_target_pd(flags_in, tgt_ren, tgt_cov, tgt_hw, tgt_board, tgt_nz, tgt_vd, tgt_s3):
                remaining_flags = []
                for f in flags_in:
                    if f == "Low renewable energy (<20%)" and tgt_ren >= 20: continue
                    if f == "Low target coverage (<50%)"  and tgt_cov >= 50: continue
                    if f == "High hazardous waste (>30%)" and tgt_hw <= 30:  continue
                    if f == "No verified emissions data"  and tgt_vd:         continue
                    if f == "Missing Scope 3 disclosure"  and tgt_s3:         continue
                    remaining_flags.append(f)
                new_uplift = float(np.clip(sum(BRSR_PD_SPREAD.get(f,0) for f in remaining_flags), 0, 0.015))
                return new_uplift, remaining_flags

            brsr_target_uplift, brsr_remaining_flags = _compute_brsr_target_pd(
                current_flags, tgt_renewable, tgt_coverage, tgt_hazwaste,
                tgt_board, tgt_netzero, tgt_verified, tgt_scope3
            )
            brsr_pd_reduction = current_brsr_pd - brsr_target_uplift

            # Preview BRSR target impact
            if brsr_ran_t:
                tp1, tp2, tp3 = st.columns(3)
                tp1.metric("Current BRSR Uplift", f"+{current_brsr_pd*10000:.0f} bps")
                tp2.metric("Target BRSR Uplift",  f"+{brsr_target_uplift*10000:.0f} bps",
                           delta=f"-{brsr_pd_reduction*10000:.0f} bps", delta_color="normal")
                tp3.metric("Flags Resolved",
                           f"{len(current_flags) - len(brsr_remaining_flags)} / {len(current_flags)}")

            if st.button("▶ Run Target Scenario", type="primary"):
                def apply_t(row):
                    if row["Year"]<=2030: em,rv,mr=s_em,s_rev,s_mar
                    elif row["Year"]<=2040: em,rv,mr=m_em,m_rev,m_mar
                    else: em,rv,mr=l_em,l_rev,l_mar
                    rt=row["Revenue"]*(1+rv/100); mt=row["EBITDA_Margin"]*(1+mr/100)
                    pt=np.clip(row["PD_Transition"]*(1-em/100),0,1)
                    et=pt*row["LGD"]*exposure_at_default/1e3
                    return pd.Series({"Revenue_Target":rt,"EBITDA_Target":rt*mt,"PD_Target":pt,"ECL_Target":et})
                df_tgt=pd.concat([df_base[["Scenario","Year"]],df_base.apply(apply_t,axis=1)],axis=1).reset_index(drop=True)
                st.session_state["df_target"]=df_tgt; st.session_state["targets_ran"]=True

                # Store BRSR target results
                st.session_state["brsr_target_uplift"]   = brsr_target_uplift
                st.session_state["brsr_pd_reduction"]    = brsr_pd_reduction
                st.session_state["brsr_remaining_flags"] = brsr_remaining_flags

                eff=(df_tgt.groupby("Scenario").agg({"PD_Target":"max"}).rename(columns={"PD_Target":"PD_Target_Max"}).reset_index()
                     .merge(df_base.groupby("Scenario").agg({"PD_Transition":"max"}).rename(columns={"PD_Transition":"PD_Base_Max"}).reset_index(),on="Scenario"))
                eff["PD_Reduction_%"]=((eff["PD_Base_Max"]-eff["PD_Target_Max"])/eff["PD_Base_Max"]*100).round(2)
                st.session_state["df_target_effect"]=eff

                st.success("✅ Targets applied — BRSR target uplift saved to Integrated Risk tab")

                # ── Results display ──
                col_r1, col_r2 = st.columns([1, 1])
                with col_r1:
                    st.markdown(f"<h4 style='color:{C['accent2']};'>Financial Target Effectiveness</h4>",unsafe_allow_html=True)
                    st.dataframe(eff.style.format({"PD_Target_Max":"{:.3%}","PD_Base_Max":"{:.3%}","PD_Reduction_%":"{:.1f}%"})
                        .background_gradient(subset=["PD_Reduction_%"],cmap="Greens"),width="stretch",hide_index=True)

                with col_r2:
                    if brsr_ran_t and len(current_flags) > 0:
                        st.markdown(f"<h4 style='color:{C['mint']};'>BRSR PD Uplift — Before vs After Targets</h4>",unsafe_allow_html=True)
                        brsr_bar_labels = ["Current Uplift", "Target Uplift"]
                        brsr_bar_vals   = [current_brsr_pd*10000, brsr_target_uplift*10000]
                        fig_brsr_tgt = go.Figure(go.Bar(
                            x=brsr_bar_labels, y=brsr_bar_vals,
                            marker_color=[C["coral"], C["mint"]],
                            text=[f"{v:.0f} bps" for v in brsr_bar_vals], textposition="outside",
                        ))
                        fig_brsr_tgt.update_layout(**_chart_layout("BRSR PD Uplift Reduction (bps)", 260),
                            yaxis_title="Basis Points")
                        _ax_style(fig_brsr_tgt)
                        st.plotly_chart(fig_brsr_tgt, width="stretch")

                # Baseline vs target PD chart
                df_plot=df_base.merge(df_tgt,on=["Scenario","Year"],how="inner")
                fig_tgt=go.Figure()
                for scen in df_plot["Scenario"].unique():
                    ds=df_plot[df_plot["Scenario"]==scen]; clr=_scen_color(scen)
                    fig_tgt.add_trace(go.Scatter(x=ds["Year"],y=ds["PD_Transition"],mode="lines",
                        name=f"{scen[:20]} Baseline",line=dict(dash="dash",color=clr,width=1.5)))
                    fig_tgt.add_trace(go.Scatter(x=ds["Year"],y=ds["PD_Target"],mode="lines+markers",
                        name=f"{scen[:20]} + Financial Targets",line=dict(color=clr,width=2.5),marker=dict(size=6)))
                # Show integrated target line (transition PD target + BRSR target uplift)
                if brsr_ran_t:
                    for scen in df_tgt["Scenario"].unique():
                        ds = df_tgt[df_tgt["Scenario"]==scen]
                        pd_integ_target = np.clip(ds["PD_Target"] + brsr_target_uplift, PD_FLOOR, PD_CAP)
                        fig_tgt.add_trace(go.Scatter(x=ds["Year"], y=pd_integ_target, mode="lines",
                            name=f"{scen[:20]} + BRSR Target",
                            line=dict(color=C["mint"], width=1.5, dash="dot")))
                fig_tgt.update_layout(**_chart_layout("PD — Baseline vs Financial Target vs Integrated (+ BRSR) Target", 340))
                fig_tgt.update_yaxes(tickformat=".1%")
                _ax_style(fig_tgt)
                st.plotly_chart(fig_tgt, width="stretch")

                if brsr_ran_t:
                    st.caption(f"**BRSR note:** Target renewable share {tgt_renewable}%, target coverage {tgt_coverage}%, target hazardous waste {tgt_hazwaste}%. These resolve {len(current_flags)-len(brsr_remaining_flags)} of {len(current_flags)} current BRSR flags, reducing PD uplift from +{current_brsr_pd*10000:.0f} bps to +{brsr_target_uplift*10000:.0f} bps. The dotted 'BRSR Target' line includes this reduction.")
            else:
                st.info("Set targets above and click **Run Target Scenario**.")

# ============================================================
# TAB 4 — INTEGRATED RISK  (INT-01)
# ============================================================
with integrated_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>🧩 Integrated Climate Risk</h2>",unsafe_allow_html=True)
    st.caption("Gaussian copula joint PD · BRSR uplift layer · Capital assessment · ICAAP metrics")

    transition_ran_i = st.session_state.get("transition_ran",False)
    physical_ran_i   = st.session_state.get("physical_ran",False)
    physical_enabled = st.session_state.get("enable_physical",False)
    brsr_ran_i       = st.session_state.get("brsr_ran",False)

    if physical_enabled and physical_ran_i and transition_ran_i: exec_mode="combined"
    elif transition_ran_i: exec_mode="transition_only"
    elif physical_enabled and physical_ran_i: exec_mode="physical_only"
    else: exec_mode="none"

    if exec_mode == "none":
        st.warning("Run Transition Risk or Physical Risk first.")
    else:

        correlation = st.slider("Transition–Physical Correlation (ρ)",0.0,0.6,default_correlation,0.05,key="corr_sl") if exec_mode=="combined" else default_correlation

        # ── BRSR uplift — must be computed BEFORE the mode banner that references _brsr_label ──
        # Use target uplift if targets have been run, else use current uplift
        brsr_pd_adj  = 0.0
        _brsr_label  = "not run"        # safe default — overwritten below if BRSR ran
        if brsr_ran_i:
            targets_ran_i = st.session_state.get("targets_ran", False)
            if targets_ran_i and "brsr_target_uplift" in st.session_state:
                brsr_pd_adj = float(st.session_state["brsr_target_uplift"])
                _brsr_label = "after targets"
            else:
                brsr_pd_adj = float(st.session_state.get("brsr_pd_adj", 0.0))
                _brsr_label = "current"

        # Mode banner — all active layers now that _brsr_label is defined
        _layers = []
        if exec_mode == "combined":       _layers.append("🔗 Transition + Physical (Copula)")
        elif exec_mode == "transition_only": _layers.append("⚡ Transition Only")
        elif exec_mode == "physical_only":   _layers.append("🌍 Physical Only")
        if brsr_ran_i:
            _brsr_tag = f"📘 BRSR (+{brsr_pd_adj*10000:.0f}bps, {_brsr_label})"
            _layers.append(_brsr_tag)
        if st.session_state.get("targets_ran", False):
            _layers.append("🎯 Targets Applied")
        _mode_str = " &nbsp;+&nbsp; ".join(_layers)
        st.markdown(f"<div style='background:{C['card']};border-left:4px solid {C['accent2']};border-radius:6px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:{C['accent2']};font-weight:600;'>{_mode_str}</div>",unsafe_allow_html=True)

        # Collect results
        pd_trans=ecl_trans=dscr_trans=stranded_t=capex_t=None
        if transition_ran_i:
            df_ti=st.session_state.get("df_transition")
            if isinstance(df_ti,pd.DataFrame) and not df_ti.empty:
                pd_trans=float(df_ti["PD_Transition"].max()); ecl_trans=float(df_ti["ECL_Transition"].max())
                dscr_trans=float(df_ti["DSCR"].min()); stranded_t=float(df_ti["Stranded_Assets"].max()); capex_t=float(df_ti["CAPEX_Gap"].max())

        pd_phys=ecl_phys=dscr_phys=rev_loss_phys=None
        if exec_mode in ["physical_only","combined"]:
            ps=st.session_state.get("phys_summary")
            if isinstance(ps,dict):
                pd_phys=float(ps.get("Physical Risk PD",0)); ecl_phys=float(ps.get("ΔECL (₹ Cr)",0))
                dscr_phys=float(ps.get("Post-Risk DSCR",1.5)); rev_loss_phys=float(ps.get("Total Revenue Loss (₹ Cr)",0))

        # Scenario-consistent physical PD
        pd_phys_scen=pd_phys
        df_pp=st.session_state.get("df_physical_projection")
        if isinstance(df_pp,pd.DataFrame) and not df_pp.empty and transition_ran_i:
            df_ti2=st.session_state.get("df_transition")
            if isinstance(df_ti2,pd.DataFrame) and not df_ti2.empty:
                sel_scen=df_ti2["Scenario"].unique()[0]
                rows=df_pp[(df_pp["Scenario"]==sel_scen)&(df_pp["Year"]==REPORTING_YEAR)]
                if not rows.empty: pd_phys_scen=float(rows["PD_Physical"].iloc[0])
        if pd_phys_scen is not None: pd_phys_scen=float(np.clip(pd_phys_scen,0,1))

        # Gaussian copula integration
        pd_copula=None
        if exec_mode=="combined" and pd_trans is not None and pd_phys_scen is not None:
            pd_copula=gaussian_copula_pd(pd_trans,pd_phys_scen,correlation)
        elif transition_ran_i and pd_trans is not None:
            pd_copula=pd_trans
        elif pd_phys_scen is not None:
            pd_copula=pd_phys_scen

        # Add BRSR uplift (additive, capped at PD_CAP)
        pd_integrated=float(np.clip(pd_copula+brsr_pd_adj,PD_FLOOR,PD_CAP)) if pd_copula is not None else None
        ecl_integrated=(ecl_trans or 0)+(ecl_phys or 0)
        dscr_integrated=min(d for d in [dscr_trans,dscr_phys] if d is not None) if any(d is not None for d in [dscr_trans,dscr_phys]) else None

        if pd_integrated is None:
            st.warning("Integrated PD could not be computed. Check module results.")

        EAD=float(exposure_at_default)
        ecl_ead_ratio=ecl_integrated/EAD if EAD>0 else 0

        # Risk score (0-100)
        pd_score  =min(100,(pd_integrated/0.20)*100)
        ecl_score =min(100,(ecl_ead_ratio/0.05)*100)
        dscr_score=100 if (dscr_integrated and dscr_integrated<1.0) else 60 if (dscr_integrated and dscr_integrated<1.2) else 30 if (dscr_integrated and dscr_integrated<1.5) else 10
        risk_score_int=round(0.4*pd_score+0.4*ecl_score+0.2*dscr_score,1)

        if risk_score_int<30: rag="🟢"; rag_color=C["mint"]; rag_text="Low Climate Risk"
        elif risk_score_int<60: rag="🟡"; rag_color=C["amber"]; rag_text="Moderate Climate Risk"
        elif risk_score_int<80: rag="🟠"; rag_color="#F97316"; rag_text="Severe Climate Risk"
        else: rag="🔴"; rag_color=C["coral"]; rag_text="Extreme Climate Risk"

        # RAG banner
        st.markdown(f"""
        <div style="background:linear-gradient(90deg,{C['card']},{C['bg_mid']});border:1px solid {rag_color};border-radius:10px;
                    padding:16px 20px;margin-bottom:16px;display:flex;align-items:center;gap:16px;">
          <div style="font-size:36px;">{rag}</div>
          <div>
            <div style="font-size:18px;font-weight:700;color:{rag_color};">{rag_text}</div>
            <div style="font-size:13px;color:{C['slate']};margin-top:3px;">
              Integrated Risk Score: {risk_score_int}/100 · PD: {pd_integrated:.2%} · ECL/EAD: {ecl_ead_ratio:.2%}
              {f' · BRSR uplift: +{brsr_pd_adj*10000:.0f}bps' if brsr_ran_i and brsr_pd_adj>0 else ''}
            </div>
          </div>
        </div>""",unsafe_allow_html=True)

        # KPI grid
        ik1,ik2,ik3,ik4,ik5=st.columns(5)
        ik1.metric("Integrated PD",f"{pd_integrated:.2%}",help="Gaussian copula + BRSR uplift")
        ik2.metric("Integrated ECL",f"₹{ecl_integrated:,.1f} Cr")
        ik3.metric("ECL / EAD",f"{ecl_ead_ratio:.2%}")
        ik4.metric("Min DSCR",f"{dscr_integrated:.2f}×" if dscr_integrated else "—")
        ik5.metric("Risk Score",f"{risk_score_int}/100")

        if exec_mode=="combined":
            st.markdown(f"<div style='color:{C['slate']};font-size:12px;margin-top:4px;'>Copula PD: {pd_copula:.2%} → +BRSR {brsr_pd_adj*10000:.0f}bps ({_brsr_label}) → Total: {pd_integrated:.2%}</div>",unsafe_allow_html=True)
        elif brsr_ran_i and brsr_pd_adj>0:
            st.markdown(f"<div style='color:{C['slate']};font-size:12px;margin-top:4px;'>Transition PD: {pd_trans:.2%} → +BRSR {brsr_pd_adj*10000:.0f}bps ({_brsr_label}) → Total: {pd_integrated:.2%}</div>",unsafe_allow_html=True)

        st.divider()

        # ── RISK DECOMPOSITION ──
        st.subheader("🧩 Risk Decomposition")
        dec_col1, dec_col2 = st.columns([1,2])
        with dec_col1:
            sources,values=[],[]
            if transition_ran_i and ecl_trans: sources.append("Transition"); values.append(ecl_trans)
            if exec_mode in ["physical_only","combined"] and ecl_phys: sources.append("Physical"); values.append(ecl_phys)
            if brsr_ran_i and brsr_pd_adj>0:
                brsr_ecl_equiv=brsr_pd_adj*LGD_0*EAD/1e3; sources.append("BRSR Operational"); values.append(brsr_ecl_equiv)
            if sources:
                fig_dec=go.Figure(go.Pie(labels=sources,values=values,
                    hole=0.45,marker=dict(colors=[C["accent2"],C["coral"],C["purple"]][:len(sources)]),
                    textinfo="label+percent",hovertemplate="%{label}: ₹%{value:.1f} Cr<extra></extra>"))
                fig_dec.update_layout(**_chart_layout("ECL Risk Decomposition",280),showlegend=False,
                    annotations=[dict(text=f"₹{ecl_integrated:.1f}Cr",x=0.5,y=0.5,font_size=14,font_color=C["white"],showarrow=False)])
                st.plotly_chart(fig_dec,width="stretch")
        with dec_col2:
            if transition_ran_i:
                df_ti3=st.session_state.get("df_transition")
                if isinstance(df_ti3,pd.DataFrame) and not df_ti3.empty:
                    fig_sc=go.Figure()
                    for scen in df_ti3["Scenario"].unique():
                        ds=df_ti3[df_ti3["Scenario"]==scen]
                        fig_sc.add_trace(go.Bar(x=["PD","ECL/EAD","DSCR Stress"],
                            y=[ds["PD_Transition"].max()*100,ds["ECL_Transition"].max()/EAD*100,max(0,(1.5-ds["DSCR"].min())/1.5*100)],
                            name=scen[:25],marker_color=_scen_color(scen)))
                    fig_sc.update_layout(**_chart_layout("Risk Metrics by Scenario (normalised %)",280),barmode="group",yaxis_title="%")
                    _ax_style(fig_sc); st.plotly_chart(fig_sc,width="stretch")

        # ── GAUSSIAN COPULA METHODOLOGY TRANSPARENCY ──
        if exec_mode=="combined":
            with st.expander("🔢 Copula Integration: Full Working",expanded=False):
                st.markdown(f"""
                **Step 1 — Marginal PDs from each engine:**
                - Transition PD: `{pd_trans:.4f}` (logit PD model, structural)
                - Physical PD: `{pd_phys_scen:.4f}` (NGFS scenario-projected, IPCC damage function)
                - Sector correlation ρ: `{correlation}`

                **Step 2 — Transform to standard normal quantiles:**
                - z_T = Φ⁻¹({pd_trans:.4f}) = `{norm.ppf(max(pd_trans,1e-6)):.4f}`
                - z_P = Φ⁻¹({pd_phys_scen:.4f}) = `{norm.ppf(max(pd_phys_scen,1e-6)):.4f}`

                **Step 3 — Bivariate normal CDF:**
                - `PD_copula = Φ₂(z_T, z_P, ρ={correlation}) = {pd_copula:.4f}`

                **Step 4 — BRSR uplift (additive):**
                - `PD_integrated = {pd_copula:.4f} + {brsr_pd_adj:.4f} (BRSR uplift) = {pd_integrated:.4f}`

                **Why copula beats simple addition:** `PD_T + PD_P = {(pd_trans or 0)+(pd_phys_scen or 0):.4f}` vs copula `{pd_copula:.4f}` — copula captures correlation structure, prevents double-counting independent defaults.
                """)

        # ── SCENARIO STRESS TABLE ──
        st.subheader("📊 Full Scenario Stress Matrix")
        if transition_ran_i:
            df_ti4=st.session_state.get("df_transition")
            if isinstance(df_ti4,pd.DataFrame) and not df_ti4.empty:
                df_stress=(df_ti4.groupby("Scenario").agg({
                    "PD_Transition":"max","ECL_Transition":"max","DSCR":"min",
                    "Carbon_Burden":"max","EBITDA_Margin":"min",
                    "Stranded_Assets":"max","CAPEX_Gap":"max",
                }).round(4).reset_index().rename(columns={
                    "PD_Transition":"Peak PD","ECL_Transition":"Peak ECL (₹Cr)",
                    "DSCR":"Min DSCR","Carbon_Burden":"Max Carbon Burden",
                    "EBITDA_Margin":"Min EBITDA Margin","Stranded_Assets":"Stranded (₹Cr)","CAPEX_Gap":"CAPEX Gap (₹Cr)"
                }))
                st.dataframe(df_stress.style
                    .format({"Peak PD":"{:.3%}","Peak ECL (₹Cr)":"{:.1f}","Min DSCR":"{:.2f}",
                             "Max Carbon Burden":"{:.2%}","Min EBITDA Margin":"{:.2%}",
                             "Stranded (₹Cr)":"{:.0f}","CAPEX Gap (₹Cr)":"{:.0f}"})
                    .background_gradient(subset=["Peak PD"],cmap="Reds")
                    .background_gradient(subset=["Min DSCR"],cmap="RdYlGn"),
                    width="stretch",hide_index=True)

        # ── ICAAP SUMMARY ──
        st.subheader("🏦 ICAAP Capital Assessment")
        capital_signal="Severe" if ecl_ead_ratio>=0.05 else "Elevated" if ecl_ead_ratio>=0.03 else "Moderate" if ecl_ead_ratio>=0.015 else "Limited"
        cap_color=C["coral"] if capital_signal in ["Severe","Elevated"] else C["amber"] if capital_signal=="Moderate" else C["mint"]
        cap1,cap2,cap3,cap4=st.columns(4)
        cap1.metric("Capital Signal",capital_signal)
        cap2.metric("Unexpected Climate Loss",f"₹{max(0,ecl_integrated-ecl_integrated*0.7):.1f} Cr",help="95th pct - mean approximation")
        cap3.metric("Climate Capital (% EAD)",f"{ecl_ead_ratio:.2%}")
        cap4.metric("ISSB S2 Reference","§15(b), §16")
        st.markdown(f"<div style='background:{C['card']};border-left:3px solid {cap_color};padding:10px 14px;border-radius:6px;font-size:13px;color:{cap_color};font-weight:600;margin-top:8px;'>Capital Stress: {capital_signal} Climate Capital Impact — {'Strategic intervention required.' if capital_signal in ['Severe','Elevated'] else 'Monitoring and capital buffer planning recommended.' if capital_signal=='Moderate' else 'Within normal tolerance.'}</div>",unsafe_allow_html=True)

        # ── ISSB S2 SUMMARY TABLE ──
        st.subheader("📋 ISSB S2–Aligned Integrated Summary")
        df_isummary=pd.DataFrame([
            {"Metric":"Integrated PD (Copula + BRSR)","Value":f"{pd_integrated:.4f}","ISSB S2":"§15(a)"},
            {"Metric":"Integrated ECL (₹ Cr)","Value":f"{ecl_integrated:.2f}","ISSB S2":"§15(b)"},
            {"Metric":"Post-Stress DSCR","Value":f"{dscr_integrated:.2f}" if dscr_integrated else "—","ISSB S2":"§15(c)"},
            {"Metric":"ECL / EAD","Value":f"{ecl_ead_ratio:.4f}","ISSB S2":"§16"},
            {"Metric":"BRSR PD Uplift (bps)","Value":f"{brsr_pd_adj*10000:.1f}","ISSB S2":"§15(a) operational"},
            {"Metric":"Capital Stress Signal","Value":capital_signal,"ISSB S2":"§16"},
            {"Metric":"Integrated Risk Score (0–100)","Value":f"{risk_score_int}","ISSB S2":"§14–16"},
            {"Metric":"Stranded Assets (₹ Cr)","Value":f"{stranded_t:.0f}" if stranded_t else "—","ISSB S2":"§22"},
            {"Metric":"CAPEX Gap (₹ Cr)","Value":f"{capex_t:.0f}" if capex_t else "—","ISSB S2":"§14"},
        ])
        st.dataframe(df_isummary.style.set_properties(**{"background-color":C["bg_dark"],"color":C["text"]}),
            width="stretch",hide_index=True)

        st.session_state["df_integrated_summary"]=df_isummary
        st.session_state["integrated_ran"]=True
        st.session_state["model_inputs"]={"company":company_name,"sector":sector,"reporting_year":REPORTING_YEAR,
            "revenue":revenue_0,"ebitda_margin":ebitda_margin_0,"interest_expense":interest_payment,
            "total_emissions":TOTAL_EMISSIONS,"exposure_at_default":EAD,"base_pd":base_pd,"lgd":LGD_0}

        # ── MONTE CARLO ──
        st.divider()
        st.subheader("🎲 Monte Carlo Climate Stress")
        run_mc=st.checkbox("Run Monte Carlo Simulation")
        if run_mc:
            n_sims=st.slider("Simulations",100,5000,1000,100)
            df_tmc=st.session_state.get("df_transition")
            if isinstance(df_tmc,pd.DataFrame) and not df_tmc.empty:
                wr=df_tmc.loc[df_tmc["PD_Transition"].idxmax()]
                cb_mc=wr["Carbon_Burden"]; pl_mc=max(0,1-wr["EBITDA_Margin"]/max(ebitda_margin_0,1e-6)); cp_mc=cb_mc*revenue_0/max(TOTAL_EMISSIONS,1)
            else: cb_mc=0.05; pl_mc=0.02; cp_mc=50.0

            pd_sim=[]; ecl_sim=[]
            for _ in range(n_sims):
                sh_c,sh_p,sh_g=np.random.multivariate_normal([0,0,0],CLIMATE_DRIVER_COV)
                cps=cp_mc*np.exp(MC_CARBON_VOL*sh_c); pds=pl_mc*np.exp(sh_p)
                cc=(TOTAL_EMISSIONS*cps*USD_INR/1e7)*(1-carbon_pass_through)
                rv=revenue_0*(1+gdp_sensitivity*sh_g)*(1-pds)
                eb=rv*ebitda_margin_0-cc; cbu=cc/max(rv,1)
                cs=BASE_CREDIT_SPREAD*(1+spread_beta*cbu)
                ds=eb/max(interest_payment*(1+cs),1e-6)
                dg=np.clip(1.5-ds,-4.0,6.0)
                lp=logit(base_pd)+alpha_dscr*dg+beta_carbon_credit*cbu
                pt=np.clip(sigmoid(lp),PD_FLOOR,PD_CAP)
                pp=np.clip(pd_phys_scen*(1+pds),0,1) if physical_ran_i and pd_phys_scen else 0
                pj=gaussian_copula_pd(pt,pp,correlation) if physical_ran_i else pt
                pj=float(np.clip(pj+brsr_pd_adj,PD_FLOOR,PD_CAP))
                lj=np.clip(LGD_0*(1+0.2*cbu+LGD_PHYSICAL_MULTIPLIER*pds),0,1)
                pd_sim.append(pj); ecl_sim.append(pj*lj*EAD/1e3)

            pd_sim=np.array(pd_sim); ecl_sim=np.array(ecl_sim)
            pd_mean=pd_sim.mean(); pd_95=np.percentile(pd_sim,95)
            ecl_mean=ecl_sim.mean(); ecl_95=np.percentile(ecl_sim,95)
            ecl_es=ecl_sim[ecl_sim>=ecl_95].mean(); ul=ecl_95-ecl_mean

            mc1,mc2,mc3,mc4=st.columns(4)
            mc1.metric("Mean PD",f"{pd_mean:.2%}"); mc2.metric("PD 95th",f"{pd_95:.2%}")
            mc3.metric("Climate VaR (₹Cr)",f"{ecl_95:,.2f}"); mc4.metric("Unexpected Loss (₹Cr)",f"{ul:,.2f}")

            fig_mc=make_subplots(rows=1,cols=2,subplot_titles=("ECL Distribution","PD vs ECL Scatter"))
            fig_mc.add_trace(go.Histogram(x=ecl_sim,nbinsx=40,marker_color=C["accent2"],opacity=0.75,name="ECL"),row=1,col=1)
            fig_mc.add_vline(x=ecl_95,line_dash="dash",line_color=C["coral"],annotation_text=f"VaR 95%: {ecl_95:.1f}",row=1,col=1)
            fig_mc.add_trace(go.Scatter(x=pd_sim,y=ecl_sim,mode="markers",marker=dict(color=C["accent3"],size=3,opacity=0.3),name="Simulation"),row=1,col=2)
            fig_mc.update_layout(**_chart_layout("Monte Carlo Results",320)); _ax_style(fig_mc,rows=1,cols=2)
            st.plotly_chart(fig_mc,width="stretch")
            st.session_state["mc_results"]={"Mean_PD":pd_mean,"PD_95":pd_95,"Mean_ECL":ecl_mean,"ECL_95":ecl_95,"Climate_Capital_VaR":ecl_95}
            log_model_run("MonteCarlo",{"company":company_name,"pd_mean":pd_mean,"pd_95":pd_95,"ecl_95":ecl_95})

        # ── REVERSE STRESS TEST ──
        st.divider()
        st.subheader("🔥 Reverse Stress Test")
        run_rev=st.checkbox("Run Reverse Stress Test")
        if run_rev:
            tgt_pd=st.slider("Target PD Threshold",0.05,0.40,0.15)
            breach=None
            for cp in np.linspace(50,500,50):
                cc=TOTAL_EMISSIONS*cp*USD_INR/1e7; cb=cc/revenue_0
                lp=logit(base_pd)+beta_carbon_credit*cb+alpha_dscr*np.clip(1.5-(dscr_trans or 1.5),-4,6)
                if sigmoid(lp)>=tgt_pd: breach=cp; break
            if breach: st.error(f"⚠️ PD exceeds {tgt_pd:.2%} when carbon price reaches **${breach:.0f}/tCO₂**")
            else: st.success("No breach in tested carbon price range (up to $500/tCO₂)")

        log_model_run("Integrated",{"company":company_name,"sector":sector,"pd_integrated":pd_integrated,
            "ecl_integrated":ecl_integrated,"risk_score":risk_score_int,"capital_signal":capital_signal})

    # ============================================================
    # TAB 6 — PLOTS  (comprehensive, all modules)
    # ============================================================
with plots_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>📈 Visual Analytics Centre</h2>",unsafe_allow_html=True)
    st.caption("All scenario charts, risk analytics, BRSR diagnostics, physical risk maps, and target comparisons in one place.")

    t_ran  = st.session_state.get("transition_ran", False)
    p_ran  = st.session_state.get("physical_ran",   False)
    b_ran  = st.session_state.get("brsr_ran",       False)
    tg_ran = st.session_state.get("targets_ran",    False)

    if not (t_ran or p_ran or b_ran):
        st.info("Run at least one engine to see charts.")
    else:

        # Scenario filter (only if transition ran)
        scen_opts = []
        if t_ran:
            df_tp = st.session_state.get("df_transition")
            if isinstance(df_tp, pd.DataFrame) and not df_tp.empty:
                scen_opts = sorted(df_tp["Scenario"].unique())
        sel_scens = st.multiselect("Filter Scenarios", options=scen_opts, default=scen_opts, key="plots_scen")

        # ── SECTION 1: TRANSITION RISK ──────────────────────────────────
        if t_ran and scen_opts:
            df_tp = st.session_state.get("df_transition")
            if isinstance(df_tp, pd.DataFrame) and not df_tp.empty:
                df_tp_f = df_tp[df_tp["Scenario"].isin(sel_scens)] if sel_scens else df_tp

                st.markdown(f"<h3 style='color:{C['accent2']};margin-top:8px;'>⚡ Transition Risk Analytics</h3>", unsafe_allow_html=True)

                # Row 1: PD + ECL
                col1, col2 = st.columns(2)
                with col1:
                    fig = go.Figure()
                    for scen in df_tp_f["Scenario"].unique():
                        ds = df_tp_f[df_tp_f["Scenario"]==scen]
                        fig.add_trace(go.Scatter(x=ds["Year"], y=ds["PD_Transition"],
                            mode="lines+markers", name=scen[:28],
                            line=dict(color=_scen_color(scen), width=2.5), marker=dict(size=7),
                            hovertemplate="Year: %{x}<br>PD: %{y:.2%}<extra>" + scen[:18] + "</extra>"))
                    fig.update_layout(**_chart_layout("Probability of Default — All Scenarios", 300))
                    fig.update_yaxes(tickformat=".1%")
                    _ax_style(fig)
                    st.plotly_chart(fig, width="stretch")

                with col2:
                    fig = go.Figure()
                    for scen in df_tp_f["Scenario"].unique():
                        ds = df_tp_f[df_tp_f["Scenario"]==scen]
                        fig.add_trace(go.Bar(x=ds["Year"], y=ds["ECL_Transition"],
                            name=scen[:28], marker_color=_scen_color(scen), opacity=0.85,
                            hovertemplate="Year: %{x}<br>ECL: ₹%{y:.1f} Cr<extra>" + scen[:18] + "</extra>"))
                    fig.update_layout(**_chart_layout("Expected Credit Loss (₹ Cr)", 300), barmode="group")
                    _ax_style(fig)
                    st.plotly_chart(fig, width="stretch")

                # Row 2: DSCR + Carbon Burden
                col3, col4 = st.columns(2)
                with col3:
                    fig = go.Figure()
                    for scen in df_tp_f["Scenario"].unique():
                        ds = df_tp_f[df_tp_f["Scenario"]==scen]
                        fig.add_trace(go.Scatter(x=ds["Year"], y=ds["DSCR"],
                            mode="lines+markers", name=scen[:28],
                            line=dict(color=_scen_color(scen), width=2), marker=dict(size=6)))
                    fig.add_hline(y=1.2, line_dash="dash", line_color=C["coral"],
                        annotation_text="1.2× covenant", annotation_font_color=C["coral"])
                    fig.add_hline(y=1.5, line_dash="dot", line_color=C["amber"],
                        annotation_text="1.5× threshold", annotation_font_color=C["amber"])
                    fig.update_layout(**_chart_layout("DSCR Stress Trajectory", 300))
                    _ax_style(fig)
                    st.plotly_chart(fig, width="stretch")

                with col4:
                    fig = go.Figure()
                    for scen in df_tp_f["Scenario"].unique():
                        ds = df_tp_f[df_tp_f["Scenario"]==scen]
                        fig.add_trace(go.Scatter(x=ds["Year"], y=ds["Carbon_Burden"]*100,
                            mode="lines+markers", name=scen[:28],
                            line=dict(color=_scen_color(scen), width=2), marker=dict(size=6),
                            fill="tozeroy", fillcolor=_hex_rgba(_scen_color(scen), 0.13)))
                    fig.update_layout(**_chart_layout("Carbon Burden (% of Revenue)", 300))
                    fig.update_yaxes(ticksuffix="%")
                    _ax_style(fig)
                    st.plotly_chart(fig, width="stretch")

                # Row 3: EBITDA Margin + Stranded Assets + CAPEX Gap (3-panel subplot)
                fig3 = make_subplots(rows=1, cols=3,
                    subplot_titles=("EBITDA Margin (%)", "Stranded Assets (₹ Cr)", "CAPEX Gap (₹ Cr)"))
                for scen in df_tp_f["Scenario"].unique():
                    ds = df_tp_f[df_tp_f["Scenario"]==scen]; clr = _scen_color(scen)
                    fig3.add_trace(go.Scatter(x=ds["Year"], y=ds["EBITDA_Margin"]*100, mode="lines+markers",
                        name=scen[:20], line=dict(color=clr, width=2), marker=dict(size=5),
                        showlegend=True), row=1, col=1)
                    fig3.add_trace(go.Scatter(x=ds["Year"], y=ds["Stranded_Assets"], mode="lines+markers",
                        name=scen[:20], line=dict(color=clr, width=2), marker=dict(size=5),
                        showlegend=False), row=1, col=2)
                    fig3.add_trace(go.Scatter(x=ds["Year"], y=ds["CAPEX_Gap"], mode="lines+markers",
                        name=scen[:20], line=dict(color=clr, width=2), marker=dict(size=5),
                        showlegend=False), row=1, col=3)
                fig3.update_layout(**_chart_layout("Financial Stress Indicators by Scenario", 280,
                    legend_override=dict(orientation="h", y=1.12)))
                _ax_style(fig3, rows=1, cols=3)
                fig3.update_yaxes(ticksuffix="%", row=1, col=1)
                st.plotly_chart(fig3, width="stretch")

                # Row 4: Carbon burden vs EBITDA scatter
                fig_sc = go.Figure()
                for scen in df_tp_f["Scenario"].unique():
                    ds = df_tp_f[df_tp_f["Scenario"]==scen]
                    fig_sc.add_trace(go.Scatter(x=ds["Carbon_Burden"]*100, y=ds["EBITDA_Margin"]*100,
                        mode="markers+lines", name=scen[:28],
                        marker=dict(color=_scen_color(scen), size=9,
                            line=dict(color=C["white"], width=0.5)),
                        line=dict(color=_scen_color(scen), width=1, dash="dot"),
                        text=ds["Year"].astype(str), textposition="top center",
                        hovertemplate="Carbon Burden: %{x:.1f}%<br>EBITDA Margin: %{y:.1f}%<extra>" + scen[:18] + "</extra>"))
                fig_sc.update_layout(**_chart_layout("Carbon Burden vs EBITDA Margin — Stress Path", 300))
                fig_sc.update_xaxes(title="Carbon Burden (%)"); fig_sc.update_yaxes(title="EBITDA Margin (%)")
                _ax_style(fig_sc)
                st.plotly_chart(fig_sc, width="stretch")

        # ── SECTION 2: TARGETS COMPARISON ──────────────────────────────
        if tg_ran:
            st.markdown(f"<h3 style='color:{C['amber']};margin-top:8px;'>🎯 Targets — Baseline vs Achievement</h3>", unsafe_allow_html=True)
            df_b2  = st.session_state.get("df_transition")
            df_tg2 = st.session_state.get("df_target")
            if isinstance(df_b2, pd.DataFrame) and isinstance(df_tg2, pd.DataFrame):
                df_pl2 = df_b2.merge(df_tg2, on=["Scenario","Year"], how="inner")
                df_pl2 = df_pl2[df_pl2["Scenario"].isin(sel_scens)] if sel_scens else df_pl2

                tc1, tc2 = st.columns(2)
                with tc1:
                    fig_tp = go.Figure()
                    for scen in df_pl2["Scenario"].unique():
                        ds = df_pl2[df_pl2["Scenario"]==scen]; clr = _scen_color(scen)
                        fig_tp.add_trace(go.Scatter(x=ds["Year"], y=ds["PD_Transition"], mode="lines",
                            name=f"{scen[:18]} Base", line=dict(dash="dash", color=clr, width=1.5)))
                        fig_tp.add_trace(go.Scatter(x=ds["Year"], y=ds["PD_Target"], mode="lines+markers",
                            name=f"{scen[:18]} Target", line=dict(color=clr, width=2.5), marker=dict(size=6)))
                    fig_tp.update_layout(**_chart_layout("PD — Baseline vs Target", 300))
                    fig_tp.update_yaxes(tickformat=".1%")
                    _ax_style(fig_tp)
                    st.plotly_chart(fig_tp, width="stretch")

                with tc2:
                    fig_te = go.Figure()
                    for scen in df_pl2["Scenario"].unique():
                        ds = df_pl2[df_pl2["Scenario"]==scen]; clr = _scen_color(scen)
                        fig_te.add_trace(go.Scatter(x=ds["Year"], y=ds["ECL_Transition"], mode="lines",
                            name=f"{scen[:18]} Base", line=dict(dash="dash", color=clr, width=1.5)))
                        fig_te.add_trace(go.Scatter(x=ds["Year"], y=ds["ECL_Target"], mode="lines+markers",
                            name=f"{scen[:18]} Target", line=dict(color=clr, width=2.5), marker=dict(size=6)))
                    fig_te.update_layout(**_chart_layout("ECL — Baseline vs Target (₹ Cr)", 300))
                    _ax_style(fig_te)
                    st.plotly_chart(fig_te, width="stretch")

                # BRSR target comparison if available
                if b_ran and "brsr_target_uplift" in st.session_state:
                    curr_up   = st.session_state.get("brsr_pd_adj", 0) * 10000
                    tgt_up    = st.session_state["brsr_target_uplift"] * 10000
                    reduction = curr_up - tgt_up
                    fig_bt = go.Figure()
                    fig_bt.add_trace(go.Bar(name="Current BRSR Uplift", x=["BRSR PD Uplift"], y=[curr_up],
                        marker_color=C["coral"], text=[f"{curr_up:.0f} bps"], textposition="outside"))
                    fig_bt.add_trace(go.Bar(name="Post-Target Uplift", x=["BRSR PD Uplift"], y=[tgt_up],
                        marker_color=C["mint"], text=[f"{tgt_up:.0f} bps"], textposition="outside"))
                    fig_bt.update_layout(**_chart_layout(f"BRSR PD Uplift: {curr_up:.0f} → {tgt_up:.0f} bps (−{reduction:.0f} bps reduction)", 250),
                        barmode="group", yaxis_title="Basis Points")
                    _ax_style(fig_bt)
                    st.plotly_chart(fig_bt, width="stretch")

        # ── SECTION 3: BRSR ANALYTICS ─────────────────────────────────
        if b_ran:
            st.markdown(f"<h3 style='color:{C['mint']};margin-top:8px;'>📘 BRSR Operational Risk Analytics</h3>", unsafe_allow_html=True)
            bs = st.session_state.get("brsr_summary")
            bf = st.session_state.get("brsr_flags")

            if isinstance(bs, pd.DataFrame) and not bs.empty:
                row_b = bs.iloc[0]
                bc1, bc2 = st.columns(2)

                # GHG intensity gauge
                with bc1:
                    ghg_val = float(row_b.get("GHG_Intensity", 0))
                    bench   = SECTOR_GHG_BENCHMARKS.get(sector, SECTOR_GHG_BENCHMARKS["Manufacturing"])
                    fig_gb = go.Figure(go.Bar(
                        x=["Your GHG", "Sector P25", "Sector P50", "Sector P75"],
                        y=[ghg_val, bench["p25"], bench["p50"], bench["p75"]],
                        marker_color=[
                            C["mint"] if ghg_val <= bench["p50"] else C["amber"] if ghg_val <= bench["p75"] else C["coral"],
                            C["mint"], C["amber"], C["coral"]
                        ],
                        text=[f"{v:.2f}" for v in [ghg_val, bench["p25"], bench["p50"], bench["p75"]]],
                        textposition="outside",
                    ))
                    fig_gb.update_layout(**_chart_layout("GHG Intensity vs Sector Benchmarks (tCO₂/₹Cr)", 280))
                    fig_gb.update_yaxes(title="tCO₂e / ₹Cr")
                    _ax_style(fig_gb)
                    st.plotly_chart(fig_gb, width="stretch")

                # BRSR PD uplift breakdown
                with bc2:
                    if isinstance(bf, pd.DataFrame) and "Flag" in bf.columns and len(bf) > 0:
                        flags_list = bf["Flag"].tolist()
                        flag_bps = {f: BRSR_PD_SPREAD.get(f, 0)*10000 for f in flags_list}
                        fig_fb = go.Figure(go.Bar(
                            x=list(flag_bps.values()), y=list(flag_bps.keys()),
                            orientation="h",
                            marker_color=[C["coral"] if v>=25 else C["amber"] if v>=15 else C["mint"] for v in flag_bps.values()],
                            text=[f"+{v:.0f} bps" for v in flag_bps.values()], textposition="outside",
                        ))
                        fig_fb.update_layout(
                            **_chart_layout(f"BRSR Flags → PD Uplift (Total: +{sum(flag_bps.values()):.0f} bps)", 280,
                                margin_override=dict(l=230, r=80, t=50, b=20)),
                        )
                        _ax_style(fig_fb)
                        st.plotly_chart(fig_fb, width="stretch")
                    else:
                        st.success("✅ No BRSR flags — zero PD uplift from operational climate risk")

                # Regulatory readiness radar
                rdness_cols = {"Renewable ≥20%": float(row_b.get("Renewable_%",0))>=20,
                               "Coverage ≥50%":  float(row_b.get("Target_Coverage_%",0))>=50,
                               "Haz Waste ≤30%": float(row_b.get("Hazardous_Waste_%",100))<=30,
                               "Water Risk OK":   row_b.get("Water_Stress_Level","High")!="High"}
                rdness_scores = [100 if v else 0 for v in rdness_cols.values()]
                rcats = list(rdness_cols.keys())
                fig_rad = go.Figure()
                fig_rad.add_trace(go.Scatterpolar(
                    r=rdness_scores + [rdness_scores[0]], theta=rcats + [rcats[0]],
                    fill="toself", fillcolor="rgba(6,182,212,0.18)",
                    line=dict(color=C["accent2"], width=2), name="Current"))
                fig_rad.add_trace(go.Scatterpolar(
                    r=[100]*(len(rcats)+1), theta=rcats+[rcats[0]],
                    line=dict(color=C["mint"], width=1, dash="dot"), name="Target"))
                fig_rad.update_layout(
                    polar=dict(radialaxis=dict(visible=True, range=[0,100]), bgcolor=C["bg_dark"]),
                    **_chart_layout("SEBI BRSR Core Readiness", 320),
                    showlegend=True,
                )
                st.plotly_chart(fig_rad, width="stretch")

        # ── SECTION 4: PHYSICAL RISK ───────────────────────────────────
        if p_ran:
            st.markdown(f"<h3 style='color:{C['coral']};margin-top:8px;'>🌍 Physical Risk Analytics</h3>", unsafe_allow_html=True)
            phys = st.session_state.get("phys_assets")
            df_pp = st.session_state.get("df_physical_projection")

            if isinstance(phys, pd.DataFrame) and not phys.empty:
                pc1, pc2 = st.columns(2)

                # Asset revenue loss bar
                with pc1:
                    fig_rl = go.Figure()
                    for haz, col, clr in [("H_flood","Flood",C["accent3"]),("H_heat","Heat",C["amber"]),("H_cyclone","Cyclone",C["coral"])]:
                        if haz in phys.columns:
                            rl = phys["revenue_loss"] * phys[haz] / phys["D_total"].replace(0,1)
                            fig_rl.add_trace(go.Bar(x=phys["asset_id"], y=rl.fillna(0), name=col, marker_color=clr))
                    fig_rl.update_layout(**_chart_layout("Revenue Loss by Asset & Hazard (₹ Cr)", 300), barmode="stack")
                    _ax_style(fig_rl)
                    st.plotly_chart(fig_rl, width="stretch")

                # Asset damage heatmap
                with pc2:
                    hm_cols = {"H_flood":"Flood","H_heat":"Heat","H_cyclone":"Cyclone"}
                    hm = phys.set_index("asset_id")[[c for c in hm_cols if c in phys.columns]].rename(columns=hm_cols)*100
                    if not hm.empty:
                        fig_hm = go.Figure(go.Heatmap(
                            z=hm.values.T, x=hm.index.tolist(), y=hm.columns.tolist(),
                            colorscale=[[0,C["mint"]],[0.35,C["amber"]],[0.65,"#F97316"],[1.0,C["coral"]]],
                            text=[[f"{v:.0f}" for v in row] for row in hm.values.T],
                            texttemplate="%{text}", colorbar=dict(title="Risk Score"),
                            zmin=0, zmax=100,
                        ))
                        fig_hm.update_layout(**_chart_layout("Asset Vulnerability Heatmap (0–100)", 280))
                        _ax_style(fig_hm)
                        st.plotly_chart(fig_hm, width="stretch")

            # NGFS scenario projections fan chart
            if isinstance(df_pp, pd.DataFrame) and not df_pp.empty:
                scens_p = sorted(df_pp["Scenario"].unique())
                fig_fan = make_subplots(
                    rows=1, cols=len(scens_p),
                    subplot_titles=[s[:22] for s in scens_p],
                    shared_yaxes=True,
                )
                for ci, scen in enumerate(scens_p, 1):
                    ds = df_pp[df_pp["Scenario"]==scen].sort_values("Year")
                    clr = _scen_color(scen)
                    # Uncertainty band
                    fig_fan.add_trace(go.Scatter(
                        x=ds["Year"].tolist()+ds["Year"].tolist()[::-1],
                        y=ds["Revenue_Loss_P90_Cr"].tolist()+ds["Revenue_Loss_P10_Cr"].tolist()[::-1],
                        fill="toself", fillcolor=_hex_rgba(clr, 0.16), line=dict(color="rgba(0,0,0,0)"),
                        name="P10–P90", showlegend=(ci==1),
                    ), row=1, col=ci)
                    fig_fan.add_trace(go.Scatter(x=ds["Year"], y=ds["Revenue_Loss_P50_Cr"],
                        mode="lines+markers", line=dict(color=clr, width=2.5),
                        marker=dict(size=6), name=f"{scen[:18]} P50", showlegend=(ci==1),
                    ), row=1, col=ci)
                fig_fan.update_layout(**_chart_layout("Physical Revenue Loss — NGFS Scenarios · P10/P50/P90 Bands", 320,
                    legend_override=dict(orientation="h", y=1.1)))
                _ax_style(fig_fan, rows=1, cols=len(scens_p))
                st.plotly_chart(fig_fan, width="stretch")

                # PD evolution across scenarios
                fig_pd_p = go.Figure()
                for scen in scens_p:
                    ds = df_pp[df_pp["Scenario"]==scen].sort_values("Year")
                    fig_pd_p.add_trace(go.Scatter(x=ds["Year"], y=ds["PD_Physical"],
                        mode="lines+markers", name=scen[:28],
                        line=dict(color=_scen_color(scen), width=2.5), marker=dict(size=7)))
                fig_pd_p.update_layout(**_chart_layout("Physical PD — NGFS Scenario Projections", 280))
                fig_pd_p.update_yaxes(tickformat=".2%")
                _ax_style(fig_pd_p)
                st.plotly_chart(fig_pd_p, width="stretch")

            # Folium map
            if isinstance(phys, pd.DataFrame) and not phys.empty:
                hchoice = st.selectbox("Hazard map layer", ["Integrated","Flood","Heat","Cyclone"], key="plots_haz")
                m = folium.Map(location=[phys["latitude"].mean(), phys["longitude"].mean()],
                    zoom_start=5, tiles="cartodbdark_matter")
                for _, r in phys.iterrows():
                    v = (r["H_flood"] if hchoice=="Flood" else r["H_heat"] if hchoice=="Heat"
                         else r["H_cyclone"] if hchoice=="Cyclone" else r["D_total"])
                    clr = "green" if v<0.2 else "orange" if v<0.4 else "red"
                    folium.CircleMarker(
                        location=[r["latitude"], r["longitude"]], radius=9,
                        color=clr, fill=True, fill_opacity=0.78,
                        popup=f"<b>{r['asset_id']}</b><br>Risk: {v:.2f}<br>Rev Loss: ₹{r['revenue_loss']:.1f}Cr"
                    ).add_to(m)
                st_folium(m, use_container_width=True, height=400)

        # ── SECTION 5: INTEGRATED RISK SUMMARY CHART ──────────────────
        if st.session_state.get("integrated_ran", False):
            st.markdown(f"<h3 style='color:{C['purple']};margin-top:8px;'>🧩 Integrated Risk Summary</h3>", unsafe_allow_html=True)
            df_int = st.session_state.get("df_integrated_summary")
            if isinstance(df_int, pd.DataFrame) and not df_int.empty:
                # Extract numeric values for waterfall/bar display
                _rows = df_int[df_int["Metric"].str.contains("PD|ECL|DSCR|Score", na=False)]
                if not _rows.empty:
                    labels = _rows["Metric"].tolist()
                    vals   = pd.to_numeric(_rows["Value"], errors="coerce").fillna(0).tolist()
                    fig_int = go.Figure(go.Bar(
                        x=[l[:35] for l in labels], y=vals,
                        marker_color=[C["accent2"],C["amber"],C["coral"],C["mint"],C["purple"]][:len(labels)],
                        text=[f"{v:.3f}" if isinstance(v,float) else str(v) for v in vals],
                        textposition="outside",
                    ))
                    fig_int.update_layout(**_chart_layout("Integrated Risk Metrics (normalised values)", 280))
                    _ax_style(fig_int)
                    st.plotly_chart(fig_int, width="stretch")

            # MC distribution if available
            mc = st.session_state.get("mc_results")
            if mc and "ECL_95" in mc:
                st.markdown(f"**Monte Carlo summary:** Mean PD={mc.get('Mean_PD',0):.2%} · PD 95th={mc.get('PD_95',0):.2%} · Climate VaR (ECL 95th)=₹{mc.get('ECL_95',0):.1f} Cr")

    # ============================================================
    # TAB 6 — BRSR CORE  (BRSR-01)
    # ============================================================
with brsr_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>📘 BRSR Core — Climate Risk Intelligence</h2>",unsafe_allow_html=True)
    st.caption("SEBI BRSR Core diagnostics with financial quantification · GHG benchmarking · PD linkage · Regulatory readiness")

    if not st.session_state.get("enable_brsr", False):
        st.info("Enable BRSR Diagnostics in the sidebar.")
    else:

        st.subheader("🔢 Operational Climate Inputs")
        bc1,bc2,bc3=st.columns(3)
        with bc1:
            st.markdown(f"<div style='color:{C['accent2']};font-weight:600;margin-bottom:8px;'>⚡ Energy</div>",unsafe_allow_html=True)
            total_energy_kwh    =st.number_input("Total Energy (kWh)",value=1_000_000_000,step=10_000_000,key="brsr_en")
            renewable_share_pct =st.slider("Renewable Share (%)",0,100,15,key="brsr_rs")
            has_verified_data   =st.checkbox("Third-Party Verified Data",value=False,key="brsr_vd")
        with bc2:
            st.markdown(f"<div style='color:{C['amber']};font-weight:600;margin-bottom:8px;'>💧 Water</div>",unsafe_allow_html=True)
            total_water_m3      =st.number_input("Water Withdrawal (m³)",value=50_000_000,step=1_000_000,key="brsr_wt")
            water_stress_region =st.selectbox("Water Stress Region",["Low","Medium","High"],key="brsr_ws")
            recycled_water_pct  =st.slider("Recycled Water (%)",0,100,10,key="brsr_rw")
        with bc3:
            st.markdown(f"<div style='color:{C['coral']};font-weight:600;margin-bottom:8px;'>♻️ Waste & Targets</div>",unsafe_allow_html=True)
            total_waste_mt      =st.number_input("Waste Generated (MT)",value=100_000,step=1_000,key="brsr_wa")
            hazardous_waste_pct =st.slider("Hazardous Waste Share (%)",0,100,20,key="brsr_hw")
            target_coverage_pct =st.slider("Target Coverage (%)",0,100,50,key="brsr_tc")

        st.markdown(f"<div style='color:{C['mint']};font-weight:600;margin:8px 0 4px;'>📋 Governance</div>",unsafe_allow_html=True)
        bg1,bg2,bg3,bg4=st.columns(4)
        has_scope3          =bg1.checkbox("Scope 3 Disclosed",value=True,key="brsr_s3")
        has_board_oversight =bg2.checkbox("Board Climate Oversight",value=False,key="brsr_bo")
        has_net_zero_target =bg3.checkbox("Net Zero / LT Target",value=False,key="brsr_nz")
        has_cbam            =bg4.checkbox("EU/UK Export Exposure",value=False,key="brsr_cb")

        # Locked public demo: BRSR inputs are fixed to the fictional company case, even if widgets render above.
        total_energy_kwh = 1_000_000_000
        renewable_share_pct = 15
        has_verified_data = False
        total_water_m3 = 50_000_000
        water_stress_region = "High"
        recycled_water_pct = 10
        total_waste_mt = 100_000
        hazardous_waste_pct = 20
        target_coverage_pct = 50
        has_scope3 = True
        has_board_oversight = False
        has_net_zero_target = False
        has_cbam = False

        with st.expander("📈 3-Year Emissions Trend (optional)"):
            te1,te2,te3=st.columns(3)
            s1_y0=te1.number_input("Scope 1 (2yr ago)",value=float(scope1*1.08),key="brsr_s1y0")
            s1_y1=te2.number_input("Scope 1 (1yr ago)",value=float(scope1*1.04),key="brsr_s1y1")
            s1_y2=te3.number_input("Scope 1 (current)",value=float(scope1),key="brsr_s1y2")
            r_y0=te1.number_input("Revenue (2yr ago, ₹Cr)",value=float(revenue_0*0.92),key="brsr_r0")
            r_y1=te2.number_input("Revenue (1yr ago, ₹Cr)",value=float(revenue_0*0.96),key="brsr_r1")
            r_y2=te3.number_input("Revenue (current, ₹Cr)",value=float(revenue_0),key="brsr_r2")

        run_brsr=st.button("▶ Run BRSR Diagnostics",type="primary",key="run_brsr_btn")
        if not run_brsr:
            st.info("Click Run BRSR Diagnostics to proceed.")
        else:

            # ── COMPUTE ──
            ghg_bench=SECTOR_GHG_BENCHMARKS.get(sector,SECTOR_GHG_BENCHMARKS["Manufacturing"])
            en_bench =SECTOR_ENERGY_BENCHMARKS.get(sector,SECTOR_ENERGY_BENCHMARKS["Manufacturing"])

            ghg_int = TOTAL_EMISSIONS/max(revenue_0,1e-6)
            s1_int  = scope1/max(revenue_0,1e-6); s2_int=scope2/max(revenue_0,1e-6); s3_int=scope3/max(revenue_0,1e-6)
            en_int  = total_energy_kwh/max(revenue_0*1e6,1)
            wa_int  = total_water_m3/max(revenue_0*1e6,1)

            def pctile(v,b):
                if v<=b["p25"]: return "Top 25% 🟢"
                if v<=b["p50"]: return "25–50th 🟡"
                if v<=b["p75"]: return "50–75th 🟠"
                return "Bottom 25% 🔴"
            ghg_rank=pctile(ghg_int,ghg_bench)

            flags=[]
            if ghg_int>ghg_bench["p75"]:         flags.append("High GHG intensity (>P75)")
            if renewable_share_pct<20:           flags.append("Low renewable energy (<20%)")
            if en_int>en_bench["p75"]:           flags.append("High energy intensity (>P75)")
            if water_stress_region=="High":      flags.append("High water stress + high withdrawal")
            if target_coverage_pct<50:          flags.append("Low target coverage (<50%)")
            if hazardous_waste_pct>30:          flags.append("High hazardous waste (>30%)")
            if not has_verified_data:           flags.append("No verified emissions data")
            if not has_scope3:                  flags.append("Missing Scope 3 disclosure")

            pd_adj=float(np.clip(sum(BRSR_PD_SPREAD.get(f,0) for f in flags),0,0.015))

            # Water stranded cost
            esc_r={"Low":0.03,"Medium":0.08,"High":0.18}[water_stress_region]
            w_cum_cost=sum(total_water_m3*42*((1+esc_r)**yr-(1+esc_r)**(yr-1))/1e7 for yr in range(1,6))
            # Energy transition risk
            fossil_kwh=total_energy_kwh*(100-renewable_share_pct)/100
            energy_tr_risk=fossil_kwh*2.0/1e7

            # Regulatory readiness (8-point)
            rdness_items={
                "Renewable ≥20%":1.0 if renewable_share_pct>=20 else 0.5 if renewable_share_pct>=10 else 0.0,
                "Target Coverage ≥50%":1.0 if target_coverage_pct>=50 else 0.5 if target_coverage_pct>=25 else 0.0,
                "Hazardous Waste ≤30%":1.0 if hazardous_waste_pct<=30 else 0.5 if hazardous_waste_pct<=40 else 0.0,
                "Water Risk Mgmt":1.0 if water_stress_region=="Low" else 0.5 if water_stress_region=="Medium" else 0.0,
                "Scope 3 Disclosed":1.0 if has_scope3 else 0.0,
                "Verified Data":1.0 if has_verified_data else 0.0,
                "Board Oversight":1.0 if has_board_oversight else 0.0,
                "Net Zero Target":1.0 if has_net_zero_target else 0.0,
            }
            readiness_score=sum(rdness_items.values())/len(rdness_items)*100

            risk_score_brsr=0
            if renewable_share_pct<10: risk_score_brsr+=20
            elif renewable_share_pct<20: risk_score_brsr+=12
            if hazardous_waste_pct>40: risk_score_brsr+=15
            elif hazardous_waste_pct>30: risk_score_brsr+=8
            if target_coverage_pct<30: risk_score_brsr+=18
            elif target_coverage_pct<50: risk_score_brsr+=10
            if water_stress_region=="High": risk_score_brsr+=18
            elif water_stress_region=="Medium": risk_score_brsr+=8
            if ghg_int>ghg_bench["p75"]: risk_score_brsr+=15
            elif ghg_int>ghg_bench["p50"]: risk_score_brsr+=7
            if not has_board_oversight: risk_score_brsr+=8
            if not has_net_zero_target: risk_score_brsr+=6
            risk_score_brsr=min(100,risk_score_brsr)

            sev_brsr="🟢 Low" if risk_score_brsr<30 else "🟡 Moderate" if risk_score_brsr<60 else "🔴 High"
            sev_color_brsr=C["mint"] if risk_score_brsr<30 else C["amber"] if risk_score_brsr<60 else C["coral"]

            brsr_summary=pd.DataFrame([{"GHG_Intensity":round(ghg_int,3),"S1_Intensity":round(s1_int,3),
                "S2_Intensity":round(s2_int,3),"S3_Intensity":round(s3_int,3),"Energy_Intensity":round(en_int,3),
                "Renewable_%":renewable_share_pct,"Water_Intensity":round(wa_int,3),"Recycled_Water_%":recycled_water_pct,
                "Hazardous_Waste_%":hazardous_waste_pct,"Target_Coverage_%":target_coverage_pct,
                "BRSR_PD_Adj_bps":round(pd_adj*10000,1),"Water_Cost_Risk_Cr":round(w_cum_cost,2),
                "Energy_TR_Risk_Cr":round(energy_tr_risk,2),"Readiness_Score":round(readiness_score,1),
                "Overall_Risk_Score":risk_score_brsr}])
            st.session_state["brsr_summary"]=brsr_summary
            st.session_state["brsr_flags"]=pd.DataFrame({"Flag":flags})
            st.session_state["brsr_pd_adj"]=pd_adj
            st.session_state["brsr_ran"]=True

            # ── DISPLAY ──
            st.markdown(f"""
            <div style="background:{C['card']};border:1px solid {sev_color_brsr};border-radius:10px;padding:14px 18px;margin-bottom:12px;display:flex;align-items:center;gap:16px;">
              <div style="font-size:30px;">{sev_brsr.split()[0]}</div>
              <div>
                <div style="font-size:16px;font-weight:700;color:{sev_color_brsr};">{sev_brsr} Operational Climate Risk · Score: {risk_score_brsr}/100</div>
                <div style="font-size:13px;color:{C['slate']};margin-top:2px;">{len(flags)} flags · PD uplift: +{pd_adj*10000:.0f}bps · Readiness: {readiness_score:.0f}% · Water cost risk (5yr): ₹{w_cum_cost:.1f}Cr</div>
              </div>
            </div>""",unsafe_allow_html=True)

            bm1,bm2,bm3,bm4,bm5=st.columns(5)
            bm1.metric("Risk Score",f"{risk_score_brsr}/100")
            bm2.metric("BRSR Readiness",f"{readiness_score:.0f}%")
            bm3.metric("PD Uplift",f"+{pd_adj*10000:.0f} bps")
            bm4.metric("Water Cost Risk (5yr)",f"₹{w_cum_cost:.1f} Cr")
            bm5.metric("Energy Transition Risk",f"₹{energy_tr_risk:.1f} Cr")

            # Chart 1: GHG benchmark
            st.subheader("📊 GHG Intensity — Sector Benchmark")
            fig_ghg=make_subplots(rows=1,cols=2,subplot_titles=("vs Sector Percentiles","Scope Breakdown"))
            bench_vals=[ghg_bench["p25"],ghg_bench["p50"],ghg_bench["p75"],ghg_int]
            bench_clrs=[C["mint"],C["amber"],"#F97316",C["accent2"] if ghg_int<=ghg_bench["p75"] else C["coral"]]
            fig_ghg.add_trace(go.Bar(x=["P25","P50","P75","This Company"],y=bench_vals,marker_color=bench_clrs,
                text=[f"{v:.2f}" for v in bench_vals],textposition="outside"),row=1,col=1)
            fig_ghg.add_trace(go.Bar(x=["Scope 1","Scope 2","Scope 3"],y=[s1_int,s2_int,s3_int],
                marker_color=[C["coral"],C["amber"],C["accent3"]],text=[f"{v:.2f}" for v in [s1_int,s2_int,s3_int]],textposition="outside"),row=1,col=2)
            fig_ghg.update_layout(**_chart_layout("",300)); _ax_style(fig_ghg,rows=1,cols=2)
            fig_ghg.update_yaxes(title="tCO₂e/₹Cr",row=1,col=1); fig_ghg.update_yaxes(title="tCO₂e/₹Cr",row=1,col=2)
            st.plotly_chart(fig_ghg,width="stretch")
            st.caption(f"Sector ({sector}): P25={ghg_bench['p25']}, P50={ghg_bench['p50']}, P75={ghg_bench['p75']} tCO₂e/₹Cr · Your intensity: **{ghg_int:.2f}** — {ghg_rank}")

            # Chart 2: PD uplift tornado
            if flags:
                st.subheader("🔗 BRSR Flags → PD Uplift")
                flag_bps={f:BRSR_PD_SPREAD.get(f,0)*10000 for f in flags}
                fig_tor=go.Figure(go.Bar(x=list(flag_bps.values()),y=list(flag_bps.keys()),orientation="h",
                    marker_color=[C["coral"] if v>=25 else C["amber"] if v>=15 else C["mint"] for v in flag_bps.values()],
                    text=[f"+{v:.0f} bps" for v in flag_bps.values()],textposition="outside"))
                fig_tor.update_layout(**_chart_layout(f"PD Uplift by BRSR Flag (Total: +{pd_adj*10000:.0f}bps)",max(220,len(flags)*45+80),
                    margin_override=dict(l=250,r=80,t=50,b=20)))
                fig_tor.update_xaxes(title="Basis Point Uplift")
                _ax_style(fig_tor); st.plotly_chart(fig_tor,width="stretch")
            else:
                st.success("✅ No BRSR flags — no PD uplift from operational climate risk")

            # Chart 3: Readiness radar
            st.subheader("🎯 SEBI BRSR Core Regulatory Readiness")
            rcol1,rcol2=st.columns([2,1])
            with rcol1:
                rcats=list(rdness_items.keys()); rvals=[rdness_items[c]*100 for c in rcats]
                fig_rad=go.Figure()
                fig_rad.add_trace(go.Scatterpolar(r=rvals+[rvals[0]],theta=rcats+[rcats[0]],fill="toself",
                    fillcolor="rgba(6,182,212,0.15)",line=dict(color=C["accent2"],width=2),name="Readiness"))
                fig_rad.add_trace(go.Scatterpolar(r=[100]*(len(rcats)+1),theta=rcats+[rcats[0]],
                    line=dict(color=C["mint"],width=1,dash="dot"),name="Full Compliance"))
                fig_rad.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,100]),bgcolor=C["bg_dark"]),
                    **_chart_layout("",360),showlegend=True)
                st.plotly_chart(fig_rad,width="stretch")
            with rcol2:
                st.dataframe(pd.DataFrame([{"Item":k,"Status":"✅" if v==1 else "⚠️" if v==0.5 else "❌"} for k,v in rdness_items.items()])
                    .style.set_properties(**{"background-color":C["bg_dark"],"color":C["text"]}),
                    width="stretch",height=320,hide_index=True)

            # Chart 4: Forward financial risk
            st.subheader("💰 5-Year Forward Operational Climate Cost")
            yr5=list(range(1,6))
            wc_ann=[total_water_m3*42*((1+esc_r)**y-(1+esc_r)**(y-1))/1e7 for y in yr5]
            ec_ann=[fossil_kwh*(0.40*y)/1e7 for y in yr5]
            fig_fwd=go.Figure()
            fig_fwd.add_trace(go.Bar(x=yr5,y=wc_ann,name="Water Cost Escalation",marker_color=C["accent3"]))
            fig_fwd.add_trace(go.Bar(x=yr5,y=ec_ann,name="Energy Carbon Surcharge",marker_color=C["amber"]))
            fig_fwd.update_layout(**_chart_layout("Projected Annual Operational Climate Cost (₹ Cr)",280),barmode="stack")
            _ax_style(fig_fwd); st.plotly_chart(fig_fwd,width="stretch")

            # Chart 5: Emissions trend
            if s1_y0>0 and r_y0>0:
                st.subheader("📈 Scope 1 Intensity Trend")
                ti=[s1_y0/r_y0,s1_y1/r_y1,s1_y2/r_y2]
                cagr=(ti[2]/ti[0])**0.5-1; dir_t="↓ Improving" if cagr<-0.02 else "↑ Worsening" if cagr>0.02 else "→ Stable"
                dir_clr=C["mint"] if cagr<-0.02 else C["coral"] if cagr>0.02 else C["amber"]
                fig_tr=go.Figure()
                fig_tr.add_trace(go.Scatter(x=["2 Years Ago","1 Year Ago","Current"],y=ti,mode="lines+markers",
                    line=dict(color=C["accent2"],width=3),marker=dict(size=10,color=C["accent"]),name="Scope 1 Intensity"))
                fig_tr.add_hline(y=ghg_bench["p50"],line_dash="dash",line_color=C["amber"],annotation_text=f"Sector P50 ({ghg_bench['p50']})",annotation_font_color=C["amber"])
                fig_tr.update_layout(**_chart_layout(f"Scope 1 GHG Intensity Trend — {dir_t} (CAGR: {cagr*100:.1f}%/yr)",260))
                _ax_style(fig_tr); st.plotly_chart(fig_tr,width="stretch")

            # Compliance table
            st.subheader("📋 BRSR Compliance Status")
            crows=[
                {"Indicator":"GHG Intensity vs Sector","Value":f"{ghg_int:.2f} tCO₂/₹Cr","Threshold":f"P50={ghg_bench['p50']}","Status":"🟢" if ghg_int<=ghg_bench["p50"] else "🟡" if ghg_int<=ghg_bench["p75"] else "🔴","Percentile":ghg_rank},
                {"Indicator":"Renewable Energy","Value":f"{renewable_share_pct}%","Threshold":"≥20%","Status":"🟢" if renewable_share_pct>=20 else "🟡" if renewable_share_pct>=10 else "🔴","Percentile":"Compliant" if renewable_share_pct>=20 else "Below"},
                {"Indicator":"Target Coverage","Value":f"{target_coverage_pct}%","Threshold":"≥50%","Status":"🟢" if target_coverage_pct>=50 else "🟡" if target_coverage_pct>=25 else "🔴","Percentile":"Compliant" if target_coverage_pct>=50 else "Below"},
                {"Indicator":"Hazardous Waste","Value":f"{hazardous_waste_pct}%","Threshold":"≤30%","Status":"🟢" if hazardous_waste_pct<=30 else "🔴","Percentile":"Compliant" if hazardous_waste_pct<=30 else "Exceeds"},
                {"Indicator":"Water Stress","Value":water_stress_region,"Threshold":"Low/Medium","Status":"🟢" if water_stress_region!="High" else "🔴","Percentile":"No elevated risk" if water_stress_region!="High" else "High risk"},
                {"Indicator":"Regulatory Readiness","Value":f"{readiness_score:.0f}%","Threshold":"≥75%","Status":"🟢" if readiness_score>=75 else "🟡" if readiness_score>=50 else "🔴","Percentile":"Ready" if readiness_score>=75 else "Needs work"},
            ]
            st.dataframe(pd.DataFrame(crows).style.set_properties(**{"background-color":C["bg_dark"],"color":C["text"]}),
                width="stretch",hide_index=True)

            # Integration link to transition engine
            if st.session_state.get("transition_ran",False):
                df_tl=st.session_state.get("df_transition")
                if isinstance(df_tl,pd.DataFrame) and not df_tl.empty:
                    st.subheader("🔗 Integration: BRSR Uplift on Transition PD")
                    cb_pk=df_tl["Carbon_Burden"].max(); pd_pk=df_tl["PD_Transition"].max()
                    ic1,ic2,ic3=st.columns(3)
                    ic1.metric("Base Transition PD",f"{pd_pk:.2%}")
                    ic2.metric("BRSR PD Uplift",f"+{pd_adj*10000:.0f} bps")
                    ic3.metric("Total PD (Transition + BRSR)",f"{min(pd_pk+pd_adj,PD_CAP):.2%}",delta=f"+{pd_adj*10000:.0f}bps",delta_color="inverse")
                    st.caption("BRSR uplift is additive operational climate risk premium. Captures risks outside carbon-price transmission chain: water scarcity, energy cost escalation, governance deficiencies, regulatory non-compliance penalties.")

            st.success("✅ BRSR Enhanced Diagnostics v1.2 completed")
            log_model_run("BRSR",{"company":company_name,"risk_score":risk_score_brsr,"pd_adj_bps":pd_adj*10000,"readiness":readiness_score})

        # ============================================================
        # TAB 7 — AI NARRATIVE
        # ============================================================
with ai_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>🤖 AI Business & Climate Intelligence Advisor</h2>", unsafe_allow_html=True)
    st.caption("Seven specialist advisors · Exact problems · Exact solutions · Implementation steps · Business impact")

    t_r = st.session_state.get("transition_ran", False)
    p_r = st.session_state.get("physical_ran",   False)
    b_r = st.session_state.get("brsr_ran",       False)
    i_r = st.session_state.get("integrated_ran", False)

    if not (t_r or p_r or b_r):
        st.info("Run at least one module (Transition Risk, Physical Risk, or BRSR) to unlock the AI advisors.")
    else:

        # ── Build rich data payload ─────────────────────────────────────
        def _build_rich_payload():
            pl = {
                "company": company_name,
                "sector": sector,
                "reporting_year": REPORTING_YEAR,
                "baseline_scenario": BASELINE_SCENARIO,
                "financials": {
                    "revenue_cr": revenue_0,
                    "ebitda_margin_pct": round(ebitda_margin_0 * 100, 1),
                    "interest_cr": interest_payment,
                    "total_assets_cr": total_assets,
                    "ead_cr": exposure_at_default,
                    "scope1_tco2e": scope1,
                    "scope2_tco2e": scope2,
                    "scope3_tco2e": scope3,
                    "total_emissions_tco2e": TOTAL_EMISSIONS,
                    "high_carbon_assets_cr": high_carbon_assets,
                    "base_pd_pct": round(base_pd * 100, 3),
                    "base_lgd_pct": round(LGD_0 * 100, 1),
                },
            }
            if t_r:
                df_ai = st.session_state.get("df_transition")
                if isinstance(df_ai, pd.DataFrame) and not df_ai.empty:
                    scen_summary = df_ai.groupby("Scenario").agg(
                        Peak_PD=("PD_Transition", "max"),
                        Peak_ECL_Cr=("ECL_Transition", "max"),
                        Min_DSCR=("DSCR", "min"),
                        Max_Carbon_Burden_pct=("Carbon_Burden", lambda x: round(x.max()*100, 2)),
                        Max_Stranded_Cr=("Stranded_Assets", "max"),
                        Max_CAPEX_Gap_Cr=("CAPEX_Gap", "max"),
                        Min_EBITDA_Margin_pct=("EBITDA_Margin", lambda x: round(x.min()*100, 2)),
                    ).round(4).to_dict()
                    # Year-wise data for worst scenario
                    worst_scen = df_ai.groupby("Scenario")["PD_Transition"].max().idxmax()
                    worst_path = df_ai[df_ai["Scenario"] == worst_scen][
                        ["Year", "Revenue", "EBITDA", "DSCR", "PD_Transition", "ECL_Transition",
                         "Carbon_Burden", "Stranded_Assets", "CAPEX_Gap"]
                    ].round(4).to_dict("records")
                    pl["transition_risk"] = {
                        "scenario_summary": scen_summary,
                        "worst_scenario": worst_scen,
                        "worst_scenario_year_by_year": worst_path,
                        "carbon_pass_through": carbon_pass_through,
                        "planned_capex_cr": planned_capex,
                        "abatement_cost_inr_per_tco2": abatement_cost,
                        "abatement_potential_pct": round(abatement_potential * 100, 0),
                        "governance_score": round(G, 3),
                    }
            if p_r:
                ps = st.session_state.get("phys_summary", {})
                pa = st.session_state.get("phys_assets")
                pp = st.session_state.get("df_physical_projection")
                pl["physical_risk"] = {
                    "summary": ps,
                    "asset_count": len(pa) if isinstance(pa, pd.DataFrame) else 0,
                    "top_risk_assets": (pa.nlargest(3, "D_total")[
                        ["asset_id","asset_type","D_total","revenue_loss","flood_depth_m","heat_days","max_wind_kmh"]
                    ].round(3).to_dict("records") if isinstance(pa, pd.DataFrame) and "D_total" in pa.columns else []),
                    "ngfs_peak_rev_loss_p50": (
                        pp["Revenue_Loss_P50_Cr"].max() if isinstance(pp, pd.DataFrame) and not pp.empty else 0
                    ),
                }
            if b_r:
                bs = st.session_state.get("brsr_summary")
                bf = st.session_state.get("brsr_flags")
                pl["brsr"] = {
                    "kpis": bs.iloc[0].to_dict() if isinstance(bs, pd.DataFrame) and not bs.empty else {},
                    "active_flags": bf["Flag"].tolist() if isinstance(bf, pd.DataFrame) and "Flag" in bf.columns else [],
                    "pd_uplift_bps": round(st.session_state.get("brsr_pd_adj", 0) * 10000, 1),
                    "readiness_score_pct": float(bs.iloc[0].get("Readiness_Score", 0)) if isinstance(bs, pd.DataFrame) and not bs.empty else 0,
                    "water_cost_risk_5yr_cr": float(bs.iloc[0].get("Water_Cost_Risk_Cr", 0)) if isinstance(bs, pd.DataFrame) and not bs.empty else 0,
                    "energy_transition_risk_cr": float(bs.iloc[0].get("Energy_TR_Risk_Cr", 0)) if isinstance(bs, pd.DataFrame) and not bs.empty else 0,
                }
            if i_r:
                di = st.session_state.get("df_integrated_summary")
                pl["integrated_risk"] = di.to_dict("records") if isinstance(di, pd.DataFrame) else []
            mc = st.session_state.get("mc_results")
            if mc:
                pl["monte_carlo"] = {
                    "mean_pd_pct": round(mc.get("Mean_PD", 0) * 100, 3),
                    "pd_95th_pct": round(mc.get("PD_95", 0) * 100, 3),
                    "mean_ecl_cr": round(mc.get("Mean_ECL", 0), 2),
                    "climate_var_95_cr": round(mc.get("ECL_95", 0), 2),
                    "unexpected_loss_cr": round(mc.get("ECL_95", 0) - mc.get("Mean_ECL", 0), 2),
                }
            return pl

        def _safe_float(v, default=0.0):
            try:
                if pd.isna(v):
                    return default
                return float(v)
            except Exception:
                return default

        def _compact_payload_for_ai(pl):
            """Create a small, advisor-ready data packet to stay below Groq TPM limits."""
            fin = pl.get("financials", {})
            out = {
                "company": pl.get("company"),
                "sector": pl.get("sector"),
                "year": pl.get("reporting_year"),
                "baseline_scenario": pl.get("baseline_scenario"),
                "financials": {
                    "revenue_cr": fin.get("revenue_cr"),
                    "ebitda_margin_pct": fin.get("ebitda_margin_pct"),
                    "interest_cr": fin.get("interest_cr"),
                    "ead_cr": fin.get("ead_cr"),
                    "base_pd_pct": fin.get("base_pd_pct"),
                    "base_lgd_pct": fin.get("base_lgd_pct"),
                    "total_emissions_tco2e": fin.get("total_emissions_tco2e"),
                    "high_carbon_assets_cr": fin.get("high_carbon_assets_cr"),
                },
            }

            tr = pl.get("transition_risk", {})
            if tr:
                scen = tr.get("scenario_summary", {})
                worst_path = tr.get("worst_scenario_year_by_year", []) or []
                key_years = []
                if worst_path:
                    key_years = [worst_path[0]]
                    if len(worst_path) > 2:
                        key_years.append(worst_path[len(worst_path)//2])
                    if len(worst_path) > 1:
                        key_years.append(worst_path[-1])
                out["transition"] = {
                    "worst_scenario": tr.get("worst_scenario"),
                    "scenario_summary": scen,
                    "key_years_only": key_years,
                    "planned_capex_cr": tr.get("planned_capex_cr"),
                    "abatement_cost_inr_per_tco2": tr.get("abatement_cost_inr_per_tco2"),
                    "abatement_potential_pct": tr.get("abatement_potential_pct"),
                    "governance_score": tr.get("governance_score"),
                }

            pr = pl.get("physical_risk", {})
            if pr:
                out["physical"] = {
                    "summary": pr.get("summary", {}),
                    "asset_count": pr.get("asset_count", 0),
                    "top_risk_assets": (pr.get("top_risk_assets", []) or [])[:2],
                    "ngfs_peak_rev_loss_p50": pr.get("ngfs_peak_rev_loss_p50", 0),
                }

            br = pl.get("brsr", {})
            if br:
                out["brsr"] = {
                    "pd_uplift_bps": br.get("pd_uplift_bps", 0),
                    "readiness_score_pct": br.get("readiness_score_pct", 0),
                    "active_flags": (br.get("active_flags", []) or [])[:6],
                    "water_cost_risk_5yr_cr": br.get("water_cost_risk_5yr_cr", 0),
                    "energy_transition_risk_cr": br.get("energy_transition_risk_cr", 0),
                    "kpis": br.get("kpis", {}),
                }

            integ = pl.get("integrated_risk", [])
            if integ:
                out["integrated"] = integ[:9]
            if pl.get("monte_carlo"):
                out["monte_carlo"] = pl.get("monte_carlo")
            if pl.get("additional_context"):
                out["additional_context"] = str(pl.get("additional_context"))[:900]
            if pl.get("specific_focus"):
                out["specific_focus"] = str(pl.get("specific_focus"))[:350]
            return out

        def _json_compact(obj, max_chars=5200):
            txt = json.dumps(obj, ensure_ascii=False, separators=(",", ":"), default=str)
            if len(txt) > max_chars:
                return txt[:max_chars] + "...[TRUNCATED_TO_FIT_GROQ_LIMIT]"
            return txt

        def _optimized_system_prompt(advisor_name, original_system):
            """Shorter prompts for high-value advisors to reduce tokens and improve consistency."""
            common = (
                "Use only the data provided. Be concise, numerical, India-focused, and action-oriented. "
                "Do not invent exact data; when estimating, label it as an estimate. "
                "Keep output under 900 words. Use markdown tables where useful."
            )
            if "CFO Action Advisor" in advisor_name:
                return common + """
ROLE: CFO climate-finance advisor for Indian corporates.
OUTPUT:
1) CFO priorities table with exactly 3 rows: Action, ₹ cost/funding, timeline, metric impact (PD/ECL/DSCR), risk if not done.
2) Financing plan: green bonds, ESG-linked loan, REC/carbon benefits, SIDBI/IFC/ADB options; include estimated ticket/rate ranges.
3) 90-day execution checklist.
4) What NOT to do: 3 mistakes.
Be specific and use the company's actual risk metrics."""
            if "Operations & Engineering Advisor" in advisor_name:
                return common + """
ROLE: Operations and decarbonisation engineering advisor for Indian heavy industry.
OUTPUT:
1) Operational root cause in 2 bullets.
2) Quick wins table: 3-4 actions, ₹ cost, annual saving, tCO₂e impact, payback, Indian vendor examples.
3) Medium-term CAPEX table: 2-3 projects, priority, cost, timeline, effect on carbon burden/DSCR/PD.
4) Year-1 implementation sequence by quarter.
Keep recommendations practical for the stated sector."""
            if "Strategic Planning Advisor" in advisor_name:
                return common + """
ROLE: Strategy advisor building a climate-resilient 5-year roadmap.
OUTPUT:
1) Strategic verdict in 3 sentences.
2) 5-year roadmap table: Year, investment decision, business-model move, metric target, milestone.
3) Competitive dynamics: who wins/loses and why.
4) 2 hidden opportunities with revenue logic.
5) 3 strategic risks to avoid.
Tie every recommendation to PD, DSCR, ECL, carbon burden, CAPEX gap or BRSR flags."""
            return common + "\n" + original_system[:1800]

        def _build_ai_messages(advisor_name, adv, payload):
            compact_payload = _compact_payload_for_ai(payload)
            data_brief = _json_compact(compact_payload, max_chars=5200)
            focus = compact_payload.get("specific_focus") or "Use the most material risks in the data."
            user_msg = (
                f"TASK: {adv['user_prefix']}\n"
                f"COMPANY: {company_name} | SECTOR: {sector} | YEAR: {REPORTING_YEAR}\n"
                f"FOCUS: {focus}\n"
                "METHOD: First silently identify the 3 most material risks from the data, then write the final answer only.\n"
                f"COMPACT_DATA_JSON:\n{data_brief}"
            )
            return [
                {"role": "system", "content": _optimized_system_prompt(advisor_name, adv["system"])},
                {"role": "user", "content": user_msg},
            ]

        # ── Advisor definitions ─────────────────────────────────────────
        ADVISORS = {
            "🔴 Crisis Diagnostician": {
                "icon": "🔴",
                "subtitle": "What is breaking, why, and how bad is it really",
                "color": C["coral"],
                "description": "Cuts through the numbers to tell you **exactly what is going wrong** in plain language — the real financial threats, the timeline, and what happens if nothing changes.",
                "system": """You are a blunt, experienced crisis diagnostic advisor combining deep expertise in corporate finance, credit risk, and climate economics. You have 25 years of experience restructuring distressed companies in carbon-intensive sectors across India and Southeast Asia.

YOUR JOB: Diagnose exactly what is going wrong with this company's finances under climate stress. Be specific, be direct, use real numbers from the data. No hedging. No regulatory jargon. Speak like a CFO who has seen companies fail.

STRUCTURE YOUR RESPONSE AS:

## 🔴 The Core Problem (2-3 sentences, brutally honest)
State the single biggest financial threat in plain language with exact numbers.

## 📉 The Cascade — What Breaks First, Then Next
Show the exact sequence: which metric hits the danger zone first, in which year, under which scenario. Use the year-by-year data. Be specific: "DSCR falls below covenant threshold of 1.2x in 2031 under Net Zero 2050, meaning the company cannot service ₹X Cr of debt without refinancing."

## 💣 The Three Landmines
The three specific things that could trigger a sudden credit event (not gradual deterioration). What triggers them. What year. What the impact would be on the bank's ECL.

## ⏰ How Much Time Is There
Give a realistic timeline: "Under current trajectory, the company has approximately X years before [specific crisis event] becomes likely."

## 🩺 Severity Verdict
Rate: [Manageable / Serious / Critical / Terminal] and explain exactly why in 3 sentences.

Use actual numbers from the data throughout. Do not use vague language like "may be impacted" — say "EBITDA will fall to ₹X Cr" or "PD increases from X% to Y%".""",
                "user_prefix": "Diagnose exactly what is going wrong with this company under climate stress. Use all numbers provided. Be direct and specific.",
            },

            "💰 CFO Action Advisor": {
                "icon": "💰",
                "subtitle": "The exact financial moves the CFO must make, with costs and timelines",
                "color": C["amber"],
                "description": "Gives the CFO a **prioritised action plan** with specific financial decisions, exact amounts, implementation timelines, and projected impact on PD, ECL, and DSCR.",
                "system": """You are a CFO advisor with 20 years of experience in capital allocation, debt restructuring, and climate transition finance at Indian conglomerates and PSU banks. You have overseen 15+ green bond issuances and climate CAPEX programmes.

YOUR JOB: Give the CFO a precise, implementable financial action plan based on the risk data. Every recommendation must include: what to do, how much it costs, how long it takes, and what it does to the key financial metrics.

STRUCTURE YOUR RESPONSE AS:

## 💰 CFO Priority #1: [Most Urgent Action — Name It Specifically]
**What:** Exactly what financial action to take (e.g., "Issue ₹X Cr green bonds to fund solar CAPEX")
**Cost/Benefit:** Exact cost, exact benefit to EBITDA/DSCR/PD
**Timeline:** Month-by-month implementation
**Risk if not done:** Specific consequence with numbers
**How to implement:** 3-4 concrete steps

## 💰 CFO Priority #2: [Second Most Urgent]
[Same structure]

## 💰 CFO Priority #3: [Third Priority]
[Same structure]

## 📊 Financial Impact Summary Table
Show before/after numbers for: DSCR, PD, ECL, Carbon Burden, CAPEX Gap if all three actions are implemented.

## ⚠️ What NOT To Do
2-3 common CFO mistakes in this situation that would make things worse. Be specific about why.

## 🏦 Financing Options Available
Specific instruments available for this sector in India: green bonds, REC monetisation, ESG-linked loans, SIDBI climate finance, IFC/ADB green facilities. Include approximate rates and ticket sizes.""",
                "user_prefix": "Give the CFO an exact, costed action plan to address these climate financial risks. Every action must have specific rupee amounts, timelines, and projected impact on metrics.",
            },

            "🏭 Operations & Engineering Advisor": {
                "icon": "🏭",
                "subtitle": "Exactly what to change in operations, with costs and carbon/cost impact",
                "color": C["mint"],
                "description": "Translates financial risk into **specific operational changes** — which equipment to replace, which processes to change, what the engineering costs are, and how much carbon and money each action saves.",
                "system": """You are an operations and engineering advisor specialising in decarbonisation of heavy industry in India. You have led 20+ energy transition projects at Steel, Power, Cement, and Oil & Gas plants. You know the specific Indian vendors, technology costs, and implementation timelines.

YOUR JOB: Turn the financial risk numbers into specific operational actions. The user needs to know exactly what to change on the ground — machinery, processes, energy sources — with real costs and real carbon/cost savings.

STRUCTURE YOUR RESPONSE AS:

## 🏭 Operational Root Cause
What is the specific operational driver of the company's climate financial risk? (e.g., "The core problem is reliance on coal-fired captive power at ₹X/kWh when grid + solar would cost ₹Y/kWh")

## ⚡ Quick Wins (0–12 months, under ₹50 Cr each)
List 3-5 specific operational changes:
- **Action:** [Specific change — e.g., "Install Variable Frequency Drives on X motors"]
- **Cost:** ₹X Cr
- **Annual saving:** ₹X Cr / X tCO₂e
- **Payback:** X months
- **Indian vendors:** [Specific companies]

## 🔧 Medium-Term Projects (1–3 years, ₹50–500 Cr)
List 3-4 specific capital projects with same detail as above.

## 🌱 Strategic Transformation (3–7 years, >₹500 Cr)
The 2-3 structural changes needed for long-term decarbonisation viability.

## 📐 CAPEX Prioritisation
Given the CAPEX gap from the model, which projects to do first to get the maximum reduction in PD and carbon burden per rupee spent.

## 🏗️ Implementation Sequencing
Month-by-month Gantt-style plan for Year 1. Be specific about what starts when.""",
                "user_prefix": "Translate these climate risk numbers into specific operational actions for this sector. Include exact costs, Indian vendors, and carbon/cost savings for each action.",
            },

            "📋 Regulatory Compliance Advisor": {
                "icon": "📋",
                "subtitle": "Exact compliance gaps, exact filing requirements, exact deadlines",
                "color": C["accent2"],
                "description": "Maps the company's current state against **every applicable Indian and global regulation** — what is non-compliant, what the penalties are, and the exact steps to achieve compliance.",
                "system": """You are a regulatory compliance specialist with deep expertise in Indian ESG and climate disclosure law: SEBI BRSR Core, RBI climate risk guidelines, Companies Act 2013 sustainability reporting, and ISSB IFRS S2. You have helped 40+ Indian listed companies achieve SEBI compliance and have appeared before RBI supervisory panels.

YOUR JOB: Tell the company exactly where it is non-compliant, what the specific regulatory consequences are, and give a precise compliance roadmap.

STRUCTURE YOUR RESPONSE AS:

## 🚨 Critical Non-Compliance Items (Act Now)
For each: The specific regulation, the specific gap, the specific penalty or consequence, and the deadline.

## ⚠️ Moderate Gaps (Address Within 6 Months)
Same format.

## ✅ Currently Compliant (Keep Doing)
Brief list of what is already in order.

## 📅 Regulatory Deadline Calendar
Month-by-month calendar of upcoming disclosure deadlines for this company:
- SEBI BRSR Core filing dates
- RBI climate stress test submission
- ICAAP climate overlay requirement
- ISSB S2 alignment dates

## 📝 Exact Disclosure Language
For each major gap, provide 2-3 sentences of **board-approved disclosure language** that accurately describes the risk without triggering regulatory concern. This is the exact text that can go into the annual report.

## 🏛️ Regulatory Engagement Strategy
Which regulators to proactively engage with, what to say, and how to get ahead of mandatory requirements rather than react to them.""",
                "user_prefix": "Identify every compliance gap against Indian and global climate regulations. Provide exact deadlines, penalties, and disclosure language that can be used directly in filings.",
            },

            "🏦 Banker & Credit Advisor": {
                "icon": "🏦",
                "subtitle": "How lenders see this company, what they will do, how to manage the relationship",
                "color": C["purple"],
                "description": "Tells you exactly **how your bank is thinking about this exposure** — whether they are likely to cut limits, raise spreads, or add covenants — and what to do to protect the credit relationship.",
                "system": """You are a senior credit risk advisor who has spent 18 years on both sides — as a corporate banker at SBI, HDFC Bank, and ICICI Bank, and as a credit risk consultant advising large Indian corporates on managing their bank relationships. You know exactly how Indian credit committees think about climate risk.

YOUR JOB: Tell the corporate exactly how their lending bank(s) are assessing this exposure right now, what credit actions they are likely to take, and how the corporate should manage the relationship.

STRUCTURE YOUR RESPONSE AS:

## 🏦 How Your Bank Is Reading These Numbers
Be specific: "A PD of X% puts this borrower in the [internal rating category] of most Indian banks. At this level, the relationship manager is likely to [specific action]."

## 📉 Credit Actions the Bank Is Likely to Take (Next 12–24 Months)
For each likely action: the trigger, the timeline, the financial impact on the company.
- Covenant review / tightening
- Margin uplift (specify basis points range)
- Limit reduction or non-renewal
- Enhanced monitoring / watch list
- Requirement for BRSR/climate disclosures

## 💬 What To Say To Your Banker (Right Now)
Exact talking points for the next banker meeting. What to proactively disclose, what narrative to lead with, what data to bring.

## 📄 Credit Documentation Strategy
Specific changes to make to the loan documentation at next renewal: covenants to negotiate, representations to include, green-linked pricing structures to propose.

## 🤝 Green Finance Opportunity
Based on these metrics, which green finance products is this company eligible for right now? Specify banks, schemes, approximate pricing vs conventional debt.

## ⚠️ What Would Trigger a Credit Crisis
The 3 specific scenarios (with numbers) that would cause the bank to move this to a stressed account. How far away is each trigger.""",
                "user_prefix": "Analyse how lending banks in India would assess this climate risk data. Tell the company exactly what credit actions to expect and how to manage the relationship.",
            },

            "📈 Investor & Equity Advisor": {
                "icon": "📈",
                "subtitle": "How markets will price this risk, what ESG investors want, how to improve valuation",
                "color": C["amber"],
                "description": "Analyses the **market and investor implications** — what ESG analysts will flag, how this affects equity valuation, what institutional investors require, and what the company must do to access green capital markets.",
                "system": """You are a capital markets and ESG investment advisor with 15 years experience covering Indian industrial stocks for FIIs, domestic mutual funds, and ESG-focused sovereign wealth funds. You have led 12 ESG-linked bond issuances and advised on 8 sustainability-linked equity stories.

YOUR JOB: Tell the company exactly how equity markets and institutional investors are assessing this climate risk, what it means for their valuation and cost of capital, and what they must do to access premium ESG capital.

STRUCTURE YOUR RESPONSE AS:

## 📊 Market Pricing of This Risk
How are equity investors likely to discount the valuation given these climate metrics? Translate PD and ECL numbers into EV/EBITDA multiple impact. Be specific: "A PD increase from X% to Y% under Net Zero scenario typically leads institutional ESG analysts to apply a Z% discount to sector multiple."

## 🔍 What ESG Analysts Will Flag in Their Reports
The specific data points that trigger negative ESG analyst notes. For each: the metric, the threshold, the typical analyst language, the index exclusion risk.

## 💹 ESG Index Inclusion / Exclusion Risk
Specific indices (MSCI ESG, FTSE4Good, S&P BSE ESG, NIFTY100 ESG) — which ones this company risks exclusion from and why, based on the BRSR flags and climate risk scores.

## 🎯 What Institutional Investors Want (Prioritised List)
The 5 specific things the 10 largest institutional investors (FIIs + domestic MFs) in this sector are asking companies to do. For each: why they want it, what it signals, how to deliver it.

## 💰 Green Capital Raising Potential
Given these metrics: Can this company issue a green bond? At what premium/discount? What ESG improvements would unlock better pricing? Specific targets to hit.

## 📣 Investor Communication Strategy
What to say at the next earnings call, analyst day, and ESG roadshow. The specific metrics to highlight, the narrative arc, the forward guidance that will be credible.""",
                "user_prefix": "Analyse how equity markets and institutional investors will assess these climate risk metrics for this Indian company. Provide specific valuation impacts and an actionable investor strategy.",
            },

            "🗺️ Strategic Planning Advisor": {
                "icon": "🗺️",
                "subtitle": "5-year business transformation roadmap with milestones and investment decisions",
                "color": C["accent2"],
                "description": "Builds a **concrete 5-year strategic roadmap** that integrates climate risk mitigation with business growth — what to invest in, what to divest, how to restructure the business model to thrive rather than just survive.",
                "system": """You are a strategy advisor combining McKinsey-grade business strategy with deep climate transition expertise. You have led 5-year strategic transformations for 3 Indian steel companies, 2 power utilities, and 1 cement group navigating the energy transition. You understand exactly how to build a business case that wins board approval and delivers shareholder value while managing climate risk.

YOUR JOB: Build a practical 5-year strategic roadmap that turns this climate risk data into a business transformation plan. Not vague strategy — specific decisions, specific investments, specific milestones.

STRUCTURE YOUR RESPONSE AS:

## 🎯 Strategic Verdict
In 3 sentences: Is this company's core business model viable through 2035 under the most likely climate scenario? What is the fundamental strategic choice it faces?

## 🗺️ The 5-Year Transformation Roadmap

### Year 1 (Stabilise): Specific actions to stop the bleeding
- Financial: [Specific decisions with ₹ amounts]
- Operational: [Specific projects to start]
- Governance: [Specific structures to put in place]
- Success metrics: [What Year 1 must achieve — specific numbers]

### Years 2–3 (Reposition): Building the new business model
- Investment decisions: What to build, buy, or partner on
- Business lines to grow vs shrink
- Key hires and capabilities to build
- Milestone: [Specific financial/climate metrics by end of Year 3]

### Years 4–5 (Lead): Competitive differentiation through climate leadership
- Market positioning
- New revenue streams from decarbonisation
- Target: [Specific PD, carbon burden, and EBITDA targets by Year 5]

## ⚔️ Competitive Dynamics
How will climate regulation reshape competitive dynamics in this sector? Who wins, who loses, and what this company must do to be a winner.

## 💡 Hidden Opportunities
2-3 business opportunities created by climate transition that this company is positioned to capture, with revenue potential estimates.

## 🚫 Strategic Risks to Avoid
The 3 most dangerous strategic mistakes companies in this situation make. Be specific about why they are tempting and why they lead to failure.""",
                "user_prefix": "Build a 5-year strategic transformation roadmap for this company based on the climate risk data. Be specific about decisions, investments, and milestones.",
            },
        }

        # ── UI Layout ───────────────────────────────────────────────────
        st.markdown(f"""
        <div style="background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:10px;
                    padding:14px 18px;margin-bottom:16px;">
            <div style="font-size:13px;color:{C['accent2']};font-weight:600;margin-bottom:6px;">
                How to use these advisors
            </div>
            <div style="font-size:12px;color:{C['slate']};line-height:1.7;">
                Each advisor looks at the <strong style="color:{C['white']};">same model results</strong>
                through a completely different lens.
                Run the one most relevant to your immediate need, or run all seven and compare.
                Every response uses your <strong style="color:{C['white']};">actual computed numbers</strong>
                — not generic advice.
            </div>
        </div>""", unsafe_allow_html=True)

        # Advisor selector
        advisor_names = list(ADVISORS.keys())
        selected_advisor = st.radio(
            "Choose your advisor",
            advisor_names,
            horizontal=False,
            label_visibility="collapsed",
        )

        adv = ADVISORS[selected_advisor]

        # Show advisor card
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,{C['bg_mid']},{C['card']});
                    border:1px solid {adv['color']};border-left:5px solid {adv['color']};
                    border-radius:10px;padding:16px 20px;margin:12px 0;">
            <div style="font-size:18px;font-weight:700;color:{adv['color']};margin-bottom:4px;">
                {selected_advisor}
            </div>
            <div style="font-size:12px;color:{C['accent2']};font-weight:600;margin-bottom:8px;">
                {adv['subtitle']}
            </div>
            <div style="font-size:12px;color:{C['slate']};line-height:1.6;">
                {adv['description']}
            </div>
        </div>""", unsafe_allow_html=True)

        # Optional context box
        with st.expander("➕ Add context to improve the response (optional)", expanded=False):
            extra_ctx = st.text_area(
                "Anything the advisor should know that's not in the model data",
                placeholder=(
                    "E.g.: We have a board-approved ₹800 Cr green CAPEX plan starting Q3 2026. "
                    "Our main lender is SBI with a loan renewal due in March 2027. "
                    "We are targeting MSCI ESG inclusion by end 2025. "
                    "Our main competitor JSW Steel has already committed to net zero by 2050."
                ),
                height=100,
                key="ai_extra_ctx",
            )
            focus_area = st.text_input(
                "Specific question or focus area (optional)",
                placeholder="E.g.: Focus specifically on what to do about the CAPEX gap and how to finance it.",
                key="ai_focus",
            )

        col_btn1, col_btn2 = st.columns([2, 1])
        with col_btn1:
            run_btn = st.button(
                f"🧠 Ask the {adv['icon']} {selected_advisor.split(' ', 1)[1]}",
                type="primary",
                width="stretch",
            )
        with col_btn2:
            clear_btn = st.button("🗑️ Clear Output", width="stretch")

        if clear_btn:
            if "ai_outputs" in st.session_state:
                st.session_state["ai_outputs"] = {}
            st.rerun()

        # ── Run the selected advisor ────────────────────────────────────
        if run_btn:
            payload = _build_rich_payload()

            # Add extra context
            extra = st.session_state.get("ai_extra_ctx", "").strip()
            focus = st.session_state.get("ai_focus", "").strip()
            if extra:
                payload["additional_context"] = extra
            if focus:
                payload["specific_focus"] = focus

            messages = _build_ai_messages(selected_advisor, adv, payload)

            with st.spinner(f"Your {selected_advisor} is analysing the compact risk brief..."):
                try:
                    client = _get_groq_client()
                    stream = client.chat.completions.create(
                        model="llama-3.1-8b-instant",
                        messages=messages,
                        temperature=0.25,
                        max_tokens=900,
                        stream=True,
                    )
                    output_parts = []
                    stream_box = st.empty()
                    for chunk in stream:
                        delta = getattr(chunk.choices[0].delta, "content", None)
                        if delta:
                            output_parts.append(delta)
                            stream_box.markdown("".join(output_parts))

                    output = "".join(output_parts).strip()
                    if "ai_outputs" not in st.session_state:
                        st.session_state["ai_outputs"] = {}
                    st.session_state["ai_outputs"][selected_advisor] = output
                    st.session_state["ai_outputs"]["_last_advisor"] = selected_advisor
                except Exception as e:
                    err = str(e)
                    if "413" in err or "Request too large" in err or "tokens" in err:
                        st.error("AI error: Groq token limit exceeded even after compression. Reduce optional context/focus text and retry, or use a higher Groq tier/model.")
                    else:
                        st.error(f"AI error: {e}. Check that GROQ_API_KEY is set in Streamlit Secrets or your environment.")

        # ── Display output ──────────────────────────────────────────────
        ai_outs = st.session_state.get("ai_outputs", {})
        last_adv = ai_outs.get("_last_advisor", selected_advisor)
        output_text = ai_outs.get(selected_advisor, "")

        if output_text:
            # Output card
            st.markdown(f"""
            <div style="background:{C['bg_dark']};border:1px solid {ADVISORS[selected_advisor]['color']};
                        border-radius:10px;padding:4px 4px 4px 4px;margin-top:8px;">
                <div style="background:{C['card']};border-radius:8px;padding:20px 24px;">
            """, unsafe_allow_html=True)
            st.markdown(output_text)
            st.markdown("</div></div>", unsafe_allow_html=True)

            # Copy / download button
            st.download_button(
                label="⬇️ Download this analysis (.txt)",
                data=output_text,
                file_name=f"ICCRE_{company_name.replace(' ','_')}_{selected_advisor.split(' ',1)[1].replace(' ','_')[:30]}_{REPORTING_YEAR}.txt",
                mime="text/plain",
                width="content",
            )

        elif not run_btn:
            st.markdown(f"""
            <div style="background:{C['card']};border:1px dashed {C['bg_ocean']};border-radius:10px;
                        padding:32px;text-align:center;margin-top:8px;">
                <div style="font-size:36px;margin-bottom:8px;">{adv['icon']}</div>
                <div style="font-size:14px;color:{C['slate']};">
                    Click the button above to get your analysis from the {selected_advisor}.
                </div>
            </div>""", unsafe_allow_html=True)

        # ── Previous outputs from other advisors ────────────────────────
        other_outputs = {k: v for k, v in ai_outs.items()
                         if k != selected_advisor and k != "_last_advisor" and v}
        if other_outputs:
            st.divider()
            st.markdown(f"<div style='font-size:12px;color:{C['slate']};font-weight:600;"
                        f"text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px;'>"
                        f"Previously generated analyses</div>", unsafe_allow_html=True)
            for prev_name, prev_text in other_outputs.items():
                prev_adv = ADVISORS.get(prev_name, {})
                with st.expander(f"{prev_name} — click to expand"):
                    st.markdown(prev_text)
                    st.download_button(
                        label=f"⬇️ Download {prev_name.split(' ',1)[1][:25]}",
                        data=prev_text,
                        file_name=f"ICCRE_{company_name.replace(' ','_')}_{prev_name.split(' ',1)[1].replace(' ','_')[:30]}_{REPORTING_YEAR}.txt",
                        mime="text/plain",
                        key=f"dl_{prev_name[:20]}",
                    )

    # ============================================================
    # TAB 8 — METHODOLOGY
    # ============================================================
    # ============================================================
    # TAB 8 — METHODOLOGY  (auditor-appropriate, IP-protected)
    # ============================================================
with methodology_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>📖 Model Methodology</h2>", unsafe_allow_html=True)
    st.caption(f"ICCRE v{MODEL_VERSION} · {ENGINE_BUILD} · Build {MODEL_BUILD_DATE} · {NGFS_DATA_VERSION}")

    st.info(
        "This section documents the model framework, key equations, regulatory alignment, "
        "and known limitations for disclosure to model validators, auditors, and regulators. "
        "Proprietary calibration data, sector-specific coefficients, and implementation details "
        "are available under NDA on request."
    )

    with st.expander("📐 Model Purpose & Scope", expanded=True):
        st.markdown("""
        **Purpose:** Quantify the financial impact of climate change on corporate credit risk under
        forward-looking scenarios aligned with the Network for Greening the Financial System (NGFS).

        **Primary outputs:** Probability of Default (PD), Loss Given Default (LGD), Expected Credit
        Loss (ECL), Debt Service Coverage Ratio (DSCR), and capital adequacy metrics under three
        NGFS climate scenarios.

        **Regulatory frameworks:** ISSB IFRS S2 (§14–16), RBI Climate Risk Management Guidelines,
        SEBI BRSR Core, Basel III capital adequacy, ICAAP.

        **Covered sectors:** Steel · Power · Cement · Oil & Gas · Manufacturing

        **Scenario coverage:** Current Policies (≈2.7°C) · NDCs (≈2.1°C) · Net Zero 2050 (≈1.5°C)

        **Data vintage:** NGFS Phase III (2023). Short-term (2025–2030) scenarios available in Phase V.
        """)

    with st.expander("🔗 Climate-to-Credit Transmission Chain"):
        st.markdown("""
        The model follows a five-stage financial transmission chain:

        **Stage 1 — Climate Scenario:** NGFS carbon price, GDP, and temperature pathways are
        extracted for the selected scenario and projection year.

        **Stage 2 — Revenue Impact:** Carbon costs and GDP shocks are transmitted to company
        revenue through three independent channels: GDP macro sensitivity, demand elasticity
        (volume reduction), and price elasticity (margin compression). A chronic physical
        temperature effect also reduces revenue proportionally to warming above baseline.

        **Stage 3 — EBITDA & DSCR:** Net carbon cost is deducted from EBITDA. A governance
        quality adjustment reflects the management quality of the transition response.
        A climate credit spread shock is applied to interest costs. DSCR is computed on the
        stressed EBITDA and stressed interest.

        **Stage 4 — Credit Model:** A structural logit PD model maps DSCR stress and carbon
        burden directly to change in default probability. The model is derived from Merton's
        structural credit framework adapted for climate-specific risk factors.

        **Stage 5 — Integration & ECL:** Transition and physical PDs are combined using a
        Gaussian copula accounting for cross-risk correlation. BRSR operational flags add a
        basis-point uplift. Final ECL = PD × LGD × EAD (IFRS 9 compliant).
        """)

    with st.expander("📐 Core Credit Model Equations"):
        st.latex(r"\text{logit}(PD_t) = \text{logit}(PD_0)"
                 r"+ \alpha_{DSCR}\cdot\underbrace{(1.5 - DSCR_t)}_{\text{signed gap}}"
                 r"+ \beta_{credit}\cdot CB_t")
        st.latex(r"PD_t = \sigma\bigl(\text{logit}(PD_t)\bigr), \quad PD_t \in [PD_{floor},\, PD_{cap}]")
        st.markdown("""
        Where:
        - **α_DSCR** — sector-specific DSCR sensitivity parameter
        - **DSCR gap** — signed deviation from 1.5× covenant threshold (positive = stress, negative = relief)
        - **β_credit** — sector-specific carbon burden sensitivity parameter
        - **CB_t** — carbon burden (net carbon cost / revenue) at time t
        - **PD_floor = 0.05%**, **PD_cap = 35%** (regulatory prudence bounds)

        The signed DSCR gap allows credit quality *improvement* for companies with DSCR > 1.5×,
        consistent with structural credit theory (Merton, 1974).
        """)

    with st.expander("🌡️ Physical Risk Damage Function (v1.2)"):
        st.latex(r"\text{Damage Multiplier} = 1 + 0.20 \cdot \Delta T + 0.04 \cdot \Delta T^2")
        st.latex(r"\sigma_{\Delta T} = 0.08 \cdot \Delta T")
        st.latex(r"P_{10} = \text{mult} - 1.645\,\sigma, \quad P_{90} = \text{mult} + 1.645\,\sigma")
        st.markdown("""
        The quadratic damage function is calibrated to economic damage literature consistent
        with IPCC AR6 findings. The linear term (0.20) captures proportional temperature-income
        effects; the quadratic term (0.04) captures accelerating damages at higher warming levels.
        Temperature anomaly ΔT is extracted from the same NGFS scenario pathway used by the
        transition engine, ensuring scenario consistency.

        Uncertainty bounds (P10–P90) grow with temperature anomaly, reflecting the greater
        model uncertainty at higher warming levels.
        """)

    with st.expander("🔀 Gaussian Copula Integration"):
        st.latex(r"PD_{copula} = \Phi_2\!\left(\Phi^{-1}(PD_T),\;\Phi^{-1}(PD_P),\;\rho\right)")
        st.latex(r"PD_{integrated} = \min\!\left(PD_{copula} + \Delta PD_{BRSR},\; PD_{cap}\right)")
        st.markdown("""
        The bivariate Gaussian copula captures the statistical dependence between transition risk
        and physical risk without assuming they are independent or perfectly correlated.
        Sector-specific correlations ρ range from 0.25 (Manufacturing) to 0.40 (Oil & Gas),
        reflecting the degree to which carbon-intensive sectors face both types of climate risk
        simultaneously.

        BRSR operational risk is additive rather than copula-integrated because it represents a
        distinct third category of climate exposure (regulatory and operational) rather than a
        correlated extension of credit or asset risk. The BRSR uplift is capped at 150 basis points.
        """)

    with st.expander("📘 BRSR → Credit Risk Linkage"):
        st.markdown("""
        BRSR Core operational flags are mapped to basis-point PD adjustments that represent
        incremental credit risk from operational climate exposures. This linkage captures risks
        outside the carbon-price transmission chain, including:

        - Water scarcity exposure in high-stress regions
        - Energy cost escalation from high fossil energy dependence
        - Governance deficiencies that reduce management quality of climate transition
        - Regulatory non-compliance risk under SEBI BRSR Core requirements

        The PD uplift is conservative: each flag is treated independently and the total is capped.
        """)

    with st.expander("🔄 v1.1 → v1.2 Model Corrections"):
        st.markdown("""
        Seven corrections were applied in v1.1. All are preserved in v1.2.

        | Fix | Nature | Financial Impact |
        |-----|--------|-----------------|
        | FIX-01 | β_carbon transition/credit separation | Prevents slider from overriding structural model |
        | FIX-02 | Physical double-count removed | Reduces PD over-estimation in high-warming scenarios |
        | FIX-03 | Revenue denominator stabilised | Removes order-of-operations sensitivity |
        | FIX-04 | Signed DSCR gap | Credit relief for strong companies (DSCR > 1.5×) |
        | FIX-05 | Two-step CPI conversion | Material for high-carbon-price scenarios |
        | FIX-06 | GDP single transmission path | Removes systematic over-estimation in disorderly transition |
        | FIX-07 | MC GDP sign consistency | MC outputs now match deterministic results |
        """)

    with st.expander("⚠️ Model Limitations & Assumptions"):
        st.markdown("""
        The following limitations apply and should be disclosed in any regulatory submission:

        1. **Calibration:** Sector parameters are calibrated to published literature. Borrower-specific
           historical default data has not been used unless the Calibration tab has been run with
           uploaded observed PD data.

        2. **Scope:** Five sectors are covered. Real estate, agriculture, infrastructure, and NBFC
           are not currently modelled.

        3. **Scenario horizon:** NGFS Phase III long-term scenarios are used. Short-term (2025–2030)
           horizon analysis requires Phase V data (available separately).

        4. **Static balance sheet:** The model does not account for dynamic feedback between climate
           stress, asset values, and new lending capacity over multi-year horizons.

        5. **Governance coefficient:** The 15% EBITDA governance sensitivity is expert-calibrated.
           Users should adjust via the governance sliders based on company-specific evidence.

        6. **Physical hazard data:** Results depend on the quality and resolution of GIS input data.
           Results should be treated as indicative for assets with limited local hazard data.

        7. **BRSR coefficients:** BRSR-to-PD spread linkages are analytically derived, not
           backtested against observed defaults. They represent regulatory risk premia, not
           historically observed credit loss differentials.
        """)

    with st.expander("📋 Scenario Registry"):
        st.dataframe(
            pd.DataFrame(SCENARIO_REGISTRY).T
            .style.set_properties(**{"background-color": C["bg_dark"], "color": C["text"]}),
            width="stretch"
        )

    with st.expander("🔢 Model Governance Record"):
        gov_df = pd.DataFrame([{
            "Parameter": "Model Version",       "Value": MODEL_VERSION},
            {"Parameter": "Parameter Version",  "Value": PARAMETER_VERSION},
            {"Parameter": "NGFS Data Version",  "Value": NGFS_DATA_VERSION},
            {"Parameter": "Engine Name",         "Value": ENGINE_BUILD},
            {"Parameter": "Build Date",          "Value": MODEL_BUILD_DATE},
            {"Parameter": "Run Log Location",    "Value": "logs/run_log.csv"},
            {"Parameter": "Audit Trail",         "Value": "SHA-256 run hash per execution"},
            {"Parameter": "Regulatory Basis",    "Value": "ISSB S2 · RBI 2024 Draft · SEBI BRSR Core · NGFS PIII"},
        ])
        st.dataframe(gov_df.style.set_properties(**{"background-color": C["bg_dark"], "color": C["text"]}),
            width="stretch", hide_index=True)

# ============================================================
# TAB 9 — VALIDATION  (improved)
# ============================================================
with validation_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>🔬 Model Validation</h2>", unsafe_allow_html=True)
    st.caption("Backtest model-implied PD against observed default rates · Bias detection · Error diagnostics")

    if not st.session_state.get("transition_ran", False):
        st.info("Run Transition Risk Engine first.")
    else:

        # ── GUIDANCE ──
        with st.expander("📋 How to use this tab", expanded=False):
            st.markdown("""
            **Purpose:** Test whether the model's projected PD matches historically observed default rates
            for comparable borrowers or sector benchmarks.

            **What you need:** A CSV file with at minimum two columns:
            - `Year` — the calendar year (e.g. 2019, 2020, 2021)
            - `Observed_PD` — the observed default rate for that year (decimal, e.g. 0.025 = 2.5%)

            **Optional columns:**
            - `Observed_ECL` — observed credit loss (₹ Cr) — used for ECL-level validation
            - `Observed_DSCR` — observed DSCR — used for transmission chain validation

            **Data sources (India):** CRISIL default studies · RBI Trend & Progress report (sector NPAs)
            · IBA credit data · Prowess/CMIE database

            **Interpretation:**
            - RMSE < 1% — Strong calibration, no action needed
            - RMSE 1–3% — Moderate error, recalibration recommended (use Calibration tab)
            - RMSE > 3% — High error, structural review required
            - Positive Bias — Model systematically over-estimates PD (conservative)
            - Negative Bias — Model systematically under-estimates PD (optimistic — investigate)
            """)

        # ── FILE UPLOAD ──
        st.subheader("📂 Upload Historical Observations")
        ufile = st.file_uploader("CSV file (Year, Observed_PD required)", type=["csv"])

        if ufile:
            df_hist = pd.read_csv(ufile)
            st.session_state["historical_data"] = df_hist

            if not {"Year", "Observed_PD"}.issubset(df_hist.columns):
                st.error("CSV must contain: `Year`, `Observed_PD`")
                st.stop()

            df_hist["Year"] = df_hist["Year"].astype(int)
            df_hist["Observed_PD"] = pd.to_numeric(df_hist["Observed_PD"], errors="coerce")
            df_hist = df_hist.dropna(subset=["Observed_PD"])

            # Select scenario to validate against
            df_trans_v = st.session_state.get("df_transition")
            scens_v = sorted(df_trans_v["Scenario"].unique())
            val_scen = st.selectbox("Validate against scenario", scens_v, key="val_scen")

            df_mod_agg = (df_trans_v[df_trans_v["Scenario"] == val_scen]
                .groupby("Year")["PD_Transition"].max().reset_index()
                .rename(columns={"PD_Transition": "Model_PD"}))
            df_cmp = df_hist.merge(df_mod_agg, on="Year", how="inner")

            if df_cmp.empty:
                st.warning("No overlapping years between uploaded data and model output.")
                st.stop()

            # ── METRICS ──
            df_cmp["Error"]    = df_cmp["Model_PD"] - df_cmp["Observed_PD"]
            df_cmp["AbsError"] = df_cmp["Error"].abs()
            df_cmp["RelError"] = df_cmp["AbsError"] / df_cmp["Observed_PD"].replace(0, np.nan)
            mae   = df_cmp["AbsError"].mean()
            rmse  = np.sqrt((df_cmp["Error"]**2).mean())
            bias  = df_cmp["Error"].mean()
            mape  = df_cmp["RelError"].mean() * 100

            # Signal
            if rmse < 0.01:   val_signal = "✅ Strong";   sig_color = C["mint"]
            elif rmse < 0.03: val_signal = "⚠️ Moderate"; sig_color = C["amber"]
            else:             val_signal = "❌ High Error"; sig_color = C["coral"]

            bias_signal = ("Conservative (over-estimates risk)" if bias > 0.005
                           else "Optimistic (under-estimates risk)" if bias < -0.005
                           else "Unbiased")

            # KPI row
            v1, v2, v3, v4, v5 = st.columns(5)
            v1.metric("Calibration Signal", val_signal)
            v2.metric("MAE",  f"{mae:.4f}",  help="Mean Absolute Error")
            v3.metric("RMSE", f"{rmse:.4f}", help="Root Mean Squared Error — primary metric")
            v4.metric("Bias", f"{bias:+.4f}", help="Positive = model over-estimates PD")
            v5.metric("MAPE", f"{mape:.1f}%", help="Mean Absolute Percentage Error")

            st.markdown(
                f"<div style='background:{C['card']};border-left:3px solid {sig_color};"
                f"border-radius:6px;padding:10px 14px;margin:8px 0;font-size:13px;'>"
                f"<strong style='color:{sig_color};'>{val_signal}</strong> &nbsp;·&nbsp; "
                f"Bias direction: <strong>{bias_signal}</strong></div>",
                unsafe_allow_html=True
            )

            # ── CHARTS ──
            st.subheader("📈 Model vs Observed PD")
            fig_v1 = make_subplots(rows=1, cols=2,
                subplot_titles=("PD Trajectory: Model vs Observed", "Prediction Error by Year"))
            # Trajectory
            fig_v1.add_trace(go.Scatter(x=df_cmp["Year"], y=df_cmp["Observed_PD"],
                mode="lines+markers", name="Observed PD",
                line=dict(color=C["mint"], width=2.5), marker=dict(size=8, symbol="diamond")), row=1, col=1)
            fig_v1.add_trace(go.Scatter(x=df_cmp["Year"], y=df_cmp["Model_PD"],
                mode="lines+markers", name="Model PD",
                line=dict(color=C["accent2"], width=2.5, dash="dash"), marker=dict(size=7)), row=1, col=1)
            # Error bars
            bar_colors = [C["coral"] if e > 0 else C["mint"] for e in df_cmp["Error"]]
            fig_v1.add_trace(go.Bar(x=df_cmp["Year"], y=df_cmp["Error"] * 100,
                name="Error (pp)", marker_color=bar_colors, opacity=0.8,
                hovertemplate="Year: %{x}<br>Error: %{y:.2f} pp<extra></extra>"), row=1, col=2)
            fig_v1.add_hline(y=0, line_dash="dot", line_color=C["slate"], row=1, col=2)
            fig_v1.update_layout(**_chart_layout("", 320, legend_override=dict(orientation="h", y=1.1)))
            fig_v1.update_yaxes(tickformat=".1%", row=1, col=1)
            fig_v1.update_yaxes(ticksuffix=" pp", title="Error (percentage points)", row=1, col=2)
            _ax_style(fig_v1, rows=1, cols=2)
            st.plotly_chart(fig_v1, width="stretch")

            # Scatter: predicted vs actual
            fig_v2 = go.Figure()
            fig_v2.add_trace(go.Scatter(
                x=df_cmp["Observed_PD"], y=df_cmp["Model_PD"],
                mode="markers+text", text=df_cmp["Year"].astype(str), textposition="top center",
                marker=dict(color=C["accent2"], size=12, line=dict(color=C["white"], width=1)),
                name="Year",
                hovertemplate="Observed: %{x:.2%}<br>Model: %{y:.2%}<extra></extra>",
            ))
            # Perfect fit line
            mn = min(df_cmp[["Observed_PD","Model_PD"]].min())
            mx = max(df_cmp[["Observed_PD","Model_PD"]].max())
            fig_v2.add_trace(go.Scatter(x=[mn, mx], y=[mn, mx],
                mode="lines", name="Perfect fit",
                line=dict(color=C["slate"], dash="dot", width=1.5)))
            fig_v2.update_layout(**_chart_layout("Predicted vs Observed PD (perfect fit = dotted line)", 300))
            fig_v2.update_xaxes(title="Observed PD", tickformat=".1%")
            fig_v2.update_yaxes(title="Model PD", tickformat=".1%")
            _ax_style(fig_v2)
            st.plotly_chart(fig_v2, width="stretch")

            # Full comparison table
            st.subheader("📋 Year-by-Year Validation Table")
            df_display = df_cmp[["Year","Observed_PD","Model_PD","Error","AbsError","RelError"]].copy()
            df_display.columns = ["Year","Observed PD","Model PD","Error","Abs Error","Rel Error"]
            st.dataframe(
                df_display.style
                    .format({"Observed PD":"{:.3%}","Model PD":"{:.3%}",
                             "Error":"{:+.3%}","Abs Error":"{:.3%}","Rel Error":"{:.1%}"})
                    .background_gradient(subset=["Abs Error"], cmap="Reds")
                    .applymap(lambda v: f"color:{C['coral']}" if isinstance(v,float) and v>0.005
                              else f"color:{C['mint']}" if isinstance(v,float) and v<-0.005 else "",
                              subset=["Error"]),
                width="stretch", hide_index=True
            )

            # ECL validation if provided
            if "Observed_ECL" in df_hist.columns:
                st.subheader("💰 ECL-Level Validation")
                df_ecl = df_hist.merge(
                    df_trans_v[df_trans_v["Scenario"]==val_scen].groupby("Year")["ECL_Transition"].max().reset_index(),
                    on="Year", how="inner"
                )
                if not df_ecl.empty:
                    ecl_rmse = np.sqrt(((df_ecl["ECL_Transition"] - df_ecl["Observed_ECL"])**2).mean())
                    ecl_bias = (df_ecl["ECL_Transition"] - df_ecl["Observed_ECL"]).mean()
                    ec1, ec2 = st.columns(2)
                    ec1.metric("ECL RMSE (₹ Cr)", f"{ecl_rmse:.2f}")
                    ec2.metric("ECL Bias (₹ Cr)", f"{ecl_bias:+.2f}")

            st.session_state["validation_results"] = {"MAE":mae,"RMSE":rmse,"Bias":bias,"MAPE":mape,"n_years":len(df_cmp)}
            log_model_run("Backtest", {"scenario":val_scen,"MAE":mae,"RMSE":rmse,"Bias":bias,"n_years":len(df_cmp)})
            st.success(f"✅ Validation complete · {len(df_cmp)} overlapping years · Scenario: {val_scen}")

            st.caption(
                "**Next step:** If RMSE > 1%, go to the **Calibration** tab to find optimised α and β_credit "
                "parameters that minimise error against this observed data."
            )
        else:
            st.info(
                "Upload a CSV with `Year` and `Observed_PD` columns. "
                "You can use sector-level NPA ratios from RBI Trend & Progress reports as a proxy "
                "if company-specific default histories are unavailable."
            )

    # ============================================================
    # TAB 10 — CALIBRATION  (improved user guidance)
    # ============================================================
with calibration_tab:
    st.markdown(f"<h2 style='color:{C['white']}'>⚙️ Parameter Calibration</h2>", unsafe_allow_html=True)
    st.caption("Optimise α (DSCR sensitivity) and β_credit (carbon burden sensitivity) to minimise error against observed default rates")

    if not st.session_state.get("transition_ran", False):
        st.info("Run Transition Risk Engine first.")
    else:

        hdata = st.session_state.get("historical_data")

        # --- FIX: strict guard + stop execution ---
        if hdata is None or not isinstance(hdata, pd.DataFrame):
            st.info(
                "Upload historical PD data in the **Validation** tab first. "
                "The Calibration tab uses that data to find optimal parameters."
            )
            st.stop()

        if "Year" not in hdata.columns:
            st.error("❌ 'Year' column missing in historical data.")
            st.stop()

        # Clean + enforce numeric year
        hdata["Year"] = pd.to_numeric(hdata["Year"], errors="coerce")
        hdata = hdata.dropna(subset=["Year"])
        # --- END FIX ---

        with st.expander("📋 What does this tab do?", expanded=True):
            st.markdown(f"""
            **Purpose:** Find the values of α (DSCR sensitivity) and β_credit (carbon burden sensitivity)
            that minimise the Root Mean Squared Error between model PD and your observed default rates.

            **How it works:**
            - The model tests {10*10:,} combinations of α and β_credit values on a grid
            - For each combination, it computes a model PD using your actual DSCR and carbon burden data
            - The combination with the lowest RMSE vs your observed PDs is recommended

            **What to do with the results:**
            - If RMSE improvement > 10%: update α and β_credit in the sidebar sliders or sector parameters
            - If improvement is 3–10%: consider updating; model is reasonably well-calibrated
            - If improvement < 3%: the default literature-based parameters are performing well for your data

            **Important:** Calibrated parameters are specific to your dataset. Do not apply parameters
            calibrated on one sector's data to a different sector without justification.
            """)

        df_trans_c = st.session_state.get("df_transition")
        scens_c = sorted(df_trans_c["Scenario"].unique())
        cal_scen = st.selectbox("Calibrate against scenario", scens_c, key="cal_scen")

        df_mc_agg = (df_trans_c[df_trans_c["Scenario"]==cal_scen]
            .groupby("Year").agg({"PD_Transition":"max","DSCR":"min","Carbon_Burden":"max"}).reset_index())
        hdata["Year"] = hdata["Year"].astype(int)
        df_cc = hdata.merge(df_mc_agg, on="Year", how="inner")

        if df_cc.empty:
            st.warning("No overlapping years between calibration data and model output.")
            st.stop()

        # Grid search parameters
        st.subheader("🔧 Search Grid Configuration")
        sg1, sg2 = st.columns(2)
        a_min, a_max = sg1.slider("α range (DSCR sensitivity)", 0.1, 2.0, (0.1, 1.5), step=0.05, key="cal_arange")
        b_min, b_max = sg2.slider("β_credit range (carbon sensitivity)", 0.1, 2.5, (0.1, 2.0), step=0.05, key="cal_brange")
        n_steps = st.slider("Grid resolution (steps per axis)", 5, 20, 10, key="cal_steps",
            help="Higher = more thorough search but slower. 10×10 = 100 combinations.")

        alpha_vals = np.linspace(a_min, a_max, n_steps)
        beta_vals  = np.linspace(b_min, b_max, n_steps)

        st.caption(f"Grid: {len(alpha_vals)} × {len(beta_vals)} = **{len(alpha_vals)*len(beta_vals):,} combinations** to test")

        if st.button("▶ Run Grid Search Calibration", type="primary"):
            with st.spinner(f"Searching {len(alpha_vals)*len(beta_vals):,} parameter combinations..."):
                best_rmse = 999.0
                best_params = (alpha_vals[len(alpha_vals)//2], beta_vals[len(beta_vals)//2])
                base_pd_c = df_cc["Observed_PD"].iloc[0]
                rmse_grid = np.zeros((len(alpha_vals), len(beta_vals)))

                for ai, a in enumerate(alpha_vals):
                    for bi, b in enumerate(beta_vals):
                        dg = np.clip(1.5 - df_cc["DSCR"], -4.0, 6.0)
                        pd_e = np.clip(sigmoid(logit(base_pd_c) + a*dg + b*df_cc["Carbon_Burden"]), PD_FLOOR, PD_CAP)
                        r = float(np.sqrt(np.mean((pd_e - df_cc["Observed_PD"])**2)))
                        rmse_grid[ai, bi] = r
                        if r < best_rmse:
                            best_rmse = r
                            best_params = (a, b)

            orig_rmse = st.session_state.get("validation_results", {}).get("RMSE", best_rmse * 1.1)
            improvement = ((orig_rmse - best_rmse) / max(orig_rmse, 1e-8)) * 100

            # Results
            if improvement > 10:   sig = "✅ Material improvement"; sig_c = C["mint"]
            elif improvement > 3:  sig = "⚠️ Moderate improvement"; sig_c = C["amber"]
            else:                  sig = "ℹ️ Marginal improvement"; sig_c = C["accent2"]

            st.markdown(f"<div style='background:{C['card']};border-left:3px solid {sig_c};border-radius:6px;padding:10px 14px;margin:8px 0;font-size:13px;color:{sig_c};font-weight:600;'>{sig}</div>", unsafe_allow_html=True)

            cr1, cr2, cr3, cr4 = st.columns(4)
            cr1.metric("Optimal α", f"{best_params[0]:.3f}", help="Update DSCR sensitivity in sector parameters")
            cr2.metric("Optimal β_credit", f"{best_params[1]:.3f}", help="Update carbon burden sensitivity in sector parameters")
            cr3.metric("Calibrated RMSE", f"{best_rmse:.4f}")
            cr4.metric("RMSE Improvement", f"{improvement:.1f}%", delta_color="normal")

            # RMSE surface heatmap
            st.subheader("📊 RMSE Surface — Parameter Search Grid")
            fig_grid = go.Figure(go.Heatmap(
                z=rmse_grid * 100,
                x=[f"{b:.2f}" for b in beta_vals],
                y=[f"{a:.2f}" for a in alpha_vals],
                colorscale=[[0.0, C["mint"]], [0.4, C["amber"]], [0.7, C["coral"]], [1.0, "#7F0000"]],
                colorbar=dict(title="RMSE (%)"),
                hovertemplate="α=%{y}<br>β=%{x}<br>RMSE=%{z:.3f}%<extra></extra>",
            ))
            # Mark optimum
            fig_grid.add_trace(go.Scatter(
                x=[f"{best_params[1]:.2f}"], y=[f"{best_params[0]:.2f}"],
                mode="markers", name="Optimum",
                marker=dict(color=C["white"], size=14, symbol="star",
                            line=dict(color=C["accent2"], width=2))
            ))
            fig_grid.update_layout(**_chart_layout("RMSE (%) by α × β_credit — lower is better", 380,
                legend_override=dict(orientation="h", y=1.05)))
            fig_grid.update_xaxes(title="β_credit"); fig_grid.update_yaxes(title="α (DSCR sensitivity)")
            st.plotly_chart(fig_grid, width="stretch")

            # Calibrated vs uncalibrated PD comparison
            dg_cal = np.clip(1.5 - df_cc["DSCR"], -4.0, 6.0)
            pd_cal = np.clip(sigmoid(logit(base_pd_c) + best_params[0]*dg_cal + best_params[1]*df_cc["Carbon_Burden"]), PD_FLOOR, PD_CAP)
            df_cc["PD_Calibrated"] = pd_cal

            fig_cal = go.Figure()
            fig_cal.add_trace(go.Scatter(x=df_cc["Year"], y=df_cc["Observed_PD"],
                mode="lines+markers", name="Observed",
                line=dict(color=C["mint"], width=2.5), marker=dict(size=9, symbol="diamond")))
            fig_cal.add_trace(go.Scatter(x=df_cc["Year"], y=df_cc["Model_PD"] if "Model_PD" in df_cc.columns else df_cc["PD_Transition"],
                mode="lines+markers", name="Original Model",
                line=dict(color=C["coral"], dash="dash", width=2), marker=dict(size=7)))
            fig_cal.add_trace(go.Scatter(x=df_cc["Year"], y=df_cc["PD_Calibrated"],
                mode="lines+markers", name="Calibrated Model",
                line=dict(color=C["accent2"], width=2.5), marker=dict(size=7)))
            fig_cal.update_layout(**_chart_layout("Calibrated vs Original Model vs Observed PD", 300))
            fig_cal.update_yaxes(tickformat=".1%")
            _ax_style(fig_cal)
            st.plotly_chart(fig_cal, width="stretch")

            # Action guidance
            st.subheader("📌 Recommended Actions")
            st.markdown(f"""
            Based on calibration results:

            | Parameter | Current (Sector Default) | Calibrated | Change |
            |-----------|--------------------------|------------|--------|
            | α (DSCR sensitivity) | {alpha_dscr:.3f} | **{best_params[0]:.3f}** | {best_params[0]-alpha_dscr:+.3f} |
            | β_credit (Carbon burden) | {beta_carbon_credit:.3f} | **{best_params[1]:.3f}** | {best_params[1]-beta_carbon_credit:+.3f} |

            {'✅ **Calibration recommended:** RMSE improves by ' + f'{improvement:.1f}%.' + ' Update sector parameters and re-run the Transition Risk Engine.' if improvement > 3 else 'ℹ️ **No action needed:** Default parameters are performing well for your data.'}
            """)

            st.session_state["calibrated_params"] = {"alpha": best_params[0], "beta_credit": best_params[1], "rmse": best_rmse, "improvement_pct": improvement}
            log_model_run("Calibration", {"scenario": cal_scen, "alpha": best_params[0], "beta": best_params[1], "rmse": best_rmse, "improvement_pct": improvement})

# ============================================================
# GET ACCESS TAB — Phase 1 Lead Capture & Product Info
# ============================================================
with access_tab:
    import urllib.parse

    st.markdown(f"<h2 style='color:{C['white']};margin-bottom:4px;'>🚀 Get Full Access to ICCRE</h2>",
                unsafe_allow_html=True)
    st.caption("Free pilot · Professional · Enterprise — Find the right tier for your team")

    # ── Value banner ─────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{C['bg_mid']},{C['card']});
                border:1px solid {C['accent']};border-radius:12px;
                padding:24px 28px;margin-bottom:20px;">
      <div style="font-size:18px;font-weight:700;color:{C['white']};margin-bottom:6px;">
        Why teams choose ICCRE
      </div>
      <div style="display:flex;gap:24px;flex-wrap:wrap;margin-top:12px;">
        <div style="flex:1;min-width:180px;">
          <div style="font-size:24px;font-weight:700;color:{C['accent2']};">20 min</div>
          <div style="font-size:11px;color:{C['slate']};">Full NGFS scenario analysis<br>vs 8 weeks with consultants</div>
        </div>
        <div style="flex:1;min-width:180px;">
          <div style="font-size:24px;font-weight:700;color:{C['mint']};">₹0</div>
          <div style="font-size:11px;color:{C['slate']};">Entry cost vs ₹40–80L<br>for a Big-4 engagement</div>
        </div>
        <div style="flex:1;min-width:180px;">
          <div style="font-size:24px;font-weight:700;color:{C['amber']};">100%</div>
          <div style="font-size:11px;color:{C['slate']};">India-native: INR, BRSR Core,<br>RBI 2024, NGFS aligned</div>
        </div>
        <div style="flex:1;min-width:180px;">
          <div style="font-size:24px;font-weight:700;color:{C['coral']};">0</div>
          <div style="font-size:11px;color:{C['slate']};">Direct Indian competitors<br>in quantitative credit risk</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Pricing tiers ─────────────────────────────────────────
    st.markdown(f"<h3 style='color:{C['white']};'>Access Tiers</h3>", unsafe_allow_html=True)

    tier_col1, tier_col2, tier_col3 = st.columns(3)

    with tier_col1:
        st.markdown(f"""
        <div style="background:{C['card']};border:1px solid {C['mint']};border-radius:10px;
                    padding:20px;height:340px;position:relative;">
          <div style="font-size:16px;font-weight:700;color:{C['mint']};margin-bottom:4px;">Free Tier</div>
          <div style="font-size:28px;font-weight:800;color:{C['white']};margin-bottom:2px;">₹ 0</div>
          <div style="font-size:11px;color:{C['slate']};margin-bottom:14px;">Forever free · Single user</div>
          <div style="font-size:11px;color:{C['off_white']};line-height:2.0;">
            ✓ All 11 analytics modules<br>
            ✓ All 3 NGFS scenarios<br>
            ✓ Excel + JSON export<br>
            ✓ AI narrative (Groq key needed)<br>
            ✓ Methodology documentation<br>
            ✓ No time limit
          </div>
          <div style="position:absolute;bottom:16px;left:16px;right:16px;
                      font-size:10px;color:{C['slate']};text-align:center;">
            Currently running on this tier
          </div>
        </div>
        """, unsafe_allow_html=True)

    with tier_col2:
        st.markdown(f"""
        <div style="background:{C['card']};border:2px solid {C['amber']};border-radius:10px;
                    padding:20px;height:340px;position:relative;">
          <div style="position:absolute;top:-12px;left:50%;transform:translateX(-50%);
                      background:{C['amber']};color:{C['bg_dark']};font-size:9px;font-weight:700;
                      padding:3px 12px;border-radius:12px;letter-spacing:.06em;white-space:nowrap;">
            MOST POPULAR
          </div>
          <div style="font-size:16px;font-weight:700;color:{C['amber']};margin-bottom:4px;">Professional</div>
          <div style="font-size:28px;font-weight:800;color:{C['white']};margin-bottom:2px;">₹ 2–5 L</div>
          <div style="font-size:11px;color:{C['slate']};margin-bottom:14px;">Per year · Up to 3 users</div>
          <div style="font-size:11px;color:{C['off_white']};line-height:2.0;">
            ✓ Everything in Free<br>
            ✓ Custom sector calibration<br>
            ✓ Calibration with your data<br>
            ✓ Quarterly parameter updates<br>
            ✓ Priority email support (4 hr)<br>
            ✓ BRSR filing language support
          </div>
        </div>
        """, unsafe_allow_html=True)

    with tier_col3:
        st.markdown(f"""
        <div style="background:{C['card']};border:1px solid {C['coral']};border-radius:10px;
                    padding:20px;height:340px;">
          <div style="font-size:16px;font-weight:700;color:{C['coral']};margin-bottom:4px;">Enterprise</div>
          <div style="font-size:28px;font-weight:800;color:{C['white']};margin-bottom:2px;">₹ 12–18 L</div>
          <div style="font-size:11px;color:{C['slate']};margin-bottom:14px;">Per year · Unlimited users</div>
          <div style="font-size:11px;color:{C['off_white']};line-height:2.0;">
            ✓ Everything in Professional<br>
            ✓ API access (v1.4)<br>
            ✓ White-label / custom branding<br>
            ✓ Private cloud deployment<br>
            ✓ Annual model validation report<br>
            ✓ Monthly strategy session
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

    # ── Contact / Request Form ────────────────────────────────
    st.markdown(f"<h3 style='color:{C['white']};'>Request a Demo or Access</h3>", unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:13px;color:{C['slate']};margin-bottom:16px;'>"
        f"Fill in your details and we will get back to you within 24 hours with a personalised demo slot or access link.</div>",
        unsafe_allow_html=True
    )

    fc1, fc2 = st.columns(2)
    with fc1:
        req_name  = st.text_input("Your Name *", placeholder="Rajesh Kumar", key="req_name")
        req_email = st.text_input("Work Email *", placeholder="rk@yourbank.com", key="req_email")
        req_org   = st.text_input("Organisation *", placeholder="XYZ Bank / ABC Corp", key="req_org")
    with fc2:
        req_role  = st.selectbox("Your Role", [
            "— Select —",
            "Bank CRO / Risk Head",
            "Credit Risk Analyst",
            "CFO / Finance Director",
            "Sustainability / ESG Officer",
            "ESG Consultant",
            "CA / CS / Compliance",
            "Researcher / Academic",
            "Other",
        ], key="req_role")
        req_tier  = st.selectbox("Interested Tier", [
            "Free Pilot (just explore)",
            "Professional (₹ 2–5 L/yr)",
            "Enterprise (₹ 12–18 L/yr)",
            "Not sure — need a demo first",
        ], key="req_tier")
        req_sector = st.selectbox("Primary Sector of Interest", [
            "Steel", "Power", "Cement", "Oil & Gas", "Manufacturing",
            "Multiple sectors", "Other",
        ], key="req_sector")

    req_msg = st.text_area(
        "What do you want to use ICCRE for? (optional)",
        placeholder=(
            "E.g.: We need to run NGFS scenario analysis for our ICAAP submission. "
            "We have 15 large Steel and Power borrowers we want to stress-test..."
        ),
        height=80,
        key="req_msg",
    )

    # Build mailto link
    if st.button("📨 Send Request", type="primary", width="content"):
        if not req_name or not req_email or not req_org or req_role == "— Select —":
            st.error("Please fill in Name, Email, Organisation, and Role before submitting.")
        else:
            subject = f"ICCRE Access Request — {req_role} at {req_org}"
            body = (
                f"Name: {req_name}\n"
                f"Email: {req_email}\n"
                f"Organisation: {req_org}\n"
                f"Role: {req_role}\n"
                f"Tier Interest: {req_tier}\n"
                f"Sector: {req_sector}\n\n"
                f"Message:\n{req_msg or 'No message provided.'}\n\n"
                f"---\nSent from ICCRE v{MODEL_VERSION} Get Access tab"
            )
            mailto = (
                f"mailto:{CONTACT_EMAIL}"
                f"?subject={urllib.parse.quote(subject)}"
                f"&body={urllib.parse.quote(body)}"
            )
            st.markdown(
                f"""
                <div style="background:{C['card']};border:1px solid {C['mint']};border-radius:10px;
                            padding:20px;text-align:center;">
                  <div style="font-size:24px;margin-bottom:8px;">✅</div>
                  <div style="font-size:15px;font-weight:600;color:{C['mint']};margin-bottom:6px;">
                    Ready to send!
                  </div>
                  <div style="font-size:12px;color:{C['slate']};margin-bottom:14px;">
                    Click the button below to open your email client with the request pre-filled.
                    We respond within 24 hours.
                  </div>
                  <a href="{mailto}"
                     style="background:{C['accent']};color:{C['bg_dark']};padding:10px 28px;
                            border-radius:8px;font-weight:700;font-size:13px;text-decoration:none;
                            display:inline-block;">
                    ✉ Open Email to Send Request
                  </a>
                  <div style="font-size:10px;color:{C['bg_ocean']};margin-top:10px;">
                    Or email directly: {CONTACT_EMAIL}
                  </div>
                </div>
                """,
                unsafe_allow_html=True
            )

    # ── Who uses ICCRE ────────────────────────────────────────
    st.divider()
    st.markdown(f"<h3 style='color:{C['white']};'>Built for</h3>", unsafe_allow_html=True)

    who_col1, who_col2, who_col3, who_col4 = st.columns(4)
    for col, emoji, title, desc in [
        (who_col1, "🏦", "Indian Banks", "ICAAP climate overlay · RBI stress test · IFRS 9 ECL · Borrower climate ratings"),
        (who_col2, "🏭", "Listed Corporates", "SEBI BRSR Core filing · Net Zero target modelling · Lender climate questionnaires"),
        (who_col3, "📊", "ESG Consultants", "Multi-client analysis · Regulatory filing support · AI narrative generation"),
        (who_col4, "🎓", "Researchers", "India-specific climate credit models · Free academic access · Open methodology"),
    ]:
        with col:
            st.markdown(f"""
            <div style="background:{C['card']};border:1px solid {C['bg_ocean']};border-radius:8px;
                        padding:14px;text-align:center;height:160px;">
              <div style="font-size:26px;margin-bottom:6px;">{emoji}</div>
              <div style="font-size:12px;font-weight:700;color:{C['white']};margin-bottom:6px;">{title}</div>
              <div style="font-size:10px;color:{C['slate']};line-height:1.6;">{desc}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── Footer strip ──────────────────────────────────────────
    st.markdown("<div style='margin-top:24px;'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:{C['bg_mid']};border-top:1px solid {C['bg_ocean']};
                padding:16px 20px;border-radius:8px;
                display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px;align-items:center;">
      <div>
        <span style="font-size:14px;font-weight:700;color:{C['accent2']};">ICCRE v{MODEL_VERSION}</span>
        <span style="font-size:11px;color:{C['slate']};margin-left:12px;">{PRODUCT_TAGLINE}</span>
      </div>
      <div style="display:flex;gap:16px;align-items:center;">
        <a href="mailto:{CONTACT_EMAIL}" style="font-size:11px;color:{C['slate']};text-decoration:none;">
          ✉ {CONTACT_EMAIL}
        </a>
        <a href="{LINKEDIN_URL}" style="font-size:11px;color:{C['slate']};text-decoration:none;">
          LinkedIn
        </a>
        <span style="font-size:10px;color:{C['bg_ocean']};">
          Build {MODEL_BUILD_DATE} · NGFS Phase III 2023 · ISSB S2 · RBI 2024
        </span>
      </div>
    </div>
    """, unsafe_allow_html=True)
